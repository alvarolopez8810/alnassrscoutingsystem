import streamlit as st
import pandas as pd
import base64
from PIL import Image
import io
import datetime
import json
from scraper_saff import scrape_schedule

# Page configuration
st.set_page_config(
    page_title="Al Nassr Academy - Scouting Department",
    page_icon="alnassracademy.png",
    layout="wide"
)

# -----------------------------
# HELPER FUNCTIONS
# -----------------------------

def create_download_buttons(df, filename_base, label_prefix="Download"):
    """
    Create download buttons for CSV and Excel formats
    
    Args:
        df: pandas DataFrame to download
        filename_base: Base name for the file (without extension)
        label_prefix: Prefix for button labels
    """
    if df.empty:
        st.warning("No data available to download")
        return
    
    col1, col2 = st.columns(2)
    
    with col1:
        # CSV Download
        csv = df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label=f"📥 {label_prefix} CSV",
            data=csv,
            file_name=f"{filename_base}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    with col2:
        # Excel Download
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
        excel_data = output.getvalue()
        
        st.download_button(
            label=f"📥 {label_prefix} EXCEL",
            data=excel_data,
            file_name=f"{filename_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# Custom CSS for the dashboard
def apply_custom_css():
    st.markdown("""
    <style>
        /* Main container */
        .main-container {
            max-width: 1200px;
            margin: 0 auto;
            padding: 1rem;
        }
        
        /* Header styling */
        .header {
            background-color: #002B5B;
            color: #F9D342;
            padding: 2rem;
            border-radius: 10px;
            text-align: center;
            margin-bottom: 2rem;
        }
        
        .logo {
            max-width: 150px;
            margin: 0 auto 1rem;
            display: block;
        }
        
        .title {
            color: #F9D342;
            font-size: 2.5rem;
            font-weight: 800;
            margin: 0.5rem 0;
        }
        
        .subtitle {
            color: white;
            font-size: 1.2rem;
            margin: 0;
        }
        
        /* Dashboard grid */
        .dashboard-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 1.5rem;
            margin-top: 2rem;
        }
        
        /* Dashboard cards */
        .dashboard-card {
            background: white;
            border-radius: 10px;
            padding: 2rem 1.5rem;
            text-align: center;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
            transition: transform 0.3s ease, box-shadow 0.3s ease;
            cursor: pointer;
            border: 1px solid #e0e0e0;
        }
        
        .dashboard-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.15);
        }
        
        .card-icon {
            font-size: 2.5rem;
            margin-bottom: 1rem;
            color: #002B5B;
        }
        
        .card-title {
            font-size: 1.2rem;
            font-weight: 600;
            margin: 0.5rem 0;
            color: #002B5B;
        }
        
        /* Leagues section */
        .leagues-section {
            margin-top: 2rem;
            padding: 1.5rem;
            background: #f8f9fa;
            border-radius: 10px;
        }
        
        /* Responsive design */
        @media (max-width: 900px) {
            .dashboard-grid {
                grid-template-columns: repeat(2, 1fr);
            }
        }
        
        @media (max-width: 600px) {
            .dashboard-grid {
                grid-template-columns: 1fr;
            }
        }
        
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.5rem;
            padding: 0.5rem;
            background: #f8f9fa;
            border-radius: 12px;
            margin-bottom: 1.5rem;
        }
        
        .stTabs [data-baseweb="tab"] {
            padding: 0.75rem 1.5rem;
            border-radius: 8px;
            font-weight: 600;
            color: #002B5B;
            transition: all 0.2s ease;
        }
        
        .stTabs [aria-selected="true"] {
            background: #002B5B;
            color: #F9D342 !important;
        }
    </style>
    """, unsafe_allow_html=True)

# Initialize session state
if 'language' not in st.session_state:
    st.session_state.language = 'en'
if 'page' not in st.session_state:
    st.session_state.page = 'home'
if 'category' not in st.session_state:
    st.session_state.category = None

# -----------------------------
# CATEGORIES
# -----------------------------
CATEGORIES = {
    'NATIONAL TEAMS': 'NATIONAL TEAMS',
    'PRO LEAGUE': 'PRO LEAGUE',
    '1 DIVISION': '1 DIVISION',
    'U21': 'U21',
    'U18': 'U18',
    'U17': 'U17',
    'U16': 'U16',
    'U15': 'U15'
}

# Mapping categories to Excel files
CATEGORY_FILES = {
    'U21': 'SaudiLeagueU21.xlsx',
    'U18': 'SaudiLeagueU18.xlsx',
    'U17': 'SaudiLeagueU17.xlsx',
    'U16': 'SaudiLeagueU16.xlsx',
    'U15': 'SaudiLeagueU15.xlsx'
}

POSITIONS = [
    "GK", "RB", "RWB", "CB", "LB", "LWB",
    "DM/6", "CM/8", "AM/10",
    "RW/WF", "LW/WF", "ST/9", "SS/9.5"
]

FOOT = ["Right", "Left", "Both"]
CONCLUSION = ["SIGN (التوقيع معه)", "MONITOR CLOSELY (متابعة دقيقة)", "DISCARD (الاستبعاد)"]

# Load data functions
@st.cache_data
def load_teams_data():
    try:
        return pd.read_excel("Saudi_Youth_Leagues.xlsx")
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return pd.DataFrame()

@st.cache_data
def load_countries():
    try:
        df = pd.read_excel("countries.xlsx")
        # Assuming the countries are in the first column
        countries = df.iloc[:, 0].dropna().tolist()
        return sorted(countries)
    except Exception as e:
        st.error(f"Error loading countries: {e}")
        return ["Saudi Arabia"]  # Default fallback

@st.cache_data
def load_league_data(category):
    """Load league and team data for a specific category (combines Premier and Division 1)"""
    try:
        dfs = []
        
        if category in CATEGORY_FILES:
            # Load Premier League file
            df_premier = pd.read_excel(CATEGORY_FILES[category])
            # Standardize column name to 'Team'
            if 'Equipos' in df_premier.columns:
                df_premier = df_premier.rename(columns={'Equipos': 'Team'})
            dfs.append(df_premier)
            
            # Try to load Division 1 file if it exists
            div1_file = CATEGORY_FILES[category].replace('.xlsx', 'Div1.xlsx')
            try:
                df_div1 = pd.read_excel(div1_file)
                # Standardize column name to 'Team'
                if 'Equipos' in df_div1.columns:
                    df_div1 = df_div1.rename(columns={'Equipos': 'Team'})
                dfs.append(df_div1)
            except FileNotFoundError:
                pass  # Division 1 file doesn't exist, that's ok
            
            # Combine all dataframes
            if dfs:
                combined_df = pd.concat(dfs, ignore_index=True)
                return combined_df
            else:
                return pd.DataFrame()
        else:
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading league data for {category}: {e}")
        return pd.DataFrame()

def add_player_to_squad(report_data, category):
    """Add player to Squad database when a report is created"""
    try:
        if category == 'U21':
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Load existing squad data
            df_squad = pd.read_excel('SaudiLeagueU21.xlsx', sheet_name='SquadU21')
            
            # Rename old columns if needed
            if 'evaluation' in df_squad.columns:
                df_squad = df_squad.rename(columns={'evaluation': 'Conclusion'})
            if 'notes' in df_squad.columns:
                df_squad = df_squad.rename(columns={'notes': 'Smart Evaluation'})
            
            # Check if player already exists (by name and team)
            player_exists = (
                (df_squad['Name'] == report_data['Player Name']) & 
                (df_squad['Team'] == report_data['Team'])
            ).any()
            
            if not player_exists:
                # Create new player entry
                new_player = {
                    'Number': report_data.get('Number', 0),
                    'Name': report_data['Player Name'],
                    'Position': report_data['Position'],
                    'Year of Birth': report_data['Birth Year'],
                    'Team': report_data['Team'],
                    'Conclusion': report_data['Conclusion'],
                    'Smart Evaluation': report_data['Smart Evaluation'],
                    'KSA Caps': 0
                }
                
                # Add to dataframe
                df_squad = pd.concat([df_squad, pd.DataFrame([new_player])], ignore_index=True)
                
                # Save to Excel
                book = openpyxl.load_workbook('SaudiLeagueU21.xlsx')
                if 'SquadU21' in book.sheetnames:
                    del book['SquadU21']
                ws = book.create_sheet('SquadU21')
                for r in dataframe_to_rows(df_squad, index=False, header=True):
                    ws.append(r)
                book.save('SaudiLeagueU21.xlsx')
                book.close()
                
                st.cache_data.clear()  # Clear cache to reload data
                return True
            else:
                return False  # Player already exists
        return False
    except Exception as e:
        st.warning(f"Note: Could not add player to Squad database: {e}")
        return False

def save_report_to_excel(report_data):
    """Save scout report to Excel file"""
    try:
        # Try to load existing reports
        try:
            df = pd.read_excel("scouts_reports.xlsx")
        except FileNotFoundError:
            # Create new DataFrame if file doesn't exist
            df = pd.DataFrame()
        
        # Append new report
        new_report = pd.DataFrame([report_data])
        df = pd.concat([df, new_report], ignore_index=True)
        
        # Save to Excel
        df.to_excel("scouts_reports.xlsx", index=False)
        
        # Also add player to Squad database if it's U21
        added_to_squad = False
        if 'Category' in report_data and report_data['Category'] == 'U21':
            added_to_squad = add_player_to_squad(report_data, 'U21')
            if not added_to_squad:
                st.info(f"ℹ️ Player '{report_data['Player Name']}' already exists in Squad Database or could not be added.")
        
        return {'success': True, 'added_to_squad': added_to_squad}
    except Exception as e:
        st.error(f"Error saving report: {e}")
        return False

@st.cache_data
def load_reports():
    """Load all scout reports from Excel"""
    try:
        df = pd.read_excel("scouts_reports.xlsx")
        return df
    except FileNotFoundError:
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading reports: {e}")
        return pd.DataFrame()

def show_leagues_section():
    st.markdown("## Leagues & Teams")
    df = load_teams_data()
    
    if not df.empty:
        # League filter
        league_filter = st.selectbox("Select League:", df['League'].unique())
        filtered_teams = df[df['League'] == league_filter]
        
        # Display teams in a table
        st.dataframe(filtered_teams, use_container_width=True)
    else:
        st.warning("No data available. Please check the Excel file.")

# -----------------------------
# LANGUAGE FUNCTIONS
# -----------------------------
def toggle_language():
    if st.session_state.language == 'en':
        st.session_state.language = 'ar'
    else:
        st.session_state.language = 'en'

def get_league_name(league_key):
    league_names = {
        'en': {
            "Saudi U-18 Premier League": "Saudi U-18 Premier League",
            "Saudi U-17 Premier League": "Saudi U-17 Premier League", 
            "Saudi U-16 Premier League": "Saudi U-16 Premier League",
            "Saudi U-15 Premier League": "Saudi U-15 Premier League",
            "Saudi Elite League U-21": "Saudi Elite League U-21"
        },
        'ar': {
            "Saudi U-18 Premier League": "الدوري السعودي الممتاز تحت 18",
            "Saudi U-17 Premier League": "الدوري السعودي الممتاز تحت 17",
            "Saudi U-16 Premier League": "الدوري السعودي الممتاز تحت 16", 
            "Saudi U-15 Premier League": "الدوري السعودي الممتاز تحت 15",
            "Saudi Elite League U-21": "الدوري السعودي النخبة تحت 21"
        }
    }
    return league_names[st.session_state.language].get(league_key, league_key)

# -----------------------------
# HOME PAGE - MAIN MENU
# -----------------------------
def show_home_page():
    # Header section
    st.markdown(
        f"""
        <div class="header">
            <img src="data:image/png;base64,{base64.b64encode(open('alnassracademy.png', 'rb').read()).decode()}" class="logo">
            <h1 class="title">AL NASSR SCOUTING DEPARTMENT</h1>
            <p class="subtitle">{'PROFESSIONAL SCOUTING SYSTEM' if st.session_state.language == 'en' else 'نظام الكشافة الاحترافي'}</p>
        </div>
        """, 
        unsafe_allow_html=True
    )
    
    # Main menu buttons
    st.markdown(f"<h2 style='text-align: center; color: #002B5B; margin: 2rem 0;'>{'Select an Option' if st.session_state.language == 'en' else 'اختر خياراً'}</h2>", unsafe_allow_html=True)
    
    # Custom CSS for home buttons
    st.markdown("""
    <style>
        .home-button-container {
            text-align: center;
            margin: 1rem 0;
        }
        .home-button-image {
            height: 120px;
            margin-bottom: 1rem;
        }
        .home-button-text {
            font-size: 1.5rem;
            font-weight: bold;
            color: #002B5B;
            margin-top: 0.5rem;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Three buttons layout
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Saudi Football button with SAFF logo
        try:
            import io
            from PIL import Image
            saff_logo = Image.open('saff.png')
            buffered = io.BytesIO()
            saff_logo.save(buffered, format="PNG")
            saff_img_str = base64.b64encode(buffered.getvalue()).decode()
            
            st.markdown(
                f'<div class="home-button-container"><img src="data:image/png;base64,{saff_img_str}" class="home-button-image"></div>',
                unsafe_allow_html=True
            )
        except:
            st.markdown('<div class="home-button-container">🏆</div>', unsafe_allow_html=True)
        
        if st.button("SAUDI FOOTBALL\n\nالكرة السعودية", key="btn_category", use_container_width=True):
            st.session_state.page = 'categories'
            st.rerun()
    
    with col2:
        # FIFA U20 World Cup button
        try:
            fifa_logo = Image.open('fifau20.png')
            buffered = io.BytesIO()
            fifa_logo.save(buffered, format="PNG")
            fifa_img_str = base64.b64encode(buffered.getvalue()).decode()
            
            st.markdown(
                f'<div class="home-button-container"><img src="data:image/png;base64,{fifa_img_str}" class="home-button-image"></div>',
                unsafe_allow_html=True
            )
        except:
            st.markdown('<div class="home-button-container">⚽</div>', unsafe_allow_html=True)
        
        if st.button("FIFA WORLD CUP U20\n\nكأس العالم تحت 20", key="btn_fifa_u20", use_container_width=True):
            st.session_state.page = 'fifa_u20'
            st.rerun()
    
    with col3:
        # Advanced Data Analysis - SAFF+ button
        try:
            saffplus_logo = Image.open('saffplus.jpg')
            buffered = io.BytesIO()
            saffplus_logo.save(buffered, format="JPEG")
            saffplus_img_str = base64.b64encode(buffered.getvalue()).decode()
            
            st.markdown(
                f'<div class="home-button-container"><img src="data:image/jpeg;base64,{saffplus_img_str}" class="home-button-image"></div>',
                unsafe_allow_html=True
            )
        except:
            st.markdown('<div class="home-button-container"><div style="font-size: 80px;">🔬</div></div>', unsafe_allow_html=True)
        
        if st.button("ADVANCED DATA\nANALYSIS - SAFF+\n\nتحليل البيانات - SAFF+", key="btn_advanced_data", use_container_width=True):
            st.session_state.page = 'advanced_data'
            st.rerun()

# -----------------------------
# CATEGORIES PAGE
# -----------------------------
def show_categories_page():
    # Header
    st.markdown(f"<h2 style='text-align: center; color: #002B5B;'>{'Select Category' if st.session_state.language == 'en' else 'اختر الفئة'}</h2>", unsafe_allow_html=True)
    st.markdown("<div style='margin: 1rem 0;'></div>", unsafe_allow_html=True)
    
    # Custom CSS for uniform image sizes
    st.markdown("""
    <style>
        .category-image-container {
            height: 200px;
            display: flex;
            align-items: center;
            justify-content: center;
            overflow: hidden;
            margin-bottom: 1rem;
            border: 2px solid #002B5B;
            border-radius: 10px;
            background-color: white;
        }
        .category-image-container img {
            max-height: 180px;
            max-width: 100%;
            object-fit: contain;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Category logo mapping - using exact file names
    category_logos = {
        'NATIONAL TEAMS': 'NationalTeam.png',
        'PRO LEAGUE': 'Senior1DIV.png',
        '1 DIVISION': 'Senior2DIV.png',
        'U21': 'U21logo.png',
        'U18': 'U18logo.png',
        'U17': 'U17logo.png',
        'U16': 'U16logo.png',
        'U15': 'U15logo.png'
    }
    
    # Create cards for each category
    col1, col2, col3, col4 = st.columns(4)
    
    categories_list = list(CATEGORIES.keys())
    
    for idx, category in enumerate(categories_list):
        col_idx = idx % 4
        with [col1, col2, col3, col4][col_idx]:
            # Try to load category logo
            logo_file = category_logos.get(category, None)
            
            if logo_file:
                try:
                    # Load and display image with uniform size
                    image = Image.open(logo_file)
                    # Convert image to base64 for HTML display
                    import io
                    import base64
                    buffered = io.BytesIO()
                    image.save(buffered, format="PNG")
                    img_str = base64.b64encode(buffered.getvalue()).decode()
                    
                    # Display image in a uniform container
                    st.markdown(
                        f'<div class="category-image-container"><img src="data:image/png;base64,{img_str}"></div>',
                        unsafe_allow_html=True
                    )
                    if st.button(f"{category}", key=f"cat_{category}", use_container_width=True):
                        st.session_state.category = category
                        st.session_state.page = 'category_view'
                        st.rerun()
                except FileNotFoundError:
                    # If logo doesn't exist, use emoji button
                    if st.button(f"⚽\n\n{category}", key=f"cat_{category}", use_container_width=True):
                        st.session_state.category = category
                        st.session_state.page = 'category_view'
                        st.rerun()
                except Exception as e:
                    # If there's any other error, show error and fallback to emoji
                    st.warning(f"Error loading {logo_file}: {str(e)}")
                    if st.button(f"⚽\n\n{category}", key=f"cat_{category}", use_container_width=True):
                        st.session_state.category = category
                        st.session_state.page = 'category_view'
                        st.rerun()
            else:
                # Fallback to emoji button
                if st.button(f"⚽\n\n{category}", key=f"cat_{category}", use_container_width=True):
                    st.session_state.category = category
                    st.session_state.page = 'category_view'
                    st.rerun()
    
    # Add spacing for remaining categories
    if len(categories_list) < 8:
        st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

# -----------------------------
# DATABASE VIEW FUNCTIONS
# -----------------------------
@st.cache_data
def load_squad_data(category):
    """Load squad data for U21 category from SquadU21 sheet"""
    try:
        if category == 'U21':
            df = pd.read_excel('SaudiLeagueU21.xlsx', sheet_name='SquadU21')
            
            # Rename old columns to new ones if they exist
            if 'evaluation' in df.columns and 'Conclusion' not in df.columns:
                df = df.rename(columns={'evaluation': 'Conclusion'})
            if 'notes' in df.columns and 'Smart Evaluation' not in df.columns:
                df = df.rename(columns={'notes': 'Smart Evaluation'})
            
            # Add KSA Caps column if it doesn't exist
            if 'KSA Caps' not in df.columns:
                df['KSA Caps'] = 0
            
            # Add League column if it doesn't exist and map teams to leagues
            if 'League' not in df.columns:
                # Default league mapping for Saudi teams
                elite_teams = [
                    'Al-Nassr FC', 'Al-Hilal SFC', 'Al-Ahli Saudi FC', 'Al-Ittihad Club',
                    'Al-Shabab FC', 'Al-Taawoun FC', 'Al-Fateh SC', 'Al-Fayha FC',
                    'Al-Raed FC', 'Al-Wehda FC', 'Damac FC', 'Ettifaq FC',
                    'Al-Khaleej FC', 'Al-Hazem FC', 'Al-Riyadh SC', 'Al-Okhdood FC',
                    'Al-Qadsiah FC', 'Al-Orobah FC'
                ]
                
                # Map teams to leagues
                df['League'] = df['Team'].apply(
                    lambda x: 'Saudi Elite League U-21' if x in elite_teams else 'Saudi First Division U-21'
                )
            
            # Ensure all required columns exist
            required_cols = ['Number', 'Name', 'Position', 'Year of Birth', 'Team', 'League', 'Conclusion', 'Smart Evaluation', 'KSA Caps']
            for col in required_cols:
                if col not in df.columns:
                    df[col] = '' if col in ['Position', 'Team', 'League', 'Conclusion', 'Smart Evaluation'] else 0
            return df
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading squad data: {e}")
        return pd.DataFrame()

def save_squad_data(df, category):
    """Save squad data back to Excel"""
    try:
        if category == 'U21':
            # Load existing workbook
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Load the existing workbook
            book = openpyxl.load_workbook('SaudiLeagueU21.xlsx')
            
            # Remove the SquadU21 sheet if it exists
            if 'SquadU21' in book.sheetnames:
                del book['SquadU21']
            
            # Create a new sheet
            ws = book.create_sheet('SquadU21')
            
            # Write the dataframe to the sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Save the workbook
            book.save('SaudiLeagueU21.xlsx')
            book.close()
            
            st.success("✅ Data saved successfully!")
            st.cache_data.clear()
            return True
    except Exception as e:
        st.error(f"Error saving data: {e}")
        return False
    return False

def show_create_report_form(category):
    """Show form to create match reports for Saudi Football categories - Smart version"""
    
    # Al Nassr FC Branded CSS
    st.markdown("""
    <style>
        /* Al Nassr Brand Colors */
        :root {
            --navy: #1a2332;
            --gold: #FFC60A;
            --gold-light: #FFD700;
            --success: #4CAF50;
            --bg-light: #f8f9fa;
            --border: #e0e0e0;
        }
        
        /* Sliders - Gold instead of Red */
        .stSlider > div > div > div > div {
            background: linear-gradient(90deg, var(--gold) 0%, var(--gold-light) 100%) !important;
        }
        .stSlider > div > div > div > div > div {
            background-color: var(--gold) !important;
            border: 3px solid var(--navy) !important;
            box-shadow: 0 0 10px rgba(255, 198, 10, 0.5) !important;
        }
        .stSlider > div > div > div {
            background-color: var(--border) !important;
        }
        
        /* Input fields - Gold focus */
        .stTextInput > div > div > input:focus,
        .stSelectbox > div > div > div:focus,
        .stTextArea > div > div > textarea:focus,
        .stNumberInput > div > div > input:focus {
            border-color: var(--gold) !important;
            box-shadow: 0 0 0 0.2rem rgba(255, 198, 10, 0.25) !important;
        }
        
        /* Textarea special styling */
        .stTextArea > div > div > textarea {
            border-left: 4px solid var(--gold) !important;
            background-color: var(--bg-light) !important;
        }
        
        /* Buttons - Gold primary */
        .stButton > button[kind="primary"] {
            background: linear-gradient(135deg, var(--gold) 0%, var(--gold-light) 100%) !important;
            color: var(--navy) !important;
            font-weight: 700 !important;
            font-size: 18px !important;
            height: 60px !important;
            border: none !important;
            border-radius: 12px !important;
            box-shadow: 0 4px 15px rgba(255, 198, 10, 0.4) !important;
            text-transform: uppercase !important;
            letter-spacing: 1px !important;
        }
        .stButton > button[kind="primary"]:hover {
            transform: translateY(-3px) !important;
            box-shadow: 0 6px 20px rgba(255, 198, 10, 0.6) !important;
        }
        
        /* Section headers */
        .section-header {
            background: var(--navy);
            color: white;
            padding: 15px 20px;
            border-left: 6px solid var(--gold);
            border-radius: 8px 8px 0 0;
            margin: 20px 0 10px 0;
            font-size: 18px;
            font-weight: 700;
        }
        
        /* Success message styling */
        .success-message {
            background: linear-gradient(135deg, rgba(76, 175, 80, 0.1) 0%, rgba(76, 175, 80, 0.05) 100%);
            border-left: 4px solid var(--success);
            padding: 12px 16px;
            border-radius: 8px;
            color: var(--success);
            font-weight: 600;
            margin: 10px 0;
        }
        
        /* Form container */
        .form-container {
            background: white;
            padding: 25px;
            border-radius: 12px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.08);
            margin-bottom: 20px;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Header with Logo
    col_logo, col_title, col_refresh = st.columns([1, 6, 2])
    
    with col_logo:
        try:
            st.image('alnassr.png', width=60)
        except:
            st.markdown("<div style='font-size: 60px;'>⚽</div>", unsafe_allow_html=True)
    
    with col_title:
        st.markdown("<h2 style='color: #1a2332; margin-top: 10px;'>Create Match Report</h2>", unsafe_allow_html=True)
    
    with col_refresh:
        if st.button("🔄 Clear Cache", help="Reload database"):
            st.cache_data.clear()
            st.rerun()
    
    # Load squad database
    df_squad = load_squad_data(category)
    
    if df_squad.empty:
        st.error("⚠️ No squad data available. Please check the database.")
        return
    
    # Initialize session state for player creation
    if 'show_create_player' not in st.session_state:
        st.session_state.show_create_player = False
    
    # Match Information Section
    st.markdown("<div class='section-header'>🏟️ Match Information</div>", unsafe_allow_html=True)
    st.markdown("<div class='form-container'>", unsafe_allow_html=True)
    col_m1, col_m2 = st.columns(2)
    
    with col_m1:
        match_date = st.date_input("📅 Match Date")
    
    with col_m2:
        scout_name = st.text_input("👤 Scout Name", placeholder="Your name")
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Player Selection Section
    st.markdown("<div class='section-header'>👕 Player Selection (Smart Database)</div>", unsafe_allow_html=True)
    st.markdown("<div class='form-container'>", unsafe_allow_html=True)
    
    # Step 1: Select League
    leagues = sorted(df_squad['League'].dropna().unique().tolist())
    selected_league = st.selectbox(
        "🏆 Select League",
        leagues,
        help="First, select the league"
    )
    
    # Get teams for match and player selection
    df_league = df_squad[df_squad['League'] == selected_league]
    teams_in_league = sorted(df_league['Team'].dropna().unique().tolist())
    
    # Match Name: Team vs Team
    st.markdown("**⚽ Match (Team vs Team)**")
    col_t1, col_vs, col_t2 = st.columns([5, 1, 5])
    
    with col_t1:
        team1 = st.selectbox("Home Team", teams_in_league, key="home_team")
    
    with col_vs:
        st.markdown("<div style='text-align:center;margin-top:28px;font-size:18px;font-weight:bold;color:#1a2332;'>VS</div>", unsafe_allow_html=True)
    
    with col_t2:
        team2 = st.selectbox("Away Team", teams_in_league, key="away_team")
    
    match_name = f"{team1} vs {team2}"
    
    st.markdown("---")
    
    # Step 2: Select player's team
    selected_team = st.selectbox(
        "🛡️ Select Player's Team",
        teams_in_league,
        help="Select the team of the player you're reporting on"
    )
    
    # Step 3: Filter players by team and show numbers
    df_team = df_league[df_league['Team'] == selected_team]
    
    # Get available numbers
    available_numbers = sorted([int(n) for n in df_team['Number'].dropna().unique() if str(n).isdigit()])
    
    col_num1, col_num2 = st.columns([3, 1])
    
    with col_num1:
        if available_numbers:
            selected_number = st.selectbox(
                "🔢 Select Player Number",
                ['Select...'] + available_numbers,
                help="Choose the player's number"
            )
        else:
            selected_number = 'Select...'
            st.warning("No players found in this team")
    
    with col_num2:
        st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
        if st.button("➕ Create New Player", help="Add a player not in database"):
            st.session_state.show_create_player = True
    
    # Auto-fill player data or allow manual entry
    player_name = ""
    player_position = ""
    player_birth_year = ""
    auto_filled = False
    
    if selected_number != 'Select...' and selected_number:
        # Find player data
        player_data = df_team[df_team['Number'] == selected_number]
        
        if not player_data.empty:
            player_name = player_data.iloc[0]['Name']
            player_position = player_data.iloc[0]['Position']
            player_birth_year = player_data.iloc[0]['Year of Birth']
            auto_filled = True
            
            st.markdown(f"<div class='success-message'>✅ <strong>Player Auto-Filled:</strong> {player_name} - {player_position} ({player_birth_year})</div>", unsafe_allow_html=True)
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Player details (editable)
    st.markdown("<div class='section-header'>ℹ️ Player Details (Editable)</div>", unsafe_allow_html=True)
    st.markdown("<div class='form-container'>", unsafe_allow_html=True)
    col_p1, col_p2, col_p3, col_p4 = st.columns(4)
    
    with col_p1:
        player_name_input = st.text_input(
            "👤 Player Name",
            value=str(player_name) if player_name else "",
            placeholder="Player full name"
        )
    
    with col_p2:
        player_position_input = st.text_input(
            "📍 Position",
            value=str(player_position) if player_position else "",
            placeholder="e.g., CM, ST, CB"
        )
    
    with col_p3:
        player_birth_year_input = st.text_input(
            "🎂 Year of Birth",
            value=str(player_birth_year) if player_birth_year and str(player_birth_year) != 'nan' else "",
            placeholder="e.g., 2004"
        )
    
    with col_p4:
        player_nationality = st.selectbox(
            "🌍 Nationality",
            ['Saudi Arabia', 'UAE', 'Qatar', 'Bahrain', 'Kuwait', 'Oman', 'Jordan', 'Iraq', 'Other'],
            index=0
        )
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Show CREATE PLAYER form if button clicked
    if st.session_state.show_create_player:
        st.markdown("---")
        st.markdown("### ➕ Create New Player")
        st.info("This player will be added to the database and VIEW DATABASE page")
        
        with st.form("create_new_player_form"):
            col_np1, col_np2 = st.columns(2)
            
            with col_np1:
                new_player_number = st.number_input("Number", min_value=1, max_value=99, value=1)
                new_player_name = st.text_input("Name", placeholder="Player full name")
                new_player_position = st.text_input("Position", placeholder="e.g., CM")
            
            with col_np2:
                new_player_birth_year = st.number_input("Year of Birth", min_value=1990, max_value=2015, value=2004)
                new_player_team = selected_team
                new_player_league = selected_league
                st.text_input("Team (auto)", value=new_player_team, disabled=True)
            
            col_submit, col_cancel = st.columns(2)
            
            with col_submit:
                submit_new_player = st.form_submit_button("💾 Add to Database", type="primary")
            
            with col_cancel:
                cancel_new_player = st.form_submit_button("❌ Cancel")
            
            if submit_new_player:
                if new_player_name and new_player_position:
                    try:
                        # Add to dataframe
                        new_row = {
                            'Number': new_player_number,
                            'Name': new_player_name,
                            'Position': new_player_position,
                            'Year of Birth': new_player_birth_year,
                            'Team': new_player_team,
                            'League': new_player_league,
                            'Conclusion': '',
                            'Smart Evaluation': '',
                            'KSA Caps': 0
                        }
                        
                        df_squad = pd.concat([df_squad, pd.DataFrame([new_row])], ignore_index=True)
                        
                        # Save to Excel
                        save_squad_data(df_squad, category)
                        
                        st.success(f"✅ Player {new_player_name} added successfully!")
                        st.session_state.show_create_player = False
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error adding player: {e}")
                else:
                    st.error("⚠️ Please fill in Name and Position")
            
            if cancel_new_player:
                st.session_state.show_create_player = False
                st.rerun()
    
    # Evaluation Section
    st.markdown("<div class='section-header'>📊 Performance Evaluation</div>", unsafe_allow_html=True)
    st.markdown("<div class='form-container'>", unsafe_allow_html=True)
    
    col_e1, col_e2, col_e3 = st.columns(3)
    
    with col_e1:
        performance = st.slider("⚡ Performance Rating", 1.0, 6.0, 3.0, 0.5)
        st.markdown(f"<div style='text-align: center; color: #FFC60A; font-size: 42px; font-weight: 700; margin-top: -10px;'>{performance}/6</div>", unsafe_allow_html=True)
    
    with col_e2:
        potential = st.slider("🌟 Potential Rating", 1.0, 6.0, 3.0, 0.5)
        st.markdown(f"<div style='text-align: center; color: #FFC60A; font-size: 42px; font-weight: 700; margin-top: -10px;'>{potential}/6</div>", unsafe_allow_html=True)
    
    with col_e3:
        conclusion = st.selectbox(
            "✅ Conclusion",
            ["A - Firmar (Sign)", 
             "B+ - Seguir para Firmar (Follow to Sign)", 
             "B - Seguir (Follow)"]
        )
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Report Section
    st.markdown("<div class='section-header'>📝 Detailed Report</div>", unsafe_allow_html=True)
    st.markdown("<div class='form-container'>", unsafe_allow_html=True)
    
    report_text = st.text_area(
        "📝 Write your detailed observations",
        placeholder="Write detailed observations about the player's performance, strengths, weaknesses, tactical awareness, technical abilities, physical attributes, mental qualities, etc...",
        height=200
    )
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Save Report Button
    if st.button("💾 Save Report", type="primary", use_container_width=True):
        # Validate required fields
        if not all([match_name, player_name_input, scout_name, selected_number != 'Select...']):
            st.error("⚠️ Please fill in all required fields: Match Name, Player Number, Player Name, and Scout Name")
        else:
            try:
                # Create report data with specific columns for saffmatchreports.xlsx
                report_data = {
                    'Scout Name': scout_name,
                    'Player Name': player_name_input,
                    'Position': player_position_input,
                    'Year of Birth': player_birth_year_input,
                    'Nationality': player_nationality,
                    'League': selected_league,
                    'Team': selected_team,
                    'Potential Rating': potential,
                    'Performance Rating': performance,
                    'Match Name': match_name,
                    'Date of the Report': match_date.strftime('%Y-%m-%d'),
                    'Conclusion': conclusion,
                    'Report Details': report_text
                }
                
                # Load existing reports from saffmatchreports.xlsx
                filename = 'saffmatchreports.xlsx'
                try:
                    df_existing = pd.read_excel(filename)
                except FileNotFoundError:
                    df_existing = pd.DataFrame()
                
                # Append new report
                df_new = pd.DataFrame([report_data])
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                
                # Save to saffmatchreports.xlsx
                df_combined.to_excel(filename, index=False)
                
                st.success(f"✅ Report saved successfully to {filename}!")
                st.balloons()
                
                # Reset create player state
                st.session_state.show_create_player = False
                
            except Exception as e:
                st.error(f"❌ Error saving report: {e}")

def save_player_photo(uploaded_file, player_name):
    """Save player photo to player_photos directory"""
    import os
    import re
    
    # Create player_photos directory if it doesn't exist
    if not os.path.exists('player_photos'):
        os.makedirs('player_photos')
    
    # Sanitize player name for filename
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', player_name)
    file_extension = uploaded_file.name.split('.')[-1]
    photo_filename = f"{safe_name}.{file_extension}"
    photo_path = os.path.join('player_photos', photo_filename)
    
    # Save the file
    with open(photo_path, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    
    return photo_filename

def get_player_photo(player_name):
    """Get player photo filename if it exists"""
    import os
    import re
    
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', player_name)
    player_photos_dir = 'player_photos'
    
    if not os.path.exists(player_photos_dir):
        return None
    
    # Check for any file with this player name
    for file in os.listdir(player_photos_dir):
        if file.startswith(safe_name):
            return file
    
    return None

def show_category_reports(category):
    """Show all reports for a specific category from saffmatchreports.xlsx - Enhanced version"""
    
    # Team logos mapping
    TEAM_LOGOS = {
        'Al-Wehda FC': 'alwehda.png',
        'Al-Jabalain': 'aljabalain.png',
        'Al-Raed FC': 'alraed.png',
        'Al-Orobah FC': 'AlOrobah.png',
        'Al-Adalah FC': 'aladalahclub.png',
        'Al-Bukayriyah FC': 'albukiryahfc.png',
        'Al-Ahli Saudi FC': 'alahli.png',
        'Al-Nassr FC': 'alnassr.png',
        'Al-Taawoun FC': 'altaawoun.png',
        'Al-Shabab FC': 'alshabab.png',
        'Al-Okhdood FC': 'alokhdood.png',
        'Al-Riyadh SC': 'alriyadh.png',
        'Al-Najmah SC': 'alnajma.png',
        'NEOM Club': 'neom.png',
        'Al-Kholood Club': 'alkholood.png',
        'Al-Fateh SC': 'alfateh.png',
        'Damac FC': 'damac.png',
        'Al-Hazem FC': 'alhazem.png',
        'Al-Fayha FC': 'alfayha.png',
        'Ettifaq FC': 'alettifaq.png',
        'Al-Hilal SFC': 'alhilal.png',
        'Al-Qadsiah FC': 'alqadsiah.png',
        'Khaleej FC': 'alkhaleej.png',
        'Al-Ittihad Club': 'alittihad.png'
    }
    
    # Enhanced CSS for View Reports
    st.markdown("""
    <style>
        .filter-container {
            background: white;
            padding: 20px;
            border-radius: 12px;
            border: 2px solid #FFC60A;
            margin-bottom: 25px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }
        .scout-header {
            background: linear-gradient(135deg, #1a2332 0%, #2a3342 100%);
            color: white;
            padding: 18px 25px;
            border-left: 8px solid #FFC60A;
            border-radius: 10px;
            margin: 30px 0 15px 0;
            font-size: 20px;
            font-weight: 700;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }
        .report-card {
            background: white;
            border: 2px solid #e0e0e0;
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 15px;
            transition: all 0.3s ease;
        }
        .report-card:hover {
            border-color: #FFC60A;
            box-shadow: 0 4px 15px rgba(255, 198, 10, 0.3);
            transform: translateY(-2px);
        }
        .player-photo {
            width: 80px;
            height: 80px;
            border-radius: 50%;
            border: 3px solid #FFC60A;
            object-fit: cover;
            box-shadow: 0 2px 8px rgba(0,0,0,0.2);
        }
        .team-badge {
            display: inline-flex;
            align-items: center;
            background: #1a2332;
            color: white;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 600;
            margin-right: 8px;
        }
        .position-badge {
            display: inline-block;
            background: #FFC60A;
            color: #1a2332;
            padding: 4px 10px;
            border-radius: 15px;
            font-size: 12px;
            font-weight: 700;
            margin-right: 8px;
        }
        .stat-box {
            background: #f8f9fa;
            border-left: 4px solid #FFC60A;
            padding: 12px;
            border-radius: 6px;
            margin: 8px 0;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("### 📊 View Reports")
    
    filename = 'saffmatchreports.xlsx'
    
    try:
        df_reports = pd.read_excel(filename)
        
        if df_reports.empty:
            st.info("📋 No reports yet. Create your first report in the CREATE REPORT tab!")
            return
        
        # Enhanced Filters Section
        st.markdown("<div class='filter-container'>", unsafe_allow_html=True)
        st.markdown("#### 🔍 Filters")
        
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        
        with col_f1:
            scouts = ['All Scouts'] + sorted(df_reports['Scout Name'].dropna().unique().tolist())
            filter_scout = st.selectbox("👤 Scout", scouts)
        
        with col_f2:
            leagues = ['All Leagues'] + sorted(df_reports['League'].dropna().unique().tolist())
            filter_league = st.selectbox("🏆 League", leagues)
        
        with col_f3:
            teams = ['All Teams'] + sorted(df_reports['Team'].dropna().unique().tolist())
            filter_team = st.selectbox("🛡️ Team", teams)
        
        with col_f4:
            conclusions = ['All'] + ["A - Firmar (Sign)", "B+ - Seguir para Firmar (Follow to Sign)", "B - Seguir (Follow)"]
            filter_conclusion = st.selectbox("✅ Conclusion", conclusions)
        
        col_apply, col_clear = st.columns([1, 1])
        with col_clear:
            if st.button("❌ Clear All Filters", use_container_width=True):
                st.rerun()
        
        st.markdown("</div>", unsafe_allow_html=True)
        
        # Apply filters
        filtered_df = df_reports.copy()
        
        if filter_scout != 'All Scouts':
            filtered_df = filtered_df[filtered_df['Scout Name'] == filter_scout]
        
        if filter_league != 'All Leagues':
            filtered_df = filtered_df[filtered_df['League'] == filter_league]
        
        if filter_team != 'All Teams':
            filtered_df = filtered_df[filtered_df['Team'] == filter_team]
        
        if filter_conclusion != 'All':
            filtered_df = filtered_df[filtered_df['Conclusion'] == filter_conclusion]
        
        st.markdown(f"<h4 style='color: #1a2332;'>📄 Showing {len(filtered_df)} reports</h4>", unsafe_allow_html=True)
        
        # Download buttons
        if not filtered_df.empty:
            st.markdown("---")
            create_download_buttons(
                filtered_df,
                filename_base=f"{category}_reports",
                label_prefix="Download"
            )
            st.markdown("---")
        
        # Display reports organized by Scout
        if filtered_df.empty:
            st.warning("🗒️ No reports found with the selected filters.")
        else:
            # Group by Scout
            scouts_in_reports = filtered_df['Scout Name'].unique()
            
            for scout in scouts_in_reports:
                scout_reports = filtered_df[filtered_df['Scout Name'] == scout]
                
                # Scout Header
                st.markdown(f"<div class='scout-header'>👤 {scout} <span style='color: #FFC60A;'>({len(scout_reports)} reports)</span></div>", unsafe_allow_html=True)
                
                # Display each report
                for idx, report in scout_reports.iterrows():
                    col_photo, col_upload, col_info = st.columns([1, 1, 4])
                    
                    # Check if player has a saved photo
                    player_photo_file = get_player_photo(report['Player Name'])
                    
                    with col_photo:
                        if player_photo_file:
                            # Show saved photo
                            try:
                                st.image(f"player_photos/{player_photo_file}", width=80, caption=report['Player Name'])
                            except:
                                # Fallback to SVG
                                st.markdown("""
                                <div style='text-align: center;'>
                                    <svg class='player-photo' xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'>
                                        <circle cx='50' cy='50' r='48' fill='#f8f9fa'/>
                                        <circle cx='50' cy='40' r='15' fill='#1a2332'/>
                                        <path d='M 25 75 Q 25 55 50 55 Q 75 55 75 75' fill='#1a2332'/>
                                    </svg>
                                </div>
                                """, unsafe_allow_html=True)
                        else:
                            # Player photo placeholder with default SVG icon
                            st.markdown("""
                            <div style='text-align: center;'>
                                <svg class='player-photo' xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'>
                                    <circle cx='50' cy='50' r='48' fill='#f8f9fa'/>
                                    <circle cx='50' cy='40' r='15' fill='#1a2332'/>
                                    <path d='M 25 75 Q 25 55 50 55 Q 75 55 75 75' fill='#1a2332'/>
                                </svg>
                            </div>
                            """, unsafe_allow_html=True)
                    
                    with col_upload:
                        # Only show uploader if no photo exists
                        if not player_photo_file:
                            # Upload player photo
                            uploaded_photo = st.file_uploader(
                                "📸 Upload Photo",
                                type=['jpg', 'jpeg', 'png'],
                                key=f"photo_upload_{idx}",
                                help="Upload player photo (JPG, PNG, max 5MB)"
                            )
                            
                            if uploaded_photo:
                                # Preview
                                st.image(uploaded_photo, width=80)
                                
                                # Save button
                                if st.button("💾 Save Photo", key=f"save_photo_{idx}", type="primary"):
                                    try:
                                        photo_filename = save_player_photo(uploaded_photo, report['Player Name'])
                                        st.success(f"✅ Photo saved!")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ Error: {e}")
                        else:
                            # Photo exists, show delete option
                            if st.button("🗑️ Delete Photo", key=f"delete_photo_{idx}"):
                                import os
                                try:
                                    os.remove(f"player_photos/{player_photo_file}")
                                    st.success("✅ Photo deleted!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Error: {e}")
                    
                    with col_info:
                        # Get team logo
                        team_logo = TEAM_LOGOS.get(report['Team'], '')
                        team_logo_html = ''
                        if team_logo:
                            try:
                                import base64
                                with open(team_logo, 'rb') as f:
                                    logo_data = base64.b64encode(f.read()).decode()
                                team_logo_html = f'<img src="data:image/png;base64,{logo_data}" style="height: 20px; vertical-align: middle; margin-right: 6px;">'
                            except:
                                team_logo_html = '🛡️ '
                        else:
                            team_logo_html = '🛡️ '
                        
                        # Player Name and Badges
                        position_html = f"<span class='position-badge'>{report['Position']}</span>"
                        team_html = f"<span class='team-badge'>{team_logo_html}{report['Team']}</span>"
                        
                        # Nationality (if exists in report, otherwise ask for it)
                        nationality = report.get('Nationality', 'Unknown')
                        nationality_icon = ''
                        if nationality == 'Saudi Arabia' or nationality == 'Unknown':
                            try:
                                with open('saff.png', 'rb') as f:
                                    nat_data = base64.b64encode(f.read()).decode()
                                nationality_icon = f'<img src="data:image/png;base64,{nat_data}" style="height: 16px; vertical-align: middle; margin-right: 4px;">'
                            except:
                                nationality_icon = '🇸🇦 '
                        
                        nationality_html = f"<span style='color: #666; font-size: 12px; margin-left: 8px;'>{nationality_icon}{nationality}</span>"
                        
                        st.markdown(f"<h3 style='color: #1a2332; margin: 0;'>{report['Player Name']}</h3>", unsafe_allow_html=True)
                        st.markdown(f"{position_html} {team_html} {nationality_html}", unsafe_allow_html=True)
                        st.markdown(f"<p style='color: #666; font-size: 13px;'>🎂 {report['Year of Birth']} • ⚽ {report['Match Name']} • 📅 {report['Date of the Report']}</p>", unsafe_allow_html=True)
                    
                    st.markdown("---")
                    
                    # Stats Section
                    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                    
                    with col_s1:
                        perf_pct = (report['Performance Rating'] / 6) * 100
                        st.markdown(f"""
                        <div class='stat-box'>
                            <div style='font-size: 11px; color: #666; font-weight: 600;'>⚡ PERFORMANCE</div>
                            <div style='font-size: 28px; color: #FFC60A; font-weight: 700;'>{report['Performance Rating']}/6</div>
                            <div style='background: #e0e0e0; height: 6px; border-radius: 3px; margin-top: 5px;'>
                                <div style='background: linear-gradient(90deg, #FFC60A, #FFD700); height: 100%; width: {perf_pct}%; border-radius: 3px;'></div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_s2:
                        pot_pct = (report['Potential Rating'] / 6) * 100
                        st.markdown(f"""
                        <div class='stat-box'>
                            <div style='font-size: 11px; color: #666; font-weight: 600;'>🌟 POTENTIAL</div>
                            <div style='font-size: 28px; color: #FFC60A; font-weight: 700;'>{report['Potential Rating']}/6</div>
                            <div style='background: #e0e0e0; height: 6px; border-radius: 3px; margin-top: 5px;'>
                                <div style='background: linear-gradient(90deg, #FFC60A, #FFD700); height: 100%; width: {pot_pct}%; border-radius: 3px;'></div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_s3:
                        conclusion_color = {'A - Firmar (Sign)': '#4CAF50', 'B+ - Seguir para Firmar (Follow to Sign)': '#2196F3', 'B - Seguir (Follow)': '#FF9800'}.get(report['Conclusion'], '#666')
                        st.markdown(f"""
                        <div class='stat-box'>
                            <div style='font-size: 11px; color: #666; font-weight: 600;'>✅ CONCLUSION</div>
                            <div style='font-size: 14px; color: {conclusion_color}; font-weight: 700; margin-top: 8px;'>{report['Conclusion'].split(' - ')[0]}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_s4:
                        # League icon
                        league_icon_html = ''
                        if 'Saudi' in str(report['League']):
                            try:
                                with open('jawwy.png', 'rb') as f:
                                    league_data = base64.b64encode(f.read()).decode()
                                league_icon_html = f'<img src="data:image/png;base64,{league_data}" style="height: 14px; vertical-align: middle; margin-right: 4px;">'
                            except:
                                league_icon_html = '🏆 '
                        else:
                            league_icon_html = '🏆 '
                        
                        st.markdown(f"""
                        <div class='stat-box'>
                            <div style='font-size: 11px; color: #666; font-weight: 600;'>{league_icon_html}LEAGUE</div>
                            <div style='font-size: 12px; color: #1a2332; font-weight: 600; margin-top: 8px;'>{report['League']}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Report Details
                    if report['Report Details'] and str(report['Report Details']) != 'nan':
                        with st.expander("📝 View Full Report"):
                            st.markdown(f"<div style='background: #fffef0; padding: 20px; border-left: 4px solid #FFC60A; border-radius: 8px;'>{report['Report Details']}</div>", unsafe_allow_html=True)
                    
                    st.markdown("<br>", unsafe_allow_html=True)
    
    except FileNotFoundError:
        st.info(f"📋 No reports found. Create your first report in the CREATE REPORT tab!")
    except Exception as e:
        st.error(f"Error loading reports: {e}")

def show_database_view(category):
    """Show database view for squad management"""
    if category != 'U21':
        st.info("📊 Database view is currently available for U21 category only.")
        return
    
    # Team logo mapping
    TEAM_LOGOS = {
        'Al-Adalah FC': 'aladalahclub.png',
        'Al-Ahli Saudi FC': 'alahli.png',
        'Al-Bukayriyah FC': 'albukiryahfc.png',
        'Al-Fateh SC': 'alfateh.png',
        'Al-Fayha FC': 'alfayha.png',
        'Al-Hazem FC': 'alhazem.png',
        'Al-Hilal SFC': 'alhilal.png',
        'Al-Ittihad Club': 'alittihad.png',
        'Al-Jabalain': 'aljabalain.png',
        'Al-Kholood Club': 'alkholood.png',
        'Al-Najmah SC': 'alnajma.png',
        'Al-Nassr FC': 'alnassr.png',
        'Al-Okhdood FC': 'alokhdood.png',
        'Al-Orobah FC': 'AlOrobah.png',
        'Al-Qadsiah FC': 'alqadsiah.png',
        'Al-Raed FC': 'alraed.png',
        'Al-Riyadh SC': 'alriyadh.png',
        'Al-Shabab FC': 'alshabab.png',
        'Al-Taawoun FC': 'altaawoun.png',
        'Al-Wehda FC': 'alwehda.png',
        'Damac FC': 'damac.png',
        'Ettifaq FC': 'alettifaq.png',
        'Khaleej FC': 'alkhaleej.png',
        'NEOM Club': 'neom.png'
    }
    
    # Conclusion color mapping (matching CREATE REPORT values)
    CONCLUSION_COLORS = {
        'SIGN (التوقيع معه)': '#4CAF50',  # Green
        'MONITOR CLOSELY (متابعة دقيقة)': '#FF9800',  # Orange
        'DISCARD (الاستبعاد)': '#F44336',  # Red
        '': '#002B5B'  # Default dark blue
    }
    
    # Load squad data
    df = load_squad_data(category)
    
    if df.empty:
        st.error("No data available. Please check the Excel file.")
        return
    
    # Custom CSS for professional scouting app style
    st.markdown("""
    <style>
        .metric-card {
            background-color: white;
            border-radius: 10px;
            padding: 1.5rem;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            text-align: center;
            margin-bottom: 1rem;
            border: 2px solid #002B5B;
        }
        .metric-value {
            font-size: 2.5rem;
            font-weight: bold;
            color: #002B5B;
        }
        .metric-label {
            font-size: 1rem;
            color: #666;
            margin-top: 0.5rem;
        }
        .team-header {
            background-color: #002B5B;
            color: white;
            padding: 1rem;
            border-radius: 8px;
            margin: 1rem 0;
            font-weight: bold;
            font-size: 1.2rem;
        }
        .stDataFrame {
            border: 2px solid #002B5B;
            border-radius: 10px;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Sidebar with counters
    with st.sidebar:
        st.markdown("### 📊 Statistics")
        
        total_teams = df['Team'].nunique()
        total_players = len(df)
        
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{total_teams}</div>
            <div class="metric-label">Total Teams</div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{total_players}</div>
            <div class="metric-label">Total U21 Players</div>
        </div>
        """, unsafe_allow_html=True)
    
    # Main content - Filters
    col_title, col_refresh = st.columns([3, 1])
    with col_title:
        st.markdown("### 🔍 Filters")
    with col_refresh:
        if st.button("🔄 Refresh Data", help="Reload data from Excel"):
            st.cache_data.clear()
            st.rerun()
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # League selector (fixed for U21)
        league = st.selectbox("Select League", ["Saudi Elite League U-21"], key="league_filter")
    
    with col2:
        # Team filter
        teams = ['All Teams'] + sorted(df['Team'].unique().tolist())
        selected_team = st.selectbox("Select Team", teams, key="team_filter")
    
    with col3:
        # Search by player name
        search_name = st.text_input("Search Player", placeholder="Enter player name...", key="search_player")
    
    # Filter dataframe
    filtered_df = df.copy()
    
    if selected_team != 'All Teams':
        filtered_df = filtered_df[filtered_df['Team'] == selected_team]
    
    if search_name:
        filtered_df = filtered_df[filtered_df['Name'].str.contains(search_name, case=False, na=False)]
    
    # Display players by team
    st.markdown("---")
    st.markdown("### 👥 Squad Database")
    
    # Info message about changes
    st.info("💡 **Tip:** After editing player details, click 'Save All Changes' at the bottom to save to Excel and refresh the view with updated colors and positions.")
    
    if filtered_df.empty:
        st.warning("No players found with the selected filters.")
        return
    
    # Group by team
    teams_to_display = filtered_df['Team'].unique()
    
    # Position options for dropdown (matching CREATE REPORT exactly)
    POSITION_OPTIONS = ['', 'GK', 'RB', 'RWB', 'CB', 'LB', 'LWB', 'DM/6', 'CM/8', 'AM/10', 'RW/WF', 'LW/WF', 'ST/9', 'SS/9.5']
    
    # Conclusion options (matching CREATE REPORT)
    CONCLUSION_OPTIONS = {
        '': '',
        '🟢 SIGN': 'SIGN (التوقيع معه)',
        '🟠 MONITOR CLOSELY': 'MONITOR CLOSELY (متابعة دقيقة)',
        '🔴 DISCARD': 'DISCARD (الاستبعاد)'
    }
    CONCLUSION_REVERSE = {v: k for k, v in CONCLUSION_OPTIONS.items()}
    
    # Store changes
    changes_made = False
    
    for team in sorted(teams_to_display):
        team_players = filtered_df[filtered_df['Team'] == team]
        
        # Team header with logo
        logo_file = TEAM_LOGOS.get(team, None)
        if logo_file:
            try:
                import io
                import base64
                team_logo = Image.open(logo_file)
                buffered = io.BytesIO()
                team_logo.save(buffered, format="PNG")
                img_str = base64.b64encode(buffered.getvalue()).decode()
                st.markdown(
                    f'<div class="team-header"><img src="data:image/png;base64,{img_str}" style="height: 40px; vertical-align: middle; margin-right: 10px;">{team} ({len(team_players)} players)</div>',
                    unsafe_allow_html=True
                )
            except:
                st.markdown(f'<div class="team-header">⚽ {team} ({len(team_players)} players)</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="team-header">⚽ {team} ({len(team_players)} players)</div>', unsafe_allow_html=True)
        
        # Create editable table for each player
        for idx, row in team_players.iterrows():
            # Get conclusion color
            conclusion_value = row['Conclusion'] if pd.notna(row['Conclusion']) else ''
            conclusion_color = CONCLUSION_COLORS.get(conclusion_value, CONCLUSION_COLORS[''])
            
            # Player info for banner
            player_pos = row['Position'] if pd.notna(row['Position']) else 'N/A'
            player_year = int(row['Year of Birth']) if pd.notna(row['Year of Birth']) else 'N/A'
            
            # Create columns for photo and banner
            col_photo_db, col_banner = st.columns([1, 5])
            
            with col_photo_db:
                # Check if player has a saved photo
                player_photo_file = get_player_photo(row['Name'])
                
                if player_photo_file:
                    # Show saved photo
                    st.image(f"player_photos/{player_photo_file}", width=80)
                    # Delete button
                    if st.button("🗑️", key=f"delete_db_photo_{idx}", help="Delete photo"):
                        import os
                        try:
                            os.remove(f"player_photos/{player_photo_file}")
                            st.success("✅ Deleted!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ {e}")
                else:
                    # Show SVG placeholder
                    st.markdown("""
                    <div style='text-align: center;'>
                        <svg width='80' height='80' xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'>
                            <circle cx='50' cy='50' r='48' fill='#f8f9fa' stroke='#FFC60A' stroke-width='3'/>
                            <circle cx='50' cy='40' r='15' fill='#1a2332'/>
                            <path d='M 25 75 Q 25 55 50 55 Q 75 55 75 75' fill='#1a2332'/>
                        </svg>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Upload option (only if no photo exists)
                    uploaded_db_photo = st.file_uploader(
                        "📸",
                        type=['jpg', 'jpeg', 'png'],
                        key=f"photo_db_{idx}",
                        help="Upload photo",
                        label_visibility="collapsed"
                    )
                    
                    if uploaded_db_photo:
                        if st.button("💾", key=f"save_db_photo_{idx}", help="Save photo"):
                            try:
                                save_player_photo(uploaded_db_photo, row['Name'])
                                st.success("✅")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ {e}")
            
            with col_banner:
                # Custom styled banner
                st.markdown(
                    f'<div style="background-color: {conclusion_color}; color: white; padding: 0.8rem 1rem; border-radius: 8px; margin: 0.5rem 0; font-weight: bold; font-size: 1.1rem;">'
                    f'#{row["Number"]} - {row["Name"]} | Position: {player_pos} | Born: {player_year}'
                    f'</div>',
                    unsafe_allow_html=True
                )
            
            with st.expander("✏️ Edit Player Details", expanded=False):
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    # Editable Number
                    new_number = st.number_input(
                        "Number",
                        min_value=0,
                        max_value=99,
                        value=int(row['Number']) if pd.notna(row['Number']) else 0,
                        key=f"number_{idx}"
                    )
                    if new_number != row['Number']:
                        df.loc[idx, 'Number'] = new_number
                        changes_made = True
                    
                    st.markdown(f"**Name:** {row['Name']}")
                
                with col2:
                    # Editable Position
                    current_pos = row['Position'] if pd.notna(row['Position']) else ''
                    new_position = st.selectbox(
                        "Position",
                        options=POSITION_OPTIONS,
                        index=POSITION_OPTIONS.index(current_pos) if current_pos in POSITION_OPTIONS else 0,
                        key=f"pos_{idx}"
                    )
                    if new_position != current_pos:
                        df.loc[idx, 'Position'] = new_position
                        changes_made = True
                
                with col3:
                    st.markdown(f"**Year of Birth:** {row['Year of Birth']}")
                    st.markdown(f"**Age:** {2025 - row['Year of Birth']}")
                
                with col4:
                    # KSA Caps
                    ksa_caps = st.number_input(
                        "KSA Caps",
                        min_value=0,
                        max_value=200,
                        value=int(row['KSA Caps']) if pd.notna(row['KSA Caps']) else 0,
                        key=f"caps_{idx}"
                    )
                    if ksa_caps != row['KSA Caps']:
                        df.loc[idx, 'KSA Caps'] = ksa_caps
                        changes_made = True
                
                # Conclusion and Smart Evaluation
                col5, col6 = st.columns(2)
                
                with col5:
                    current_conclusion = row['Conclusion'] if pd.notna(row['Conclusion']) else ''
                    conclusion_display = CONCLUSION_REVERSE.get(current_conclusion, '')
                    new_conclusion_display = st.selectbox(
                        "Conclusion",
                        options=list(CONCLUSION_OPTIONS.keys()),
                        index=list(CONCLUSION_OPTIONS.keys()).index(conclusion_display) if conclusion_display in CONCLUSION_OPTIONS else 0,
                        key=f"conclusion_{idx}"
                    )
                    new_conclusion = CONCLUSION_OPTIONS[new_conclusion_display]
                    if new_conclusion != current_conclusion:
                        df.loc[idx, 'Conclusion'] = new_conclusion
                        changes_made = True
                
                with col6:
                    # Smart Evaluation
                    current_smart_eval = row['Smart Evaluation'] if pd.notna(row['Smart Evaluation']) else ''
                    new_smart_eval = st.text_area(
                        "Smart Evaluation",
                        value=current_smart_eval,
                        key=f"smart_eval_{idx}",
                        height=100
                    )
                    if new_smart_eval != current_smart_eval:
                        df.loc[idx, 'Smart Evaluation'] = new_smart_eval
                        changes_made = True
        
        st.markdown("<br>", unsafe_allow_html=True)
    
    # Save button
    if changes_made:
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("💾 Save All Changes", use_container_width=True, type="primary"):
                if save_squad_data(df, category):
                    st.rerun()

# -----------------------------
# ADVANCED DATA ANALYSIS - SAFF+
# -----------------------------
def show_advanced_data_analysis(category):
    """Show advanced data analysis from u18u17.xlsx with dependent filters"""
    
    # Header
    st.markdown("<h2 style='text-align: center; color: #002B5B;'>🔬 ADVANCED DATA ANALYSIS - SAFF+</h2>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #666;'>Analysis of U18 and U17 players data</p>", unsafe_allow_html=True)
    
    try:
        # Load data from u18u17.xlsx
        df_u18 = pd.read_excel('u18u17.xlsx', sheet_name='Datos Completos U18')
        df_u17 = pd.read_excel('u18u17.xlsx', sheet_name='Datos Completos U17')
        df_both = pd.read_excel('u18u17.xlsx', sheet_name='Jugadores en Ambas Ligas')
    except Exception as e:
        st.error(f"❌ Error loading data: {e}")
        st.info("📝 Please ensure u18u17.xlsx file exists with sheets: 'Datos Completos U18', 'Datos Completos U17', 'Jugadores en Ambas Ligas'")
        return
    
    # Function to extract Year of Birth from birthday column
    def extract_year_of_birth(birthday_value):
        """Extract year from birthday field - handles various formats"""
        if pd.isna(birthday_value):
            return None
        try:
            # Convert to string and extract first 4 digits (year)
            birthday_str = str(birthday_value)
            # Try to find 4 consecutive digits (year)
            import re
            year_match = re.search(r'(\d{4})', birthday_str)
            if year_match:
                year = int(year_match.group(1))
                # Validate year is reasonable (between 1990 and 2015 for youth players)
                if 1990 <= year <= 2015:
                    return year
        except:
            pass
        return None
    
    # Add Year of Birth column to all dataframes and convert numeric columns
    for df in [df_u18, df_u17, df_both]:
        if 'birthday' in df.columns:
            df['Year of Birth'] = df['birthday'].apply(extract_year_of_birth)
        
        # Convert numeric columns to proper types
        numeric_columns = ['MP', 'MSt', 'MSub', 'Goals', 'Cards', 'Minutes', 
                          'U18_MP', 'U18_Minutes', 'U18_Goals', 'U18_Cards',
                          'U17_MP', 'U17_Minutes', 'U17_Goals', 'U17_Cards',
                          'Total_MP', 'Total_Minutes', 'Total_Goals']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    
    # For Both Leagues, add position from U18 data
    if not df_both.empty and 'position' not in df_both.columns:
        # Create a mapping of player_name to position from U18 data
        if not df_u18.empty and 'position' in df_u18.columns and 'player_name' in df_u18.columns:
            position_map = df_u18.set_index('player_name')['position'].to_dict()
            df_both['position'] = df_both['player_name'].map(position_map)
    
    # Custom CSS
    st.markdown("""
    <style>
        .filter-header {
            background: linear-gradient(135deg, #1a2332 0%, #2a3342 100%);
            color: white;
            padding: 15px;
            border-radius: 8px;
            border-left: 5px solid #FFC60A;
            margin-bottom: 20px;
            font-weight: 600;
        }
        .stat-card {
            background: white;
            padding: 20px;
            border-radius: 10px;
            border: 2px solid #FFC60A;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            text-align: center;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Data source selector
    st.markdown("<div class='filter-header'>🏆 Select Data Source</div>", unsafe_allow_html=True)
    league_option = st.radio(
        "",
        ["U18 League", "U17 League", "Players in Both Leagues"],
        horizontal=True,
        label_visibility="collapsed"
    )
    
    # Select appropriate dataframe
    if league_option == "U18 League":
        df_original = df_u18.copy()
        league_title = "U18"
    elif league_option == "U17 League":
        df_original = df_u17.copy()
        league_title = "U17"
    else:
        df_original = df_both.copy()
        league_title = "Both Leagues"
    
    if df_original.empty:
        st.warning("⚠️ No data available")
        return
    
    st.markdown(f"<h3 style='color: #1a2332;'>📄 {league_title} - Player Statistics</h3>", unsafe_allow_html=True)
    
    # Initialize session state for filters
    if 'clear_filters' not in st.session_state:
        st.session_state.clear_filters = False
    
    # Sidebar filters with dependencies
    with st.sidebar:
        st.markdown("### 🔍 Filters")
        st.markdown("---")
        
        # Start with full dataset
        filtered_df = df_original.copy()
        
        # League filter
        if 'league' in filtered_df.columns:
            leagues_available = ['All'] + sorted(filtered_df['league'].dropna().unique().tolist())
            selected_league = st.selectbox("🏆 League", leagues_available, key="filter_league")
            if selected_league != 'All':
                filtered_df = filtered_df[filtered_df['league'] == selected_league]
        
        # Team filter (dependent on League)
        team_col = None
        if 'team' in filtered_df.columns:
            team_col = 'team'
        elif 'team_U18' in filtered_df.columns:
            team_col = 'team_U18'
        
        if team_col:
            teams_available = ['All'] + sorted(filtered_df[team_col].dropna().unique().tolist())
            selected_team = st.selectbox("🛡️ Team", teams_available, key="filter_team")
            if selected_team != 'All':
                if team_col == 'team_U18':
                    # For Both Leagues, filter by either U18 or U17 team
                    filtered_df = filtered_df[(filtered_df['team_U18'] == selected_team) | (filtered_df['team_U17'] == selected_team)]
                else:
                    filtered_df = filtered_df[filtered_df[team_col] == selected_team]
        
        # Position filter (dependent on previous filters)
        if 'position' in filtered_df.columns:
            positions_available = ['All'] + sorted([p for p in filtered_df['position'].dropna().unique().tolist() if str(p).strip()])
            selected_position = st.selectbox("🎯 Position", positions_available, key="filter_position")
            if selected_position != 'All':
                filtered_df = filtered_df[filtered_df['position'] == selected_position]
        
        # Jersey filter (dependent on previous filters)
        if 'jersey' in filtered_df.columns:
            jerseys_available = ['All'] + sorted([int(j) for j in filtered_df['jersey'].dropna().unique().tolist() if str(j).strip() and str(j) != 'nan'])
            if jerseys_available and len(jerseys_available) > 1:
                selected_jersey = st.selectbox("🔢 Jersey", jerseys_available, key="filter_jersey")
                if selected_jersey != 'All':
                    filtered_df = filtered_df[filtered_df['jersey'] == selected_jersey]
        
        # Year of Birth filter (dependent on previous filters)
        if 'Year of Birth' in filtered_df.columns:
            years_available = sorted([y for y in filtered_df['Year of Birth'].dropna().unique().tolist()], reverse=True)
            if years_available:
                years_options = ['All'] + [int(y) for y in years_available]
                selected_year = st.selectbox("🎂 Year of Birth", years_options, key="filter_year")
                if selected_year != 'All':
                    filtered_df = filtered_df[filtered_df['Year of Birth'] == selected_year]
        
        # Matches Played (MP) slider
        mp_col = None
        if 'MP' in filtered_df.columns:
            mp_col = 'MP'
        elif 'Total_MP' in filtered_df.columns:
            mp_col = 'Total_MP'
        
        if mp_col:
            mp_values = filtered_df[mp_col].dropna()
            if not mp_values.empty:
                min_mp = int(mp_values.min())
                max_mp = int(mp_values.max())
                if min_mp < max_mp:
                    selected_mp = st.slider(
                        "⚽ Matches Played (MP)",
                        min_value=min_mp,
                        max_value=max_mp,
                        value=(min_mp, max_mp),
                        key="filter_mp"
                    )
                    filtered_df = filtered_df[(filtered_df[mp_col] >= selected_mp[0]) & (filtered_df[mp_col] <= selected_mp[1])]
                    st.caption(f"Range: {selected_mp[0]} - {selected_mp[1]} matches")
        
        # Minutes slider
        minutes_col = None
        if 'Minutes' in filtered_df.columns:
            minutes_col = 'Minutes'
        elif 'Total_Minutes' in filtered_df.columns:
            minutes_col = 'Total_Minutes'
        
        if minutes_col:
            minutes_values = filtered_df[minutes_col].dropna()
            if not minutes_values.empty:
                min_minutes = int(minutes_values.min())
                max_minutes = int(minutes_values.max())
                if min_minutes < max_minutes:
                    selected_minutes = st.slider(
                        "⏱️ Minutes" if minutes_col != 'Total_Minutes' else "⏱️ Total Minutes",
                        min_value=min_minutes,
                        max_value=max_minutes,
                        value=(min_minutes, max_minutes),
                        key="filter_minutes"
                    )
                    filtered_df = filtered_df[(filtered_df[minutes_col] >= selected_minutes[0]) & (filtered_df[minutes_col] <= selected_minutes[1])]
                    st.caption(f"Range: {selected_minutes[0]} - {selected_minutes[1]} min")
        
        # Additional filters for "Players in Both Leagues"
        if league_option == "Players in Both Leagues":
            st.markdown("---")
            st.markdown("**🔄 Both Leagues Filters**")
            
            # U18 MP filter
            if 'U18_MP' in filtered_df.columns:
                u18_mp_values = filtered_df['U18_MP'].dropna()
                if not u18_mp_values.empty:
                    min_u18_mp = int(u18_mp_values.min())
                    max_u18_mp = int(u18_mp_values.max())
                    if min_u18_mp < max_u18_mp:
                        selected_u18_mp = st.slider(
                            "⚽ U18 Matches Played",
                            min_value=min_u18_mp,
                            max_value=max_u18_mp,
                            value=(min_u18_mp, max_u18_mp),
                            key="filter_u18_mp"
                        )
                        filtered_df = filtered_df[(filtered_df['U18_MP'] >= selected_u18_mp[0]) & (filtered_df['U18_MP'] <= selected_u18_mp[1])]
            
            # U17 MP filter
            if 'U17_MP' in filtered_df.columns:
                u17_mp_values = filtered_df['U17_MP'].dropna()
                if not u17_mp_values.empty:
                    min_u17_mp = int(u17_mp_values.min())
                    max_u17_mp = int(u17_mp_values.max())
                    if min_u17_mp < max_u17_mp:
                        selected_u17_mp = st.slider(
                            "⚽ U17 Matches Played",
                            min_value=min_u17_mp,
                            max_value=max_u17_mp,
                            value=(min_u17_mp, max_u17_mp),
                            key="filter_u17_mp"
                        )
                        filtered_df = filtered_df[(filtered_df['U17_MP'] >= selected_u17_mp[0]) & (filtered_df['U17_MP'] <= selected_u17_mp[1])]
            
            # U18 Minutes filter
            if 'U18_Minutes' in filtered_df.columns:
                u18_min_values = filtered_df['U18_Minutes'].dropna()
                if not u18_min_values.empty:
                    min_u18_min = int(u18_min_values.min())
                    max_u18_min = int(u18_min_values.max())
                    if min_u18_min < max_u18_min:
                        selected_u18_min = st.slider(
                            "⏱️ U18 Minutes",
                            min_value=min_u18_min,
                            max_value=max_u18_min,
                            value=(min_u18_min, max_u18_min),
                            key="filter_u18_min"
                        )
                        filtered_df = filtered_df[(filtered_df['U18_Minutes'] >= selected_u18_min[0]) & (filtered_df['U18_Minutes'] <= selected_u18_min[1])]
            
            # U17 Minutes filter
            if 'U17_Minutes' in filtered_df.columns:
                u17_min_values = filtered_df['U17_Minutes'].dropna()
                if not u17_min_values.empty:
                    min_u17_min = int(u17_min_values.min())
                    max_u17_min = int(u17_min_values.max())
                    if min_u17_min < max_u17_min:
                        selected_u17_min = st.slider(
                            "⏱️ U17 Minutes",
                            min_value=min_u17_min,
                            max_value=max_u17_min,
                            value=(min_u17_min, max_u17_min),
                            key="filter_u17_min"
                        )
                        filtered_df = filtered_df[(filtered_df['U17_Minutes'] >= selected_u17_min[0]) & (filtered_df['U17_Minutes'] <= selected_u17_min[1])]
            
            # U18 Goals filter
            if 'U18_Goals' in filtered_df.columns:
                u18_goals_values = filtered_df['U18_Goals'].dropna()
                if not u18_goals_values.empty:
                    min_u18_goals = int(u18_goals_values.min())
                    max_u18_goals = int(u18_goals_values.max())
                    if min_u18_goals < max_u18_goals:
                        selected_u18_goals = st.slider(
                            "⚽ U18 Goals",
                            min_value=min_u18_goals,
                            max_value=max_u18_goals,
                            value=(min_u18_goals, max_u18_goals),
                            key="filter_u18_goals"
                        )
                        filtered_df = filtered_df[(filtered_df['U18_Goals'] >= selected_u18_goals[0]) & (filtered_df['U18_Goals'] <= selected_u18_goals[1])]
            
            # U17 Goals filter
            if 'U17_Goals' in filtered_df.columns:
                u17_goals_values = filtered_df['U17_Goals'].dropna()
                if not u17_goals_values.empty:
                    min_u17_goals = int(u17_goals_values.min())
                    max_u17_goals = int(u17_goals_values.max())
                    if min_u17_goals < max_u17_goals:
                        selected_u17_goals = st.slider(
                            "⚽ U17 Goals",
                            min_value=min_u17_goals,
                            max_value=max_u17_goals,
                            value=(min_u17_goals, max_u17_goals),
                            key="filter_u17_goals"
                        )
                        filtered_df = filtered_df[(filtered_df['U17_Goals'] >= selected_u17_goals[0]) & (filtered_df['U17_Goals'] <= selected_u17_goals[1])]
            
            # Total Goals filter
            if 'Total_Goals' in filtered_df.columns:
                total_goals_values = filtered_df['Total_Goals'].dropna()
                if not total_goals_values.empty:
                    min_total_goals = int(total_goals_values.min())
                    max_total_goals = int(total_goals_values.max())
                    if min_total_goals < max_total_goals:
                        selected_total_goals = st.slider(
                            "🎯 Total Goals",
                            min_value=min_total_goals,
                            max_value=max_total_goals,
                            value=(min_total_goals, max_total_goals),
                            key="filter_total_goals"
                        )
                        filtered_df = filtered_df[(filtered_df['Total_Goals'] >= selected_total_goals[0]) & (filtered_df['Total_Goals'] <= selected_total_goals[1])]
        
        # Quick search by player name
        st.markdown("---")
        player_search = st.text_input("🔍 Quick search (Player name)", placeholder="Type player name...", key="filter_search")
        if player_search:
            if 'player_name' in filtered_df.columns:
                filtered_df = filtered_df[filtered_df['player_name'].str.contains(player_search, case=False, na=False)]
        
        # Clear filters button
        st.markdown("---")
        if st.button("🗑️ Clear All Filters", use_container_width=True):
            # Clear all filter keys
            filter_keys = ['filter_league', 'filter_team', 'filter_position', 'filter_jersey', 'filter_year', 
                          'filter_mp', 'filter_minutes', 'filter_search',
                          'filter_u18_mp', 'filter_u17_mp', 'filter_u18_min', 'filter_u17_min',
                          'filter_u18_goals', 'filter_u17_goals', 'filter_total_goals']
            for key in filter_keys:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()
    
    # Main content
    st.markdown("---")
    
    # Results count
    total_original = len(df_original)
    total_filtered = len(filtered_df)
    st.markdown(f"<h4 style='color: #1a2332;'>📄 Showing {total_filtered} of {total_original} players</h4>", unsafe_allow_html=True)
    
    if filtered_df.empty:
        st.warning("🗒️ No players found with the selected filters.")
    else:
        # Download buttons
        create_download_buttons(
            filtered_df,
            filename_base=f"SAFF_Plus_{league_title}_filtered",
            label_prefix="Download Data"
        )
        
        st.markdown("---")
        
        # Sort by Goals descending if column exists
        goals_col = None
        if 'Goals' in filtered_df.columns:
            goals_col = 'Goals'
        elif 'Total_Goals' in filtered_df.columns:
            goals_col = 'Total_Goals'
        
        if goals_col:
            filtered_df = filtered_df.sort_values(goals_col, ascending=False)
        
        # Display data table with pagination
        st.dataframe(
            filtered_df,
            use_container_width=True,
            height=500
        )
        
        # Statistics summary
        st.markdown("---")
        st.markdown("### 📊 Statistics Summary")
        
        col_s1, col_s2, col_s3, col_s4 = st.columns(4)
        
        with col_s1:
            goals_col = 'Goals' if 'Goals' in filtered_df.columns else 'Total_Goals' if 'Total_Goals' in filtered_df.columns else None
            if goals_col:
                try:
                    total_goals = int(filtered_df[goals_col].sum())
                    avg_goals = filtered_df[goals_col].mean()
                    st.markdown(f"""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>⚽</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>{total_goals}</div>
                        <div style='color: #666;'>Total Goals</div>
                        <div style='font-size: 12px; color: #999;'>Avg: {avg_goals:.2f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                except:
                    st.markdown("""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>⚽</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>N/A</div>
                        <div style='color: #666;'>Total Goals</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        with col_s2:
            minutes_col = 'Minutes' if 'Minutes' in filtered_df.columns else 'Total_Minutes' if 'Total_Minutes' in filtered_df.columns else None
            if minutes_col:
                try:
                    avg_minutes = filtered_df[minutes_col].mean()
                    max_minutes = filtered_df[minutes_col].max()
                    st.markdown(f"""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>⏱️</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>{int(avg_minutes)}</div>
                        <div style='color: #666;'>Avg Minutes</div>
                        <div style='font-size: 12px; color: #999;'>Max: {int(max_minutes)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                except:
                    st.markdown("""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>⏱️</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>N/A</div>
                        <div style='color: #666;'>Avg Minutes</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        with col_s3:
            if 'Cards' in filtered_df.columns:
                try:
                    total_cards = int(filtered_df['Cards'].sum())
                    st.markdown(f"""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>🟨</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>{total_cards}</div>
                        <div style='color: #666;'>Total Cards</div>
                    </div>
                    """, unsafe_allow_html=True)
                except:
                    st.markdown("""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>🟨</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>N/A</div>
                        <div style='color: #666;'>Total Cards</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        with col_s4:
            mp_col = 'MP' if 'MP' in filtered_df.columns else 'Total_MP' if 'Total_MP' in filtered_df.columns else None
            if mp_col:
                try:
                    total_mp = int(filtered_df[mp_col].sum())
                    avg_mp = filtered_df[mp_col].mean()
                    st.markdown(f"""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>⚽</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>{total_mp}</div>
                        <div style='color: #666;'>Total Matches</div>
                        <div style='font-size: 12px; color: #999;'>Avg: {avg_mp:.1f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                except:
                    st.markdown("""
                    <div class='stat-card'>
                        <div style='font-size: 32px; color: #FFC60A;'>⚽</div>
                        <div style='font-size: 28px; font-weight: bold; color: #002B5B;'>N/A</div>
                        <div style='color: #666;'>Total Matches</div>
                    </div>
                    """, unsafe_allow_html=True)

# -----------------------------
# CATEGORY VIEW (with tabs)
# -----------------------------
def show_category_view():
    category = st.session_state.category
    
    if not category:
        st.session_state.page = 'categories'
        st.rerun()
        return
    
    # Back button and Header
    col_back, col_title = st.columns([1, 5])
    
    with col_back:
        if st.button("⬅️ Back" if st.session_state.language == 'en' else "⬅️ رجوع", 
                    key="btn_back_to_categories",
                    help="Return to categories" if st.session_state.language == 'en' else "العودة إلى الفئات"):
            st.session_state.page = 'categories'
            st.rerun()
    
    with col_title:
        # Show U21 SAUDI LEAGUES with logo for U21 category
        if category == 'U21':
            try:
                from PIL import Image
                import io
                u21_logo = Image.open('U21logo.png')
                u21_logo.thumbnail((50, 50))
                buffered = io.BytesIO()
                if u21_logo.mode in ('RGBA', 'LA', 'P'):
                    u21_logo.save(buffered, format="PNG")
                else:
                    u21_logo = u21_logo.convert('RGB')
                    u21_logo.save(buffered, format="PNG")
                u21_logo_str = base64.b64encode(buffered.getvalue()).decode()
                logo_html = f'<img src="data:image/png;base64,{u21_logo_str}" style="height:45px; vertical-align:middle; margin-right:15px;">'
            except:
                logo_html = ''
            
            st.markdown(f"<h2 style='text-align: center; color: #002B5B;'>{logo_html}U21 SAUDI LEAGUES</h2>", unsafe_allow_html=True)
        else:
            st.markdown(f"<h2 style='text-align: center; color: #002B5B;'>{category}</h2>", unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Tabs - Show CALENDAR only for U21
    if category == 'U21':
        if st.session_state.language == 'en':
            tabs = st.tabs(["📝 CREATE REPORT", "📊 VIEW REPORTS", "🗄️ VIEW DATABASE", "📈 ANALYTICS", "📅 CALENDAR - SCHEDULE"])
        else:
            tabs = st.tabs(["📝 إنشاء تقرير", "📊 عرض التقارير", "🗄️ قاعدة البيانات", "📈 التحليلات", "📅 التقويم - الجدول"])
    else:
        if st.session_state.language == 'en':
            tabs = st.tabs(["📝 CREATE REPORT", "📊 VIEW REPORTS", "🗄️ VIEW DATABASE", "📈 ANALYTICS"])
        else:
            tabs = st.tabs(["📝 إنشاء تقرير", "📊 عرض التقارير", "🗄️ قاعدة البيانات", "📈 التحليلات"])
    
    with tabs[0]:
        show_create_report_form(category)
    
    with tabs[1]:
        show_category_reports(category)
    
    with tabs[2]:
        show_database_view(category)
    
    with tabs[3]:
        if st.session_state.language == 'en':
            st.info("📈 Analytics will be implemented soon.")
        else:
            st.info("📈 سيتم تنفيذ التحليلات قريباً.")
    
    # Calendar tab - only for U21
    if category == 'U21':
        with tabs[4]:
            show_calendar_schedule(category)

# -----------------------------
# FIFA WORLD CUP U20 VIEW
# -----------------------------
def show_fifa_u20_view():
    """Show FIFA U20 World Cup section with 5 tabs"""
    
    # Country flags mapping (used across multiple tabs)
    COUNTRY_FLAGS = {
        'Arabia Saudita': 'saff.png',
        'Argentina': 'afa.png',
        'Australia': 'australia.png',
        'Brasil': 'brasil.png',
        'Chile': 'chile.png',
        'Colombia': 'colombiau20.png',
        'Corea del Sur': 'korea.png',
        'Cuba': 'cuba.png',
        'Egipto': 'egipto.webp',
        'España': 'spain.png',
        'Estados Unidos': 'usa.png',
        'Francia': 'francia.png',
        'Italia': 'italia.png',
        'Japón': 'japan.png',
        'Marruecos': 'marruecos.png',
        'México': 'mexico.svg',
        'Nigeria': 'nigeria.png',
        'Noruega': 'noruega.png',
        'Nueva Caledonia': 'nuevacaledonia.png',
        'Nueva Zelanda': 'nuevazelanda.png',
        'Panamá': 'panama.png',
        'Paraguay': 'paraguay.webp',
        'Sudáfrica': 'sudafrica.png',
        'Ucrania': 'ucrania.png'
    }
    
    # Back button and Header
    col_back, col_title = st.columns([1, 5])
    
    with col_back:
        if st.button("⬅️ Back" if st.session_state.language == 'en' else "⬅️ رجوع", 
                    key="btn_back_to_home_fifa",
                    help="Return to home" if st.session_state.language == 'en' else "العودة إلى الرئيسية"):
            st.session_state.page = 'home'
            st.rerun()
    
    with col_title:
        # Load FIFA U20 icon
        try:
            from PIL import Image
            import io
            
            icon_img = Image.open('iconfifau20.png')
            icon_img.thumbnail((40, 40))  # Resize icon
            buffered = io.BytesIO()
            
            if icon_img.mode in ('RGBA', 'LA', 'P'):
                icon_img.save(buffered, format="PNG")
                img_type = "png"
            else:
                icon_img = icon_img.convert('RGB')
                icon_img.save(buffered, format="PNG")
                img_type = "png"
            
            icon_str = base64.b64encode(buffered.getvalue()).decode()
            icon_html = f'<img src="data:image/{img_type};base64,{icon_str}" style="height:35px; vertical-align:middle; margin-right:10px;">'
        except:
            icon_html = '⚽ '  # Fallback to emoji if image fails
        
        if st.session_state.language == 'en':
            st.markdown(f"<h2 style='color: #002B5B;'>{icon_html}FIFA WORLD CUP U20</h2>", unsafe_allow_html=True)
        else:
            st.markdown(f"<h2 style='color: #002B5B;'>{icon_html}كأس العالم تحت 20</h2>", unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Create 7 tabs
    if st.session_state.language == 'en':
        tabs = st.tabs([
            "📝 CREATE MATCH REPORT",
            "👤 CREATE INDIVIDUAL REPORT",
            "📊 INDIVIDUAL REPORTS",
            "📅 SCHEDULE - MATCHES",
            "🗄️ PLAYER DATABASE",
            "📈 MATCH REPORTS",
            "⚽ CAMPOGRAMAS"
        ])
    else:
        tabs = st.tabs([
            "📝 إنشاء تقرير مباراة",
            "👤 إنشاء تقرير فردي",
            "📊 تقارير فردية",
            "📅 التقويم - المباريات",
            "🗄️ قاعدة بيانات اللاعبين",
            "📈 تقارير المباريات",
            "⚽ رسوم الملعب"
        ])
    
    # Tab 1: CREATE MATCH REPORT - Al Nassr Design
    with tabs[0]:
        # Al Nassr Custom CSS for Form
        st.markdown("""
            <style>
            /* Al Nassr Form Styling */
            .stTextInput > div > div > input,
            .stDateInput > div > div > input,
            .stSelectbox > div > div > select {
                background-color: #f5f5f5 !important;
                border: 2px solid #e0e0e0 !important;
                border-radius: 6px !important;
                transition: all 0.3s ease !important;
            }
            
            .stTextInput > div > div > input:focus,
            .stDateInput > div > div > input:focus,
            .stSelectbox > div > div > select:focus {
                border-color: #FFC60A !important;
                box-shadow: 0 0 0 3px rgba(255,198,10,0.1) !important;
            }
            
            .alnassr-header {
                background: #1a2332;
                color: white;
                padding: 15px 20px;
                border-radius: 8px;
                margin: 20px 0 15px 0;
                font-size: 16px;
                font-weight: 600;
                display: flex;
                align-items: center;
                gap: 10px;
            }
            
            .alnassr-title {
                display: flex;
                align-items: center;
                gap: 15px;
                margin-bottom: 30px;
            }
            </style>
        """, unsafe_allow_html=True)
        
        # Load Al Nassr Logo
        try:
            from PIL import Image
            import io
            logo_img = Image.open('alnassr.png')
            logo_img.thumbnail((50, 50))
            buffered = io.BytesIO()
            if logo_img.mode in ('RGBA', 'LA', 'P'):
                logo_img.save(buffered, format="PNG")
            else:
                logo_img = logo_img.convert('RGB')
                logo_img.save(buffered, format="PNG")
            logo_str = base64.b64encode(buffered.getvalue()).decode()
            logo_html = f'<img src="data:image/png;base64,{logo_str}" style="height:50px; vertical-align:middle;">'
        except:
            logo_html = '🦅'
        
        # Page Title with Logo
        st.markdown(f"""
            <div class="alnassr-title">
                {logo_html}
                <div>
                    <h2 style="margin:0; color:#1a2332;">{"Create Match Report" if st.session_state.language == 'en' else "إنشاء تقرير مباراة"}</h2>
                    <p style="margin:0; color:#666; font-size:14px;">FIFA U20 World Cup</p>
                </div>
            </div>
        """, unsafe_allow_html=True)
        
        # List of 24 countries
        COUNTRIES_U20 = [
            'Arabia Saudita', 'Argentina', 'Australia', 'Brasil', 'Chile', 'Colombia',
            'Corea del Sur', 'Cuba', 'Egipto', 'España', 'Estados Unidos', 'Francia',
            'Italia', 'Japón', 'Marruecos', 'México', 'Nigeria', 'Noruega',
            'Nueva Caledonia', 'Nueva Zelanda', 'Panamá', 'Paraguay', 'Sudáfrica', 'Ucrania'
        ]
        
        # Match Information Section - Navy Header
        st.markdown("""
            <div class="alnassr-header">
                ℹ️ Match Information
            </div>
        """ if st.session_state.language == 'en' else """
            <div class="alnassr-header">
                ℹ️ معلومات المباراة
            </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            match_date = st.date_input(
                "Match Date" if st.session_state.language == 'en' else "تاريخ المباراة",
                value=datetime.date.today(),
                key="fifa_match_date"
            )
        
        with col2:
            scout_name = st.text_input(
                "Scout Name" if st.session_state.language == 'en' else "اسم الكشاف",
                placeholder="Your name" if st.session_state.language == 'en' else "اسمك",
                key="fifa_scout_name"
            )
        
        with col3:
            match_phase = st.selectbox(
                "Phase/Round" if st.session_state.language == 'en' else "الجولة/المرحلة",
                ["", "Group Stage", "Round of 16", "Quarter Finals", "Semi Finals", "Final"],
                key="fifa_match_phase"
            )
        
        # Team Selection Section - Navy Header
        st.markdown("""
            <div class="alnassr-header">
                ⚽ Team Selection
            </div>
        """ if st.session_state.language == 'en' else """
            <div class="alnassr-header">
                ⚽ اختيار الفرق
            </div>
        """, unsafe_allow_html=True)
        
        col_home, col_away = st.columns(2)
        
        with col_home:
            home_team = st.selectbox(
                "🏠 Home Team" if st.session_state.language == 'en' else "🏠 الفريق المحلي",
                [""] + COUNTRIES_U20,
                key="fifa_home_team"
            )
        
        with col_away:
            away_team = st.selectbox(
                "✈️ Away Team" if st.session_state.language == 'en' else "✈️ الفريق الزائر",
                [""] + COUNTRIES_U20,
                key="fifa_away_team"
            )
        
        # Info message if teams not selected
        if not home_team or not away_team:
            st.markdown("""
                <div style="background: #f0f7ff; border-left: 4px solid #FFC60A; padding: 15px; border-radius: 6px; margin-top: 15px;">
                    <p style="margin: 0; color: #1a2332; font-size: 14px;">
                        <strong>ℹ️ Info:</strong> Please select both home and away teams to configure lineups
                    </p>
                </div>
            """ if st.session_state.language == 'en' else """
                <div style="background: #f0f7ff; border-left: 4px solid #FFC60A; padding: 15px; border-radius: 6px; margin-top: 15px;">
                    <p style="margin: 0; color: #1a2332; font-size: 14px;">
                        <strong>ℹ️ معلومة:</strong> يرجى اختيار الفريق المحلي والزائر لإعداد التشكيلات
                    </p>
                </div>
            """, unsafe_allow_html=True)
        
        # Load player data if both teams are selected
        if home_team and away_team:
            try:
                df_players = pd.read_excel('WorldCupU20playersdata.xlsx')
                
                # Detectar nombres de columnas
                country_col_match = None
                for col in df_players.columns:
                    if col.lower() in ['país', 'pais', 'country', 'equipo', 'team']:
                        country_col_match = col
                        break
                
                position_col_match = None
                for col in df_players.columns:
                    if 'position' in col.lower() or 'posición' in col.lower():
                        position_col_match = col
                        break
                
                name_col_match = None
                for col in df_players.columns:
                    if col.lower() in ['nombre', 'name', 'player']:
                        name_col_match = col
                        break
                
                # Position order for sorting
                POSITION_ORDER = {'GK': 1, 'RB': 2, 'CB': 3, 'LB': 4, 'DM': 5, 'CM': 6, 'CAM': 7, 'RW': 8, 'LW': 9, 'ST': 10}
                
                # Get players for home team sorted by position
                home_team_df = df_players[df_players[country_col_match] == home_team].copy()
                home_team_df['pos_order'] = home_team_df[position_col_match].map(POSITION_ORDER).fillna(99)
                home_team_df = home_team_df.sort_values('pos_order')
                home_players = home_team_df[name_col_match].tolist()
                
                # Get players for away team sorted by position
                away_team_df = df_players[df_players[country_col_match] == away_team].copy()
                away_team_df['pos_order'] = away_team_df[position_col_match].map(POSITION_ORDER).fillna(99)
                away_team_df = away_team_df.sort_values('pos_order')
                away_players = away_team_df[name_col_match].tolist()
                
                # Create player-to-position mapping
                home_player_positions = dict(zip(home_team_df[name_col_match], home_team_df[position_col_match]))
                away_player_positions = dict(zip(away_team_df[name_col_match], away_team_df[position_col_match]))
                
                st.markdown("---")
                
                # Initialize session state for players if not exists
                if 'home_match_players' not in st.session_state:
                    st.session_state.home_match_players = []
                if 'away_match_players' not in st.session_state:
                    st.session_state.away_match_players = []
                
                # Players Section
                col_lineup_home, col_lineup_away = st.columns(2)
                
                # HOME TEAM PLAYERS
                with col_lineup_home:
                    # Get home team flag/logo
                    home_flag_file = COUNTRY_FLAGS.get(home_team, None)
                    home_flag_html = ''
                    
                    if home_flag_file:
                        try:
                            from PIL import Image
                            import io
                            
                            img = Image.open(home_flag_file)
                            img.thumbnail((40, 40))  # Resize to small
                            buffered = io.BytesIO()
                            
                            if img.mode in ('RGBA', 'LA', 'P'):
                                img.save(buffered, format="PNG")
                                img_type = "png"
                            else:
                                img = img.convert('RGB')
                                img.save(buffered, format="PNG")
                                img_type = "png"
                            
                            img_str = base64.b64encode(buffered.getvalue()).decode()
                            home_flag_html = f'<img src="data:image/{img_type};base64,{img_str}" style="height:30px; vertical-align:middle; margin-right:10px;">'
                        except:
                            home_flag_html = '🏠 '
                    else:
                        home_flag_html = '🏠 '
                    
                    st.markdown(f"### {home_flag_html}{home_team}", unsafe_allow_html=True)
                    
                    # Button to add player
                    if st.button("➕ Add Player" if st.session_state.language == 'en' else "➕ إضافة لاعب", key="add_home_player"):
                        st.session_state.home_match_players.append({
                            'id': len(st.session_state.home_match_players),
                            'name': '',
                            'number': 1,
                            'position': '',
                            'starter': 'Sí',
                            'minutes': 90,
                            'performance': 1,
                            'potential': 1,
                            'report': ''
                        })
                        st.rerun()
                    
                    st.markdown("---")
                    
                    # Display added players
                    for idx, player_data in enumerate(st.session_state.home_match_players):
                        # Get current name for display
                        current_name = player_data.get('name', 'Not selected')
                        display_name = current_name if current_name else 'Not selected'
                        
                        with st.expander(f"👤 Player {idx+1}: {display_name}", expanded=True):
                            col1, col2, col3 = st.columns([3, 1, 2])
                            
                            with col1:
                                selected_player = st.selectbox(
                                    "Player Name",
                                    [""] + home_players,
                                    index=0 if not player_data['name'] else (home_players.index(player_data['name']) + 1 if player_data['name'] in home_players else 0),
                                    key=f"home_p_name_{idx}"
                                )
                                if selected_player != player_data.get('name', ''):
                                    st.session_state.home_match_players[idx]['name'] = selected_player
                                    
                                    # Auto-completar datos del jugador cuando se selecciona
                                    if selected_player:
                                        # 1. Buscar número anterior en reportes previos
                                        try:
                                            prev_reports = pd.read_excel('fifa_u20_player_reports.xlsx')
                                            player_prev_reports = prev_reports[prev_reports['Player Name'] == selected_player]
                                            if not player_prev_reports.empty:
                                                # Obtener el número más reciente
                                                last_number = player_prev_reports.iloc[-1]['Number']
                                                if pd.notna(last_number):
                                                    st.session_state.home_match_players[idx]['number'] = int(last_number)
                                        except:
                                            pass
                                        
                                        # 2. Buscar año de nacimiento en la base de datos
                                        try:
                                            df_players_db = pd.read_excel('WorldCupU20playersdata.xlsx')
                                            player_data_db = df_players_db[df_players_db['Nombre'].str.strip().str.lower() == selected_player.strip().lower()]
                                            if not player_data_db.empty:
                                                year_value = player_data_db.iloc[0].get('Año', '')
                                                if pd.notna(year_value):
                                                    birth_year_val = str(year_value)[:4]  # Tomar solo los primeros 4 dígitos
                                                    try:
                                                        st.session_state.home_match_players[idx]['birth_year'] = int(birth_year_val)
                                                    except:
                                                        pass
                                        except:
                                            pass
                                    
                                    st.rerun()
                            
                            with col2:
                                player_number = st.number_input(
                                    "#",
                                    min_value=1,
                                    max_value=99,
                                    value=player_data.get('number', 1),
                                    key=f"home_p_num_{idx}"
                                )
                                st.session_state.home_match_players[idx]['number'] = player_number
                            
                            with col3:
                                # Auto-fill position
                                default_pos = ""
                                position_index = 0
                                if selected_player:
                                    default_pos = home_player_positions.get(selected_player, "")
                                    positions_list = ["", "GK", "RB", "CB", "LB", "DM", "CM", "CAM", "RW", "LW", "ST"]
                                    if default_pos in positions_list:
                                        position_index = positions_list.index(default_pos)
                                    elif player_data.get('position') in positions_list:
                                        position_index = positions_list.index(player_data['position'])
                                
                                player_position = st.selectbox(
                                    "Position",
                                    ["", "GK", "RB", "CB", "LB", "DM", "CM", "CAM", "RW", "LW", "ST"],
                                    index=position_index,
                                    key=f"home_p_pos_{idx}"
                                )
                                st.session_state.home_match_players[idx]['position'] = player_position
                            
                            # Birth Year
                            birth_year = st.number_input(
                                "🎂 AÑO (Birth Year)",
                                min_value=1990,
                                max_value=2015,
                                value=player_data.get('birth_year', 2005),
                                step=1,
                                key=f"home_p_year_{idx}"
                            )
                            st.session_state.home_match_players[idx]['birth_year'] = birth_year
                            
                            # Titular and Minutes
                            col_starter, col_minutes = st.columns(2)
                            
                            with col_starter:
                                starter = st.selectbox(
                                    "🎽 Titular",
                                    ["Sí", "No"],
                                    index=0 if player_data.get('starter', 'Sí') == 'Sí' else 1,
                                    key=f"home_p_starter_{idx}"
                                )
                                st.session_state.home_match_players[idx]['starter'] = starter
                            
                            with col_minutes:
                                minutes = st.number_input(
                                    "⏱️ Minutes",
                                    min_value=0,
                                    max_value=120,
                                    value=player_data.get('minutes', 90),
                                    key=f"home_p_minutes_{idx}"
                                )
                                st.session_state.home_match_players[idx]['minutes'] = minutes
                            
                            # Performance and Potential
                            col_perf, col_pot = st.columns(2)
                            
                            with col_perf:
                                performance = st.selectbox(
                                    "🎯 Performance (1-6)",
                                    [1, 2, 3, 4, 5, 6],
                                    index=player_data.get('performance', 1) - 1,
                                    key=f"home_p_perf_{idx}"
                                )
                                st.session_state.home_match_players[idx]['performance'] = performance
                            
                            with col_pot:
                                potential = st.selectbox(
                                    "⭐ Potential (1-6)",
                                    [1, 2, 3, 4, 5, 6],
                                    index=player_data.get('potential', 1) - 1,
                                    key=f"home_p_pot_{idx}"
                                )
                                st.session_state.home_match_players[idx]['potential'] = potential
                            
                            # FIRMAR/CONCLUSION
                            conclusion_options = [
                                "A - Firmar (Sign)",
                                "B+ - Seguir para Firmar (Follow to Sign)",
                                "B - Seguir (Follow)"
                            ]
                            default_conclusion_idx = 0
                            if 'conclusion' in player_data and player_data['conclusion'] in conclusion_options:
                                default_conclusion_idx = conclusion_options.index(player_data['conclusion'])
                            
                            conclusion = st.selectbox(
                                "✅ FIRMAR/CONCLUSION",
                                conclusion_options,
                                index=default_conclusion_idx,
                                key=f"home_p_conclusion_{idx}"
                            )
                            st.session_state.home_match_players[idx]['conclusion'] = conclusion
                            
                            # Report text area
                            report_text = st.text_area(
                                "📝 Report",
                                value=player_data.get('report', ''),
                                height=150,
                                key=f"home_p_report_{idx}"
                            )
                            st.session_state.home_match_players[idx]['report'] = report_text
                            
                            # Remove button
                            if st.button("🗑️ Remove", key=f"remove_home_{idx}"):
                                st.session_state.home_match_players.pop(idx)
                                st.rerun()
                
                # AWAY TEAM PLAYERS
                with col_lineup_away:
                    # Get away team flag/logo
                    away_flag_file = COUNTRY_FLAGS.get(away_team, None)
                    away_flag_html = ''
                    
                    if away_flag_file:
                        try:
                            from PIL import Image
                            import io
                            
                            img = Image.open(away_flag_file)
                            img.thumbnail((40, 40))  # Resize to small
                            buffered = io.BytesIO()
                            
                            if img.mode in ('RGBA', 'LA', 'P'):
                                img.save(buffered, format="PNG")
                                img_type = "png"
                            else:
                                img = img.convert('RGB')
                                img.save(buffered, format="PNG")
                                img_type = "png"
                            
                            img_str = base64.b64encode(buffered.getvalue()).decode()
                            away_flag_html = f'<img src="data:image/{img_type};base64,{img_str}" style="height:30px; vertical-align:middle; margin-right:10px;">'
                        except:
                            away_flag_html = '✈️ '
                    else:
                        away_flag_html = '✈️ '
                    
                    st.markdown(f"### {away_flag_html}{away_team}", unsafe_allow_html=True)
                    
                    # Button to add player
                    if st.button("➕ Add Player" if st.session_state.language == 'en' else "➕ إضافة لاعب", key="add_away_player"):
                        st.session_state.away_match_players.append({
                            'id': len(st.session_state.away_match_players),
                            'name': '',
                            'number': 1,
                            'position': '',
                            'starter': 'Sí',
                            'minutes': 90,
                            'performance': 1,
                            'potential': 1,
                            'report': ''
                        })
                        st.rerun()
                    
                    st.markdown("---")
                    
                    # Display added players
                    for idx, player_data in enumerate(st.session_state.away_match_players):
                        # Get current name for display
                        current_name = player_data.get('name', 'Not selected')
                        display_name = current_name if current_name else 'Not selected'
                        
                        with st.expander(f"👤 Player {idx+1}: {display_name}", expanded=True):
                            col1, col2, col3 = st.columns([3, 1, 2])
                            
                            with col1:
                                selected_player = st.selectbox(
                                    "Player Name",
                                    [""] + away_players,
                                    index=0 if not player_data['name'] else (away_players.index(player_data['name']) + 1 if player_data['name'] in away_players else 0),
                                    key=f"away_p_name_{idx}"
                                )
                                if selected_player != player_data.get('name', ''):
                                    st.session_state.away_match_players[idx]['name'] = selected_player
                                    
                                    # Auto-completar datos del jugador cuando se selecciona
                                    if selected_player:
                                        # 1. Buscar número anterior en reportes previos
                                        try:
                                            prev_reports = pd.read_excel('fifa_u20_player_reports.xlsx')
                                            player_prev_reports = prev_reports[prev_reports['Player Name'] == selected_player]
                                            if not player_prev_reports.empty:
                                                # Obtener el número más reciente
                                                last_number = player_prev_reports.iloc[-1]['Number']
                                                if pd.notna(last_number):
                                                    st.session_state.away_match_players[idx]['number'] = int(last_number)
                                        except:
                                            pass
                                        
                                        # 2. Buscar año de nacimiento en la base de datos
                                        try:
                                            df_players_db = pd.read_excel('WorldCupU20playersdata.xlsx')
                                            player_data_db = df_players_db[df_players_db['Nombre'].str.strip().str.lower() == selected_player.strip().lower()]
                                            if not player_data_db.empty:
                                                year_value = player_data_db.iloc[0].get('Año', '')
                                                if pd.notna(year_value):
                                                    birth_year_val = str(year_value)[:4]  # Tomar solo los primeros 4 dígitos
                                                    try:
                                                        st.session_state.away_match_players[idx]['birth_year'] = int(birth_year_val)
                                                    except:
                                                        pass
                                        except:
                                            pass
                                    
                                    st.rerun()
                            
                            with col2:
                                player_number = st.number_input(
                                    "#",
                                    min_value=1,
                                    max_value=99,
                                    value=player_data.get('number', 1),
                                    key=f"away_p_num_{idx}"
                                )
                                st.session_state.away_match_players[idx]['number'] = player_number
                            
                            with col3:
                                # Auto-fill position
                                default_pos = ""
                                position_index = 0
                                if selected_player:
                                    default_pos = away_player_positions.get(selected_player, "")
                                    positions_list = ["", "GK", "RB", "CB", "LB", "DM", "CM", "CAM", "RW", "LW", "ST"]
                                    if default_pos in positions_list:
                                        position_index = positions_list.index(default_pos)
                                    elif player_data.get('position') in positions_list:
                                        position_index = positions_list.index(player_data['position'])
                                
                                player_position = st.selectbox(
                                    "Position",
                                    ["", "GK", "RB", "CB", "LB", "DM", "CM", "CAM", "RW", "LW", "ST"],
                                    index=position_index,
                                    key=f"away_p_pos_{idx}"
                                )
                                st.session_state.away_match_players[idx]['position'] = player_position
                            
                            # Birth Year
                            birth_year_away = st.number_input(
                                "🎂 AÑO (Birth Year)",
                                min_value=1990,
                                max_value=2015,
                                value=player_data.get('birth_year', 2005),
                                step=1,
                                key=f"away_p_year_{idx}"
                            )
                            st.session_state.away_match_players[idx]['birth_year'] = birth_year_away
                            
                            # Titular and Minutes
                            col_starter, col_minutes = st.columns(2)
                            
                            with col_starter:
                                starter = st.selectbox(
                                    "🎽 Titular",
                                    ["Sí", "No"],
                                    index=0 if player_data.get('starter', 'Sí') == 'Sí' else 1,
                                    key=f"away_p_starter_{idx}"
                                )
                                st.session_state.away_match_players[idx]['starter'] = starter
                            
                            with col_minutes:
                                minutes = st.number_input(
                                    "⏱️ Minutes",
                                    min_value=0,
                                    max_value=120,
                                    value=player_data.get('minutes', 90),
                                    key=f"away_p_minutes_{idx}"
                                )
                                st.session_state.away_match_players[idx]['minutes'] = minutes
                            
                            # Performance and Potential
                            col_perf, col_pot = st.columns(2)
                            
                            with col_perf:
                                performance = st.selectbox(
                                    "🎯 Performance (1-6)",
                                    [1, 2, 3, 4, 5, 6],
                                    index=player_data.get('performance', 1) - 1,
                                    key=f"away_p_perf_{idx}"
                                )
                                st.session_state.away_match_players[idx]['performance'] = performance
                            
                            with col_pot:
                                potential = st.selectbox(
                                    "⭐ Potential (1-6)",
                                    [1, 2, 3, 4, 5, 6],
                                    index=player_data.get('potential', 1) - 1,
                                    key=f"away_p_pot_{idx}"
                                )
                                st.session_state.away_match_players[idx]['potential'] = potential
                            
                            # FIRMAR/CONCLUSION
                            conclusion_options_away = [
                                "A - Firmar (Sign)",
                                "B+ - Seguir para Firmar (Follow to Sign)",
                                "B - Seguir (Follow)"
                            ]
                            default_conclusion_idx_away = 0
                            if 'conclusion' in player_data and player_data['conclusion'] in conclusion_options_away:
                                default_conclusion_idx_away = conclusion_options_away.index(player_data['conclusion'])
                            
                            conclusion_away = st.selectbox(
                                "✅ FIRMAR/CONCLUSION",
                                conclusion_options_away,
                                index=default_conclusion_idx_away,
                                key=f"away_p_conclusion_{idx}"
                            )
                            st.session_state.away_match_players[idx]['conclusion'] = conclusion_away
                            
                            # Report text area
                            report_text = st.text_area(
                                "📝 Report",
                                value=player_data.get('report', ''),
                                height=150,
                                key=f"away_p_report_{idx}"
                            )
                            st.session_state.away_match_players[idx]['report'] = report_text
                            
                            # Remove button
                            if st.button("🗑️ Remove", key=f"remove_away_{idx}"):
                                st.session_state.away_match_players.pop(idx)
                                st.rerun()
                
                st.markdown("---")
                
                # Save button
                if st.button("💾 Save Match Report" if st.session_state.language == 'en' else "💾 حفظ تقرير المباراة", type="primary"):
                    if not scout_name or not match_phase:
                        st.error("Please fill all required fields" if st.session_state.language == 'en' else "يرجى ملء جميع الحقول المطلوبة")
                    elif len(st.session_state.home_match_players) == 0 and len(st.session_state.away_match_players) == 0:
                        st.warning("⚠️ Please add at least one player" if st.session_state.language == 'en' else "⚠️ يرجى إضافة لاعب واحد على الأقل")
                    else:
                        # Prepare player reports for saving
                        player_reports_list = []
                        
                        # Add home team players
                        for player in st.session_state.home_match_players:
                            if player['name']:  # Only save if player is selected
                                player_reports_list.append({
                                    'Date': str(match_date),
                                    'Scout': scout_name,
                                    'Phase': match_phase,
                                    'Match': f"{home_team} vs {away_team}",
                                    'Team': home_team,
                                    'Player Name': player['name'],
                                    'Number': player['number'],
                                    'Position': player['position'],
                                    'Birth Year': player.get('birth_year', ''),
                                    'Starter': player['starter'],
                                    'Minutes': player['minutes'],
                                    'Performance': player['performance'],
                                    'Potential': player['potential'],
                                    'Conclusion': player.get('conclusion', 'B - Seguir (Follow)'),
                                    'Report': player['report']
                                })
                        
                        # Add away team players
                        for player in st.session_state.away_match_players:
                            if player['name']:  # Only save if player is selected
                                player_reports_list.append({
                                    'Date': str(match_date),
                                    'Scout': scout_name,
                                    'Phase': match_phase,
                                    'Match': f"{home_team} vs {away_team}",
                                    'Team': away_team,
                                    'Player Name': player['name'],
                                    'Number': player['number'],
                                    'Position': player['position'],
                                    'Birth Year': player.get('birth_year', ''),
                                    'Starter': player['starter'],
                                    'Minutes': player['minutes'],
                                    'Performance': player['performance'],
                                    'Potential': player['potential'],
                                    'Conclusion': player.get('conclusion', 'B - Seguir (Follow)'),
                                    'Report': player['report']
                                })
                        
                        if player_reports_list:
                            # Save directly to local Excel
                            try:
                                df_new_reports = pd.DataFrame(player_reports_list)
                                
                                # Load existing reports and append new ones
                                try:
                                    df_existing = pd.read_excel('fifa_u20_player_reports.xlsx')
                                    df_combined = pd.concat([df_existing, df_new_reports], ignore_index=True)
                                except FileNotFoundError:
                                    df_combined = df_new_reports
                                
                                # Save to Excel
                                df_combined.to_excel('fifa_u20_player_reports.xlsx', index=False)
                                
                                st.success(f"✅ Match report saved! {len(player_reports_list)} player reports added." if st.session_state.language == 'en' else f"✅ تم حفظ تقرير المباراة! تم إضافة {len(player_reports_list)} تقرير لاعب.")
                                st.balloons()
                                
                                # Clear session state
                                st.session_state.home_match_players = []
                                st.session_state.away_match_players = []
                                st.rerun()
                                
                            except Exception as e:
                                st.error(f"❌ Error saving report: {e}")
                        else:
                            st.warning("⚠️ No players selected to save" if st.session_state.language == 'en' else "⚠️ لا يوجد لاعبون محددون للحفظ")
            
            except FileNotFoundError:
                st.error("❌ WorldCupU20playersdata.xlsx file not found" if st.session_state.language == 'en' else "❌ ملف WorldCupU20playersdata.xlsx غير موجود")
            except Exception as e:
                st.error(f"Error: {e}")
        else:
            st.info("👆 Please select both home and away teams to configure lineups" if st.session_state.language == 'en' else "👆 يرجى اختيار الفريق المحلي والزائر لإعداد التشكيلات")
    
    # Tab 2: CREATE INDIVIDUAL REPORT
    with tabs[1]:
        st.markdown("### 👤 Create Individual Player Report")
        
        # Load player database
        try:
            df_players = pd.read_excel('WorldCupU20playersdata.xlsx')
            
            # Detectar nombres de columnas
            country_col_ind = None
            for col in df_players.columns:
                if col.lower() in ['país', 'pais', 'country', 'equipo', 'team']:
                    country_col_ind = col
                    break
            
            position_col_ind = None
            for col in df_players.columns:
                if 'position' in col.lower() or 'posición' in col.lower():
                    position_col_ind = col
                    break
            
            name_col_ind = None
            for col in df_players.columns:
                if col.lower() in ['nombre', 'name', 'player']:
                    name_col_ind = col
                    break
            
            # Team selection
            st.markdown("---")
            teams = sorted(df_players[country_col_ind].unique().tolist())
            selected_team = st.selectbox(
                "🌍 Select Team / Selección",
                [""] + teams,
                key="individual_team_select"
            )
            
            if selected_team:
                # Filter players by team
                team_players = df_players[df_players[country_col_ind] == selected_team].copy()
                
                # Sort by position order (GK first, ST last)
                position_order = {'Portero': 1, 'Lateral Derecho': 2, 'Lateral Izquierdo': 3, 'Defensa Central': 4, 
                                'Pivote': 5, 'Mediocentro': 6, 'Mediocentro Ofensivo': 7, 
                                'Extremo Derecho': 8, 'Extremo Izquierdo': 9, 'Delantero Centro': 10}
                team_players['pos_order'] = team_players[position_col_ind].map(position_order).fillna(99)
                team_players = team_players.sort_values('pos_order')
                
                player_names = team_players[name_col_ind].tolist()
                
                selected_player = st.selectbox(
                    "👤 Select Player / Jugador",
                    [""] + player_names,
                    key="individual_player_select"
                )
                
                if selected_player:
                    # Get player data
                    player_data = team_players[team_players[name_col_ind] == selected_player].iloc[0]
                    
                    st.markdown("---")
                    st.markdown("### 📋 Player Information")
                    
                    # Display auto-filled data
                    col_info1, col_info2, col_info3, col_info4 = st.columns(4)
                    
                    with col_info1:
                        st.markdown(f"**Name:** {player_data.get(name_col_ind, 'N/A')}")
                    with col_info2:
                        st.markdown(f"**Position:** {player_data.get(position_col_ind, 'N/A')}")
                    with col_info3:
                        birth_date = str(player_data.get('Año', 'N/A'))[:10]  # Remove time if present
                        st.markdown(f"**Birth Date:** {birth_date}")
                    with col_info4:
                        contract = str(player_data.get('Fin Contrato', 'N/A'))[:10]  # Remove time
                        st.markdown(f"**Contract:** {contract}")
                    
                    st.markdown("---")
                    st.markdown("### 📝 Additional Information")
                    
                    # Scout name
                    scout_name = st.text_input("👤 Scout Name / Nombre del Scout", placeholder="Your name", key="ind_scout_name")
                    
                    # Agent info
                    col_agent1, col_agent2 = st.columns(2)
                    with col_agent1:
                        agent_name = st.text_input("🤝 Agent Name / Nombre del Agente", key="ind_agent_name")
                    with col_agent2:
                        agent_phone = st.text_input("📞 Agent Phone / Teléfono del Agente", key="ind_agent_phone")
                    
                    # Check if player already has a photo in previous reports
                    existing_photo_path = None
                    try:
                        df_existing_reports = pd.read_excel('fifa_u20_individual_reports.xlsx')
                        player_reports = df_existing_reports[df_existing_reports['Player'] == selected_player]
                        if not player_reports.empty:
                            # Get the most recent report with a photo
                            for idx, report in player_reports.iterrows():
                                if report.get('Photo') and str(report['Photo']) != 'nan' and str(report['Photo']).strip():
                                    existing_photo_path = report['Photo']
                                    break
                    except:
                        pass
                    
                    # Show existing photo if available
                    if existing_photo_path:
                        st.info(f"📸 Este jugador ya tiene una foto guardada. Se usará automáticamente.")
                        try:
                            from PIL import Image
                            existing_img = Image.open(existing_photo_path)
                            st.image(existing_img, width=200, caption=f"Foto actual de {selected_player}")
                        except:
                            st.warning("⚠️ No se pudo cargar la foto guardada. Puedes subir una nueva.")
                            existing_photo_path = None
                    
                    # Photo upload
                    uploaded_photo = st.file_uploader(
                        "📷 Upload New Player Photo / Subir Nueva Foto del Jugador" if existing_photo_path else "📷 Upload Player Photo / Subir Foto del Jugador",
                        type=['jpg', 'jpeg', 'png'],
                        key="ind_photo_upload",
                        help="Sube una nueva foto solo si quieres cambiar la actual" if existing_photo_path else None
                    )
                    
                    if uploaded_photo:
                        st.success("✅ Nueva foto subida correctamente")
                    
                    st.markdown("---")
                    st.markdown("### ⚙️ Technical Evaluation")
                    
                    # Technical ratings
                    col_eval1, col_eval2, col_eval3 = st.columns(3)
                    
                    # Helper functions for colors
                    def get_color_for_value(value):
                        if value <= 2:
                            return "#FF4444"  # Red
                        elif value <= 4:
                            return "#FFA500"  # Orange
                        else:
                            return "#00C851"  # Green
                    
                    def get_color_for_perfil(perfil_text):
                        if "TOP WORLD CLASS" in perfil_text:
                            return "#00C851"  # Green
                        elif "ELITE CHAMPIONS LEAGUE" in perfil_text:
                            return "#4285F4"  # Blue
                        elif "FIRST PRO LEVEL" in perfil_text:
                            return "#8B4513"  # Brown
                        else:
                            return "#FFA500"  # Orange
                    
                    with col_eval1:
                        rendimiento = st.selectbox(
                            "🎯 Rendimiento (Performance)",
                            [1, 2, 3, 4, 5, 6],
                            index=2,
                            key="ind_rendimiento"
                        )
                        rend_color = get_color_for_value(rendimiento)
                        rend_pct = (rendimiento / 6) * 100
                        st.markdown(f'''
                            <div style="background-color: #e0e0e0; border-radius: 10px; height: 20px; overflow: hidden;">
                                <div style="background-color: {rend_color}; height: 100%; width: {rend_pct}%; border-radius: 10px;"></div>
                            </div>
                        ''', unsafe_allow_html=True)
                    
                    with col_eval2:
                        potencial = st.selectbox(
                            "⭐ Potencial (Potential)",
                            [1, 2, 3, 4, 5, 6],
                            index=2,
                            key="ind_potencial"
                        )
                        pot_color = get_color_for_value(potencial)
                        pot_pct = (potencial / 6) * 100
                        st.markdown(f'''
                            <div style="background-color: #e0e0e0; border-radius: 10px; height: 20px; overflow: hidden;">
                                <div style="background-color: {pot_color}; height: 100%; width: {pot_pct}%; border-radius: 10px;"></div>
                            </div>
                        ''', unsafe_allow_html=True)
                    
                    with col_eval3:
                        perfil_options = [
                            "1 - BACKUP SPL - TOP 4",
                            "2 - STARTER SPL - TOP 4",
                            "3 - STAND OUT SPL - TOP4",
                            "4 - FIRST PRO LEVEL",
                            "5 - ELITE CHAMPIONS LEAGUE",
                            "6 - TOP WORLD CLASS"
                        ]
                        perfil_selected = st.selectbox(
                            "📊 Perfil (Profile)",
                            perfil_options,
                            index=2,
                            key="ind_perfil"
                        )
                        # Extract numeric value
                        perfil = int(perfil_selected.split(' - ')[0])
                        perfil_color = get_color_for_perfil(perfil_selected)
                        perfil_pct = (perfil / 6) * 100
                        st.markdown(f'''
                            <div style="background-color: #e0e0e0; border-radius: 10px; height: 20px; overflow: hidden;">
                                <div style="background-color: {perfil_color}; height: 100%; width: {perfil_pct}%; border-radius: 10px;"></div>
                            </div>
                        ''', unsafe_allow_html=True)
                    
                    # Technical comment
                    comentario_tecnico = st.text_area(
                        "💬 Technical Comment / Comentario Técnico",
                        height=150,
                        placeholder="Enter detailed technical evaluation...",
                        key="ind_comentario"
                    )
                    
                    st.markdown("---")
                    st.markdown("### 📌 Conclusion")
                    
                    conclusion = st.radio(
                        "Final Decision / Decisión Final",
                        ["A - Firmar (Sign)", "B+ - Seguir para Firmar (Follow to Sign)", "B - Seguir (Follow)"],
                        key="ind_conclusion"
                    )
                    
                    st.markdown("---")
                    
                    # Save button
                    if st.button("💾 Save Individual Report", type="primary", use_container_width=True):
                        # Prepare report data
                        report_data = {
                            'Date': datetime.date.today().strftime('%Y-%m-%d'),
                            'Scout': scout_name,
                            'Team': selected_team,
                            'Player': selected_player,
                            'Position': player_data.get('Position principal', 'N/A'),
                            'Birth Date': birth_date,
                            'Contract': contract,
                            'Agent': agent_name,
                            'Agent Phone': agent_phone,
                            'Rendimiento': rendimiento,
                            'Potencial': potencial,
                            'Perfil': perfil_selected,  # Save full description
                            'Technical Comment': comentario_tecnico,
                            'Conclusion': conclusion
                        }
                        
                        # Save photo: use new uploaded photo, or keep existing one
                        if uploaded_photo:
                            # Create directory if it doesn't exist
                            import os
                            os.makedirs('player_photos', exist_ok=True)
                            
                            # Save photo to disk
                            photo_filename = f"player_photos/{selected_player.replace(' ', '_')}.jpg"
                            with open(photo_filename, 'wb') as f:
                                f.write(uploaded_photo.getbuffer())
                            report_data['Photo'] = photo_filename
                        elif existing_photo_path:
                            # Use existing photo from previous report
                            report_data['Photo'] = existing_photo_path
                        else:
                            report_data['Photo'] = ''
                        
                        # Save to Excel
                        try:
                            df_individual = pd.DataFrame([report_data])
                            
                            # Check if file exists
                            try:
                                df_existing = pd.read_excel('fifa_u20_individual_reports.xlsx')
                                df_combined = pd.concat([df_existing, df_individual], ignore_index=True)
                            except FileNotFoundError:
                                df_combined = df_individual
                            
                            df_combined.to_excel('fifa_u20_individual_reports.xlsx', index=False)
                            
                            st.success("✅ Individual report saved successfully!")
                            st.balloons()
                        except Exception as e:
                            st.error(f"Error saving report: {e}")
        
        except FileNotFoundError:
            st.error("❌ WorldCupU20playersdata.xlsx not found. Please add the player database file.")
        except Exception as e:
            st.error(f"Error: {e}")
    
    # Tab 3: INDIVIDUAL REPORTS (View saved reports)
    with tabs[2]:
        # Load Al Nassr logo for title
        try:
            from PIL import Image
            import io
            logo_img = Image.open('alnassr.png')
            logo_img.thumbnail((40, 40))
            buffered = io.BytesIO()
            if logo_img.mode in ('RGBA', 'LA', 'P'):
                logo_img.save(buffered, format="PNG")
            else:
                logo_img = logo_img.convert('RGB')
                logo_img.save(buffered, format="PNG")
            logo_str = base64.b64encode(buffered.getvalue()).decode()
            logo_html = f'<img src="data:image/png;base64,{logo_str}" style="height:32px; vertical-align:middle; margin-right:10px;">'
        except:
            logo_html = '📋'  # Fallback to emoji if logo not found
        
        st.markdown(f"<h2 style='text-align: center; color: #002B5B;'>{logo_html} INDIVIDUAL REPORTS</h2>", unsafe_allow_html=True)
        
        try:
            # Load individual reports
            df_individual_reports = pd.read_excel('fifa_u20_individual_reports.xlsx')
            
            # Check if navigated from player database
            filter_player_name = st.session_state.get('filter_player', None)
            if filter_player_name:
                st.info(f"👤 Showing reports for: **{filter_player_name}**")
                # Clear the filter after showing
                if st.button("❌ Clear Filter"):
                    st.session_state.filter_player = None
                    st.rerun()
            
            if not df_individual_reports.empty:
                # Filters
                st.markdown("---")
                st.markdown("🔍 **Filtros / Filters**")
                col_f1, col_f2, col_f3 = st.columns(3)
                
                with col_f1:
                    scouts_filter = ['All Scouts'] + sorted(df_individual_reports['Scout'].dropna().unique().tolist())
                    selected_scout_filter = st.selectbox("👤 Scout", scouts_filter, key="ind_reports_scout_filter")
                
                with col_f2:
                    # Filtrar equipos según el scout seleccionado
                    if selected_scout_filter != 'All Scouts':
                        available_teams = df_individual_reports[df_individual_reports['Scout'] == selected_scout_filter]['Team'].unique()
                    else:
                        available_teams = df_individual_reports['Team'].unique()
                    teams_filter = ['All Teams'] + sorted(available_teams.tolist())
                    selected_team_filter = st.selectbox("🌍 Team", teams_filter, key="ind_reports_team_filter")
                
                with col_f3:
                    # Filtrar jugadores según scout y equipo seleccionados
                    filtered_for_players = df_individual_reports.copy()
                    if selected_scout_filter != 'All Scouts':
                        filtered_for_players = filtered_for_players[filtered_for_players['Scout'] == selected_scout_filter]
                    if selected_team_filter != 'All Teams':
                        filtered_for_players = filtered_for_players[filtered_for_players['Team'] == selected_team_filter]
                    
                    players_filter = ['All Players'] + sorted(filtered_for_players['Player'].unique().tolist())
                    # Auto-select player if navigated from player database
                    default_player_index = 0
                    if filter_player_name and filter_player_name in players_filter:
                        default_player_index = players_filter.index(filter_player_name)
                    selected_player_filter = st.selectbox("👤 Player", players_filter, index=default_player_index, key="ind_reports_player_filter")
                
                # Apply filters
                filtered_reports = df_individual_reports.copy()
                if selected_scout_filter != 'All Scouts':
                    filtered_reports = filtered_reports[filtered_reports['Scout'] == selected_scout_filter]
                if selected_team_filter != 'All Teams':
                    filtered_reports = filtered_reports[filtered_reports['Team'] == selected_team_filter]
                if selected_player_filter != 'All Players':
                    filtered_reports = filtered_reports[filtered_reports['Player'] == selected_player_filter]
                
                st.markdown("---")
                
                # Download buttons for Individual Reports
                st.markdown("### 📥 Descargar Datos / Download Data")
                create_download_buttons(
                    filtered_reports, 
                    filename_base="fifa_u20_individual_reports",
                    label_prefix="Descargar / Download"
                )
                st.markdown("---")
                
                # === VISUALIZACIÓN DE MEDIAS SI UN JUGADOR TIENE MÚLTIPLES INFORMES ===
                if selected_player_filter != 'All Players':
                    player_reports = filtered_reports[filtered_reports['Player'] == selected_player_filter]
                    unique_scouts = player_reports['Scout'].nunique()
                    
                    if unique_scouts >= 2:
                        st.markdown(f"### 📊 {selected_player_filter} - Media de {unique_scouts} Scouts")
                        st.markdown("")
                        
                        # Calcular medias
                        avg_rendimiento = player_reports['Rendimiento'].mean()
                        avg_potencial = player_reports['Potencial'].mean()
                        
                        # Mostrar medias en 2 columnas con número grande y barra de progreso
                        col_avg1, col_avg2 = st.columns(2)
                        
                        with col_avg1:
                            st.markdown("""
                                <div style="background: white; padding: 30px; border-radius: 15px; box-shadow: 0 4px 20px rgba(0,0,0,0.06); text-align: center;">
                                    <p style="font-size: 14px; color: #999; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 10px;">🎯 RENDIMIENTO - MEDIA</p>
                                    <h1 style="font-size: 72px; font-weight: 700; color: #1B2845; margin: 10px 0; line-height: 1;">{:.1f}</h1>
                                    <p style="font-size: 14px; color: #666; margin-top: 5px;">Sobre 6.0</p>
                                </div>
                            """.format(avg_rendimiento), unsafe_allow_html=True)
                            st.markdown("<br>", unsafe_allow_html=True)
                            progress_rend = avg_rendimiento / 6.0
                            st.progress(progress_rend)
                        
                        with col_avg2:
                            st.markdown("""
                                <div style="background: white; padding: 30px; border-radius: 15px; box-shadow: 0 4px 20px rgba(0,0,0,0.06); text-align: center;">
                                    <p style="font-size: 14px; color: #999; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 10px;">⭐ POTENCIAL - MEDIA</p>
                                    <h1 style="font-size: 72px; font-weight: 700; color: #1B2845; margin: 10px 0; line-height: 1;">{:.1f}</h1>
                                    <p style="font-size: 14px; color: #666; margin-top: 5px;">Sobre 6.0</p>
                                </div>
                            """.format(avg_potencial), unsafe_allow_html=True)
                            st.markdown("<br>", unsafe_allow_html=True)
                            progress_pot = avg_potencial / 6.0
                            st.progress(progress_pot)
                        
                        st.markdown("---")
                        st.markdown(f"### 📝 Detalles de {len(player_reports)} Informes Individuales")
                    else:
                        st.markdown(f"### 📊 {len(filtered_reports)} Individual Report(s)")
                else:
                    st.markdown(f"### 📊 {len(filtered_reports)} Individual Report(s)")
                
                # Display reports
                for idx, report in filtered_reports.iterrows():
                    # Get scout name if available
                    scout_name = report.get('Scout', 'Unknown')
                    with st.expander(
                        f"👤 {report['Player']} ({report['Team']}) | {report['Conclusion']} | 📅 {report['Date']} | 👤 Scout: {scout_name}",
                        expanded=False  # Colapsado por defecto
                    ):
                        # === CSS FINAL LIMPIO Y MINIMALISTA ===
                        st.markdown("""
                            <style>
                            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&display=swap');
                            body { background-color: #fafafa; font-family: 'Inter', sans-serif; }
                            
                            .player-photo-final { 
                                width: 120px; 
                                height: 120px; 
                                border-radius: 50%; 
                                border: 3px solid #FFD700; 
                                object-fit: cover; 
                                margin: 0 auto; 
                                display: block;
                                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                            }
                            .player-name-final { 
                                font-size: 32px; 
                                font-weight: 700; 
                                color: #1B2845; 
                                text-align: center; 
                                margin: 15px 0 8px 0; 
                            }
                            .player-subtitle-final { 
                                font-size: 15px; 
                                color: #6c757d; 
                                text-align: center; 
                                margin-bottom: 0;
                            }
                            
                            .metric-card-final { 
                                background: white; 
                                padding: 25px 20px; 
                                border-radius: 12px; 
                                box-shadow: 0 4px 20px rgba(0,0,0,0.06); 
                                text-align: center;
                                transition: transform 0.3s ease, box-shadow 0.3s ease;
                                height: 100%;
                            }
                            .metric-card-final:hover {
                                transform: translateY(-3px);
                                box-shadow: 0 6px 25px rgba(0,0,0,0.12);
                            }
                            .metric-number-final { 
                                font-size: 48px; 
                                font-weight: 700; 
                                color: #1B2845; 
                                line-height: 1; 
                                margin-bottom: 10px;
                            }
                            .metric-label-final { 
                                font-size: 12px; 
                                color: #999; 
                                text-transform: uppercase; 
                                letter-spacing: 2px; 
                                margin-bottom: 15px;
                                font-weight: 600;
                            }
                            .stars-final { 
                                font-size: 20px; 
                                display: flex; 
                                justify-content: center; 
                                gap: 5px;
                            }
                            .star-gold { color: #FFD700; filter: drop-shadow(0 2px 3px rgba(255,215,0,0.3)); }
                            .star-gray { color: #e9ecef; }
                            
                            .secondary-info-final {
                                background: white;
                                padding: 20px;
                                border-radius: 12px;
                                box-shadow: 0 4px 20px rgba(0,0,0,0.06);
                                display: flex;
                                justify-content: center;
                                gap: 30px;
                                flex-wrap: wrap;
                                margin: 40px 0;
                            }
                            .secondary-info-final span {
                                font-size: 14px;
                                color: #495057;
                            }
                            
                            .tech-comment-final {
                                background: #f8f9fa;
                                padding: 25px;
                                border-radius: 12px;
                                border-left: 4px solid #FFD700;
                                box-shadow: 0 2px 15px rgba(0,0,0,0.05);
                                height: 100%;
                                display: flex;
                                align-items: center;
                            }
                            .tech-comment-final p {
                                font-style: italic;
                                color: #495057;
                                font-size: 16px;
                                margin: 0;
                                line-height: 1.6;
                            }
                            
                            .conclusion-final {
                                background: #f8f9fa;
                                padding: 25px 35px;
                                border-radius: 12px;
                                box-shadow: 0 2px 15px rgba(0,0,0,0.05);
                                height: 100%;
                                display: flex;
                                align-items: center;
                                justify-content: center;
                                text-align: center;
                            }
                            .conclusion-final p {
                                font-size: 28px;
                                font-weight: 700;
                                text-transform: uppercase;
                                margin: 0;
                                line-height: 1.3;
                            }
                            .conclusion-green { color: #28a745; }
                            .conclusion-blue { color: #1B2845; }
                            .conclusion-orange { color: #ff8c00; }
                            
                            @keyframes fadeInUp {
                                from {
                                    opacity: 0;
                                    transform: translateY(20px);
                                }
                                to {
                                    opacity: 1;
                                    transform: translateY(0);
                                }
                            }
                            .animate-fade-in {
                                animation: fadeInUp 0.6s ease-out;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        
                        # === HEADER SECTION (2 COLUMNAS) ===
                        col_left, col_right = st.columns([250, 1000], gap="large")
                        
                        # COLUMNA IZQUIERDA: Foto + Nombre + Subtítulo
                        with col_left:
                            # Foto circular
                            if report.get('Photo') and report['Photo']:
                                try:
                                    from PIL import Image
                                    import io
                                    img = Image.open(report['Photo'])
                                    # Convertir a base64 para aplicar clase CSS
                                    buffered = io.BytesIO()
                                    if img.mode in ('RGBA', 'LA', 'P'):
                                        img.save(buffered, format="PNG")
                                    else:
                                        img = img.convert('RGB')
                                        img.save(buffered, format="PNG")
                                    img_str = base64.b64encode(buffered.getvalue()).decode()
                                    st.markdown(f'<img src="data:image/png;base64,{img_str}" class="player-photo-final">', unsafe_allow_html=True)
                                except:
                                    # Usar logo Al Nassr como placeholder
                                    try:
                                        from PIL import Image
                                        import io
                                        logo_img = Image.open('alnassr.png')
                                        buffered = io.BytesIO()
                                        if logo_img.mode in ('RGBA', 'LA', 'P'):
                                            logo_img.save(buffered, format="PNG")
                                        else:
                                            logo_img = logo_img.convert('RGB')
                                            logo_img.save(buffered, format="PNG")
                                        logo_str = base64.b64encode(buffered.getvalue()).decode()
                                        st.markdown(f'<img src="data:image/png;base64,{logo_str}" class="player-photo-final">', unsafe_allow_html=True)
                                    except:
                                        st.markdown('<div style="width:120px; height:120px; border-radius:50%; border:3px solid #FFD700; display:flex; align-items:center; justify-content:center; margin:0 auto; background:#f0f0f0; font-size:50px;">👤</div>', unsafe_allow_html=True)
                            else:
                                # Usar logo Al Nassr como placeholder
                                try:
                                    from PIL import Image
                                    import io
                                    logo_img = Image.open('alnassr.png')
                                    buffered = io.BytesIO()
                                    if logo_img.mode in ('RGBA', 'LA', 'P'):
                                        logo_img.save(buffered, format="PNG")
                                    else:
                                        logo_img = logo_img.convert('RGB')
                                        logo_img.save(buffered, format="PNG")
                                    logo_str = base64.b64encode(buffered.getvalue()).decode()
                                    st.markdown(f'<img src="data:image/png;base64,{logo_str}" class="player-photo-final">', unsafe_allow_html=True)
                                except:
                                    st.markdown('<div style="width:120px; height:120px; border-radius:50%; border:3px solid #FFD700; display:flex; align-items:center; justify-content:center; margin:0 auto; background:#f0f0f0; font-size:50px;">👤</div>', unsafe_allow_html=True)
                            
                            # Nombre
                            st.markdown(f'<p class="player-name-final">{report["Player"]}</p>', unsafe_allow_html=True)
                            
                            # Subtítulo con posición, año, país y bandera
                            birth_year = str(report['Birth Date'])[:4] if report.get('Birth Date') else 'N/A'
                            position = report.get('Position', 'N/A')
                            team = report['Team']
                            
                            # Get country flag
                            flag_html = ''
                            flag_file = COUNTRY_FLAGS.get(team, None)
                            if flag_file:
                                try:
                                    from PIL import Image
                                    import io
                                    flag_img = Image.open(flag_file)
                                    flag_img.thumbnail((20, 20))
                                    buffered = io.BytesIO()
                                    if flag_img.mode in ('RGBA', 'LA', 'P'):
                                        flag_img.save(buffered, format="PNG")
                                    else:
                                        flag_img = flag_img.convert('RGB')
                                        flag_img.save(buffered, format="PNG")
                                    flag_str = base64.b64encode(buffered.getvalue()).decode()
                                    flag_html = f' <img src="data:image/png;base64,{flag_str}" style="height:16px; vertical-align:middle;">'
                                except:
                                    pass
                            
                            st.markdown(f'<p class="player-subtitle-final">{position} • {birth_year} • {team}{flag_html}</p>', unsafe_allow_html=True)
                        
                        # COLUMNA DERECHA: Grid de 3 métricas
                        with col_right:
                            col_m1, col_m2, col_m3 = st.columns(3)
                            
                            # Métrica 1: Rendimiento
                            with col_m1:
                                rendimiento = int(report['Rendimiento'])
                                stars_html = ''.join([f'<span class="star-gold">⭐</span>' for _ in range(rendimiento)]) + ''.join([f'<span class="star-gray">☆</span>' for _ in range(6 - rendimiento)])
                                st.markdown(f"""
                                    <div class="metric-card-final animate-fade-in">
                                        <div class="metric-number-final">{rendimiento}</div>
                                        <div class="metric-label-final">Rendimiento</div>
                                        <div class="stars-final">{stars_html}</div>
                                    </div>
                                """, unsafe_allow_html=True)
                            
                            # Métrica 2: Potencial
                            with col_m2:
                                potencial = int(report['Potencial'])
                                stars_html = ''.join([f'<span class="star-gold">⭐</span>' for _ in range(potencial)]) + ''.join([f'<span class="star-gray">☆</span>' for _ in range(6 - potencial)])
                                st.markdown(f"""
                                    <div class="metric-card-final animate-fade-in">
                                        <div class="metric-number-final">{potencial}</div>
                                        <div class="metric-label-final">Potencial</div>
                                        <div class="stars-final">{stars_html}</div>
                                    </div>
                                """, unsafe_allow_html=True)
                            
                            # Métrica 3: Perfil
                            with col_m3:
                                perfil_val = report['Perfil']
                                if isinstance(perfil_val, str) and '-' in perfil_val:
                                    perfil_num = int(perfil_val.split(' - ')[0])
                                else:
                                    perfil_num = int(perfil_val)
                                stars_html = ''.join([f'<span class="star-gold">⭐</span>' for _ in range(perfil_num)]) + ''.join([f'<span class="star-gray">☆</span>' for _ in range(6 - perfil_num)])
                                st.markdown(f"""
                                    <div class="metric-card-final animate-fade-in">
                                        <div class="metric-number-final">{perfil_num}</div>
                                        <div class="metric-label-final">Perfil</div>
                                        <div class="stars-final">{stars_html}</div>
                                    </div>
                                """, unsafe_allow_html=True)
                        
                        # === INFORMACIÓN SECUNDARIA ===
                        contract = report.get('Contract', 'N/A')
                        agent = report.get('Agent', 'N/A')
                        phone = report.get('Agent Phone', 'N/A')
                        report_date = report.get('Date', 'N/A')
                        scout = report.get('Scout', 'N/A')
                        
                        st.markdown(f"""
                            <div class="secondary-info-final animate-fade-in">
                                <span>📄 {contract}</span>
                                <span>🤝 {agent}</span>
                                <span>📞 {phone}</span>
                                <span>📅 {report_date}</span>
                                <span>👤 {scout}</span>
                            </div>
                        """, unsafe_allow_html=True)
                        
                        # === SECCIÓN INFERIOR: COMENTARIO TÉCNICO Y CONCLUSIÓN (2 COLUMNAS) ===
                        col_comment, col_conclusion = st.columns(2)
                        
                        # COLUMNA IZQUIERDA: Comentario técnico
                        with col_comment:
                            tech_comment = report.get('Technical Comment', None)
                            if tech_comment and str(tech_comment) != 'nan' and str(tech_comment).strip():
                                st.markdown(f"""
                                    <div class="tech-comment-final animate-fade-in">
                                        <p>{tech_comment}</p>
                                    </div>
                                """, unsafe_allow_html=True)
                            else:
                                st.markdown("""
                                    <div class="tech-comment-final animate-fade-in">
                                        <p style="color: #999;">No technical comment available.</p>
                                    </div>
                                """, unsafe_allow_html=True)
                        
                        # COLUMNA DERECHA: Conclusión
                        with col_conclusion:
                            conclusion = report.get('Conclusion', '')
                            
                            # Determinar color del texto según conclusión
                            if conclusion.startswith('A'):
                                conclusion_class = 'conclusion-green'
                            elif conclusion.startswith('B+'):
                                conclusion_class = 'conclusion-blue'
                            elif conclusion.startswith('B') and not conclusion.startswith('B+'):
                                conclusion_class = 'conclusion-orange'
                            else:
                                conclusion_class = 'conclusion-blue'  # Default
                            
                            st.markdown(f"""
                                <div class="conclusion-final animate-fade-in">
                                    <p class="{conclusion_class}">{conclusion}</p>
                                </div>
                            """, unsafe_allow_html=True)
            else:
                st.info("🔍 No individual reports found. Create one in the CREATE INDIVIDUAL REPORT tab.")
        
        except FileNotFoundError:
            st.info("📊 No individual reports yet. Create your first report!")
        except Exception as e:
            st.error(f"Error loading reports: {e}")
    
    # Tab 4: SCHEDULE - MATCHES
    with tabs[3]:
        st.markdown("<h2 style='text-align: center; color: #002B5B;'>📅 SCHEDULE & MATCHES</h2>", unsafe_allow_html=True)
        
        try:
            import json
            with open('datos_liga_crudos.json', 'r', encoding='utf-8') as f:
                liga_data = json.load(f)
            
            # Tabs for different views
            schedule_tabs = st.tabs(["📅 Matches", "🏆 Table"])
            
            # Tab 1: Matches
            with schedule_tabs[0]:
                st.markdown("### ⚽ Matches Calendar")
                
                # Get all matches
                all_matches = liga_data.get('matches', {}).get('allMatches', [])
                
                if all_matches:
                    # Group filter
                    groups = sorted(set([m.get('group', 'N/A') for m in all_matches if m.get('group')]))
                    selected_group = st.selectbox("🎯 Group", ['All Groups'] + groups)
                    
                    # Filter matches
                    filtered_matches = all_matches if selected_group == 'All Groups' else [m for m in all_matches if m.get('group') == selected_group]
                    
                    st.markdown("---")
                    
                    # Display matches
                    for match in filtered_matches:
                        home = match.get('home', {})
                        away = match.get('away', {})
                        status = match.get('status', {})
                        
                        col1, col2, col3 = st.columns([2, 1, 2])
                        
                        with col1:
                            st.markdown(f"**{home.get('name', 'TBD')}**")
                        with col2:
                            if status.get('finished'):
                                st.markdown(f"<div style='text-align:center; font-weight:bold; color:#002B5B;'>{status.get('scoreStr', 'vs')}</div>", unsafe_allow_html=True)
                            else:
                                st.markdown(f"<div style='text-align:center;'>vs</div>", unsafe_allow_html=True)
                        with col3:
                            st.markdown(f"**{away.get('name', 'TBD')}**")
                        
                        # Match details
                        col_d1, col_d2, col_d3 = st.columns(3)
                        with col_d1:
                            st.markdown(f"🏆 Group {match.get('group', 'N/A')}")
                        with col_d2:
                            st.markdown(f"📅 {status.get('utcTime', '')[:10]}")
                        with col_d3:
                            st.markdown(f"⏰ {status.get('reason', {}).get('short', 'Scheduled')}")
                        
                        st.markdown("<hr style='margin: 1rem 0; border-top: 1px solid #eee;'>", unsafe_allow_html=True)
                else:
                    st.info("🔍 No matches found")
            
            # Tab 2: Table
            with schedule_tabs[1]:
                st.markdown("### 🏆 Group Tables")
                
                tables = liga_data.get('table', [])
                
                for table_data in tables:
                    if table_data.get('data', {}).get('tables'):
                        for group_table in table_data['data']['tables']:
                            group_name = group_table.get('leagueName', 'Group')
                            st.markdown(f"#### {group_name}")
                            
                            table_all = group_table.get('table', {}).get('all', [])
                            
                            if table_all:
                                df_table = pd.DataFrame(table_all)
                                display_cols = ['name', 'played', 'wins', 'draws', 'losses', 'scoresStr', 'pts']
                                available_cols = [col for col in display_cols if col in df_table.columns]
                                
                                st.dataframe(
                                    df_table[available_cols].reset_index(drop=True),
                                    use_container_width=True,
                                    hide_index=True
                                )
                            
                            st.markdown("---")
        
        except FileNotFoundError:
            st.error("❌ datos_liga_crudos.json not found")
        except Exception as e:
            st.error(f"Error loading schedule: {e}")
    
    # Tab 5: PLAYER DATABASE
    with tabs[4]:
        # Load Al Nassr logo for title
        try:
            from PIL import Image
            import io
            logo_title = Image.open('alnassr.png')
            logo_title.thumbnail((30, 30))
            buffered_title = io.BytesIO()
            if logo_title.mode in ('RGBA', 'LA', 'P'):
                logo_title.save(buffered_title, format="PNG")
            else:
                logo_title = logo_title.convert('RGB')
                logo_title.save(buffered_title, format="PNG")
            logo_title_str = base64.b64encode(buffered_title.getvalue()).decode()
            logo_title_html = f'<img src="data:image/png;base64,{logo_title_str}" style="height:24px; vertical-align:middle; margin-right:8px;">'
        except:
            logo_title_html = '🗄️'
        
        if st.session_state.language == 'en':
            st.markdown(f"<h3>{logo_title_html} Player Database - FIFA U20 World Cup</h3>", unsafe_allow_html=True)
        else:
            st.markdown(f"<h3>{logo_title_html} قاعدة بيانات اللاعبين - كأس العالم تحت 20</h3>", unsafe_allow_html=True)
        
        # Custom CSS for player cards
        st.markdown("""
        <style>
            .player-card {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 1rem;
                border-radius: 10px;
                color: white;
                margin-bottom: 0.5rem;
            }
            .player-card-header {
                font-size: 1.2rem;
                font-weight: bold;
            }
            .player-badge {
                display: inline-block;
                background-color: rgba(255,255,255,0.2);
                padding: 0.3rem 0.6rem;
                border-radius: 15px;
                margin-right: 0.5rem;
                font-size: 0.9rem;
            }
        </style>
        """, unsafe_allow_html=True)
        
        # Load player data
        try:
            df_players = pd.read_excel('WorldCupU20playersdata.xlsx')
            
            # Detectar nombres de columnas automáticamente
            # Buscar columna de país/equipo
            country_col = None
            for col in df_players.columns:
                if col.lower() in ['país', 'pais', 'paíós', 'country', 'equipo', 'team']:
                    country_col = col
                    break
            
            # Buscar columna de posición
            position_col = None
            for col in df_players.columns:
                if 'position' in col.lower() or 'posición' in col.lower():
                    position_col = col
                    break
            
            # Buscar columna de nombre
            name_col = None
            for col in df_players.columns:
                if col.lower() in ['nombre', 'name', 'player']:
                    name_col = col
                    break
            
            # Verificar que existan las columnas necesarias
            if not country_col or not position_col or not name_col:
                st.error("⚠️ No se encontraron las columnas necesarias en el Excel. Asegúrate de que existan: País/Country, Position, Nombre/Name")
                raise ValueError("Columnas necesarias no encontradas")
            
            # Load match reports from local Excel
            try:
                df_reports = pd.read_excel('fifa_u20_player_reports.xlsx')
            except FileNotFoundError:
                df_reports = pd.DataFrame()
            except Exception as e:
                df_reports = pd.DataFrame()
            
            # Try to load individual reports
            try:
                df_individual_reports = pd.read_excel('fifa_u20_individual_reports.xlsx')
            except FileNotFoundError:
                df_individual_reports = pd.DataFrame()
            except Exception as e:
                df_individual_reports = pd.DataFrame()
            
            # Control para mostrar todos los jugadores
            if 'show_all_players' not in st.session_state:
                st.session_state.show_all_players = False
            
            # Botón para mostrar todos
            col_btn1, col_btn2 = st.columns([1, 5])
            with col_btn1:
                if st.button("👥 Mostrar todos" if not st.session_state.show_all_players else "❌ Ocultar", key="toggle_show_all_players"):
                    st.session_state.show_all_players = not st.session_state.show_all_players
                    st.rerun()
            
            # Filters
            st.markdown("---")
            col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)
            
            with col_filter1:
                # Search by player name
                search_player = st.text_input(
                    "🔍 Search Player" if st.session_state.language == 'en' else "🔍 ابحث عن لاعب",
                    placeholder="Enter player name..." if st.session_state.language == 'en' else "أدخل اسم اللاعب...",
                    key="fifa_u20_search_player"
                )
            
            with col_filter2:
                # Position filter
                all_positions = ['All Positions'] + sorted(df_players[position_col].dropna().unique().tolist())
                selected_position = st.selectbox(
                    "⚽ Position" if st.session_state.language == 'en' else "⚽ المركز",
                    all_positions,
                    key="fifa_u20_position_filter"
                )
            
            with col_filter3:
                # Country filter
                all_countries = ['All Teams'] + sorted(df_players[country_col].unique().tolist())
                selected_country = st.selectbox(
                    "🌍 Team" if st.session_state.language == 'en' else "🌍 الفريق",
                    all_countries,
                    key="fifa_u20_country_filter"
                )
            
            with col_filter4:
                # Conclusion filter
                conclusion_options = ['Todas', 'A - Firmar', 'B+ - Seguir para Firmar', 'B - Seguir']
                selected_conclusion = st.selectbox(
                    "🎯 Conclusión",
                    conclusion_options,
                    key="fifa_u20_conclusion_filter"
                )
            
            # Checkbox para mostrar solo jugadores con informe
            col_check1, col_check2 = st.columns([2, 4])
            with col_check1:
                only_with_reports = st.checkbox(
                    "📊 Mostrar solo jugadores con informe",
                    value=False,
                    key="only_with_reports_filter"
                )
            
            # Filter dataframe
            filtered_df = df_players.copy()
            
            if selected_country != 'All Teams':
                filtered_df = filtered_df[filtered_df[country_col] == selected_country]
            
            if selected_position != 'All Positions':
                filtered_df = filtered_df[filtered_df[position_col] == selected_position]
            
            if search_player:
                filtered_df = filtered_df[filtered_df[name_col].str.contains(search_player, case=False, na=False)]
            
            # Filtrar por jugadores con informe
            if only_with_reports:
                players_with_any_report = []
                
                # Añadir jugadores con match reports
                if not df_reports.empty:
                    players_with_any_report.extend(df_reports['Player Name'].unique().tolist())
                
                # Añadir jugadores con individual reports
                if not df_individual_reports.empty:
                    players_with_any_report.extend(df_individual_reports['Player'].unique().tolist())
                
                # Filtrar solo jugadores con informe
                if players_with_any_report:
                    filtered_df = filtered_df[filtered_df[name_col].isin(players_with_any_report)]
                else:
                    filtered_df = pd.DataFrame()  # No hay jugadores con informe
            
            # Filtrar por conclusión
            if selected_conclusion != 'Todas':
                players_with_conclusion = []
                
                # Buscar en match reports
                if not df_reports.empty:
                    if selected_conclusion == 'A - Firmar':
                        conclusion_reports = df_reports[df_reports['Conclusion'].str.contains('A -|A-|A ', case=False, na=False, regex=True)]
                    elif selected_conclusion == 'B+ - Seguir para Firmar':
                        conclusion_reports = df_reports[df_reports['Conclusion'].str.contains(r'B\+', case=False, na=False, regex=True)]
                    elif selected_conclusion == 'B - Seguir':
                        conclusion_reports = df_reports[
                            (df_reports['Conclusion'].str.contains('B -|B-|B ', case=False, na=False, regex=True)) &
                            (~df_reports['Conclusion'].str.contains(r'B\+', case=False, na=False, regex=True))
                        ]
                    else:
                        conclusion_reports = pd.DataFrame()
                    
                    if not conclusion_reports.empty:
                        players_with_conclusion.extend(conclusion_reports['Player Name'].unique().tolist())
                
                # Buscar en individual reports
                if not df_individual_reports.empty:
                    if selected_conclusion == 'A - Firmar':
                        conclusion_ind_reports = df_individual_reports[df_individual_reports['Conclusion'].str.contains('A -|A-|A ', case=False, na=False, regex=True)]
                    elif selected_conclusion == 'B+ - Seguir para Firmar':
                        conclusion_ind_reports = df_individual_reports[df_individual_reports['Conclusion'].str.contains(r'B\+', case=False, na=False, regex=True)]
                    elif selected_conclusion == 'B - Seguir':
                        conclusion_ind_reports = df_individual_reports[
                            (df_individual_reports['Conclusion'].str.contains('B -|B-|B ', case=False, na=False, regex=True)) &
                            (~df_individual_reports['Conclusion'].str.contains(r'B\+', case=False, na=False, regex=True))
                        ]
                    else:
                        conclusion_ind_reports = pd.DataFrame()
                    
                    if not conclusion_ind_reports.empty:
                        players_with_conclusion.extend(conclusion_ind_reports['Player'].unique().tolist())
                
                # Filtrar solo jugadores con esa conclusión
                if players_with_conclusion:
                    filtered_df = filtered_df[filtered_df[name_col].isin(players_with_conclusion)]
                else:
                    filtered_df = pd.DataFrame()  # No hay jugadores con esa conclusión
            
            # Si no se ha activado "Mostrar todos" y no hay filtros activos, no mostrar nada
            has_active_filters = (
                search_player or 
                selected_position != 'All Positions' or 
                selected_country != 'All Teams' or
                only_with_reports or
                selected_conclusion != 'Todas'
            )
            
            if not st.session_state.show_all_players and not has_active_filters:
                filtered_df = pd.DataFrame()  # No mostrar nada
            
            st.markdown("---")
            
            # Display stats
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            with col_stat1:
                st.metric("👥 Total Players", len(filtered_df))
            with col_stat2:
                st.metric("🌍 Teams", filtered_df[country_col].nunique() if not filtered_df.empty else 0)
            with col_stat3:
                # Count players with match reports
                if not df_reports.empty and not filtered_df.empty:
                    players_with_match_reports = filtered_df[name_col].isin(df_reports['Player Name']).sum()
                    st.metric("⚽ Match Reports", players_with_match_reports)
                else:
                    st.metric("⚽ Match Reports", 0)
            with col_stat4:
                # Count players with individual reports
                if not df_individual_reports.empty and not filtered_df.empty:
                    players_with_individual_reports = filtered_df[name_col].isin(df_individual_reports['Player']).sum()
                    st.metric("📋 Individual Reports", players_with_individual_reports)
                else:
                    st.metric("📋 Individual Reports", 0)
            
            st.markdown("---")
            
            # Download buttons for Player Database
            if not filtered_df.empty:
                st.markdown("### 📥 Descargar Datos / Download Data")
                create_download_buttons(
                    filtered_df, 
                    filename_base="fifa_u20_player_database",
                    label_prefix="Descargar / Download"
                )
                st.markdown("---")
            
            if filtered_df.empty:
                if not st.session_state.show_all_players and not has_active_filters:
                    st.info("👥 Haz clic en 'Mostrar todos' o usa los filtros para ver los jugadores" if st.session_state.language == 'en' else "👥 انقر على 'إظهار الكل' أو استخدم الفلاتر لمشاهدة اللاعبين")
                else:
                    st.warning("❌ No se encontraron jugadores con los filtros seleccionados" if st.session_state.language == 'en' else "❌ لا يوجد لاعبون بالفلاتر المحددة")
            else:
                # Group by country/team
                for country in sorted(filtered_df[country_col].unique()):
                    country_players = filtered_df[filtered_df[country_col] == country]
                    
                    # Country header with flag/logo image using PIL
                    flag_file = COUNTRY_FLAGS.get(country, None)
                    flag_html = ''
                    
                    if flag_file:
                        try:
                            # Use PIL to open and encode image
                            from PIL import Image
                            import io
                            
                            img = Image.open(flag_file)
                            buffered = io.BytesIO()
                            
                            # Convert to RGB if necessary and save as PNG
                            if img.mode in ('RGBA', 'LA', 'P'):
                                # Keep transparency
                                img.save(buffered, format="PNG")
                                img_type = "png"
                            else:
                                img = img.convert('RGB')
                                img.save(buffered, format="PNG")
                                img_type = "png"
                            
                            img_str = base64.b64encode(buffered.getvalue()).decode()
                            flag_html = f'<img src="data:image/{img_type};base64,{img_str}" style="height:40px; vertical-align:middle; margin-right:10px;">'
                        except Exception as e:
                            flag_html = '🏴 '
                    else:
                        flag_html = '🏴 '
                    
                    # Count players with match reports in this team
                    team_players_with_match_reports = 0
                    if not df_reports.empty:
                        team_players_with_match_reports = country_players[name_col].isin(df_reports['Player Name']).sum()
                    
                    # Count players with individual reports in this team
                    team_players_with_individual_reports = 0
                    if not df_individual_reports.empty:
                        team_players_with_individual_reports = country_players[name_col].isin(df_individual_reports['Player']).sum()
                    
                    # Display country header
                    st.markdown(
                        f'<div style="background-color: #002B5B; color: white; padding: 1rem; border-radius: 8px; margin: 1rem 0; font-weight: bold; font-size: 1.2rem;">{flag_html}{country} U20 ({len(country_players)} players | ⚽ {team_players_with_match_reports} match reports | 📋 {team_players_with_individual_reports} individual reports)</div>',
                        unsafe_allow_html=True
                    )
                    
                    # Display players as expandable cards within this team
                    for idx, player in country_players.iterrows():
                        player_name = player[name_col]
                        player_position = player.get(position_col, 'N/A')
                        player_team = player[country_col]
                        player_age = player.get('Edad', 'N/A')
                        player_club = player.get('Equipo', 'N/A')
                        
                        # Check if player has match reports
                        has_match_reports = False
                        player_reports = pd.DataFrame()
                        if not df_reports.empty:
                            player_reports = df_reports[df_reports['Player Name'] == player_name]
                            has_match_reports = len(player_reports) > 0
                        
                        # Check if player has individual reports
                        has_individual_reports = False
                        player_individual_reports = pd.DataFrame()
                        if not df_individual_reports.empty:
                            # Try exact match first
                            player_individual_reports = df_individual_reports[df_individual_reports['Player'] == player_name]
                            # If no exact match, try case-insensitive and stripped
                            if len(player_individual_reports) == 0:
                                player_individual_reports = df_individual_reports[
                                    df_individual_reports['Player'].str.strip().str.lower() == player_name.strip().lower()
                                ]
                            has_individual_reports = len(player_individual_reports) > 0
                        
                        # Card header with status indicator
                        has_any_report = has_match_reports or has_individual_reports
                        
                        # Determinar emoji según conclusión
                        status_emoji = ""
                        
                        if has_any_report:
                            # Primero verificar si tiene informe individual con conclusión
                            if has_individual_reports:
                                # Get the most recent individual report
                                latest_ind = player_individual_reports.iloc[-1]
                                conclusion = str(latest_ind.get('Conclusion', '')).strip().upper()
                                
                                # ⭐️ si tiene A - Firmar (Sign)
                                if 'A -' in conclusion or 'A-' in conclusion or conclusion.startswith('A '):
                                    status_emoji = "⭐️"
                                # 🟢 si tiene B+ - Seguir para Firmar (Follow to Sign)
                                elif 'B+' in conclusion:
                                    status_emoji = "🟢"
                                # ☑️ si tiene cualquier otro informe
                                else:
                                    status_emoji = "☑️"
                            # Si solo tiene match reports
                            elif has_match_reports:
                                # Verificar conclusiones de match reports
                                if not player_reports.empty:
                                    latest_match = player_reports.iloc[-1]
                                    match_conclusion = str(latest_match.get('Conclusion', '')).strip().upper()
                                    
                                    if 'A -' in match_conclusion or 'A-' in match_conclusion or match_conclusion.startswith('A '):
                                        status_emoji = "⭐️"
                                    elif 'B+' in match_conclusion:
                                        status_emoji = "🟢"
                                    else:
                                        status_emoji = "☑️"
                                else:
                                    status_emoji = "☑️"
                        
                        # Siempre usar fondo blanco (sin color)
                        card_bg_color = "white"
                        
                        # Al Nassr minimalist player profile
                        with st.container():
                            # Get player photo
                            player_photo_base64 = ""
                            try:
                                from PIL import Image
                                import io
                                import os
                                
                                player_photo_file = f"player_photos/{player_name.replace(' ', '_')}.jpg"
                                if not os.path.exists(player_photo_file):
                                    player_photo_file = 'profile1.jpg'
                                
                                if os.path.exists(player_photo_file):
                                    img = Image.open(player_photo_file)
                                    img.thumbnail((120, 120), Image.Resampling.LANCZOS)
                                    buffered = io.BytesIO()
                                    
                                    if img.mode in ('RGBA', 'LA', 'P'):
                                        img.save(buffered, format="PNG")
                                    else:
                                        img = img.convert('RGB')
                                        img.save(buffered, format="PNG")
                                    
                                    player_photo_base64 = base64.b64encode(buffered.getvalue()).decode()
                            except:
                                pass
                            
                            # Birth year
                            birth_year = str(player.get('Año', ''))[:4] if player.get('Año') else 'N/A'
                            contract_date = str(player.get('Fin Contrato', 'N/A'))[:10]
                            
                            # Expander con el nombre del jugador
                            with st.expander(
                                f"{status_emoji} **{player_name}**",
                                expanded=False
                            ):
                                # Al Nassr Header Section
                                photo_html = f'<img src="data:image/png;base64,{player_photo_base64}" style="width:120px; height:120px; border-radius:50%; object-fit:cover; border:3px solid #FFC60A;">' if player_photo_base64 else '<div style="width:120px; height:120px; border-radius:50%; background:#FFC60A; display:flex; align-items:center; justify-content:center; font-size:48px; border:3px solid #FFC60A;">👤</div>'
                                
                                st.markdown(f"""
                                    <div style="background: #1a2332; padding: 30px; border-radius: 8px; margin-bottom: 30px;">
                                        <div style="display: flex; align-items: center; gap: 30px; flex-wrap: wrap;">
                                            <div>{photo_html}</div>
                                            <div style="flex: 1; min-width: 300px;">
                                                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 20px;">
                                                    <div>
                                                        <div style="color: #FFC60A; font-size: 12px; text-transform: uppercase; font-weight: 600; margin-bottom: 5px;">POSITION</div>
                                                        <div style="color: white; font-size: 24px; font-weight: 600;">{player_position}</div>
                                                    </div>
                                                    <div>
                                                        <div style="color: #FFC60A; font-size: 12px; text-transform: uppercase; font-weight: 600; margin-bottom: 5px;">TEAM</div>
                                                        <div style="color: white; font-size: 24px; font-weight: 600;">{player.get('Equipo', 'N/A')}</div>
                                                    </div>
                                                    <div>
                                                        <div style="color: #FFC60A; font-size: 12px; text-transform: uppercase; font-weight: 600; margin-bottom: 5px;">AGE</div>
                                                        <div style="color: white; font-size: 24px; font-weight: 600;">{player_age}</div>
                                                    </div>
                                                    <div>
                                                        <div style="color: #FFC60A; font-size: 12px; text-transform: uppercase; font-weight: 600; margin-bottom: 5px;">YEAR</div>
                                                        <div style="color: white; font-size: 24px; font-weight: 600;">{birth_year}</div>
                                                    </div>
                                                    <div>
                                                        <div style="color: #FFC60A; font-size: 12px; text-transform: uppercase; font-weight: 600; margin-bottom: 5px;">CONTRACT</div>
                                                        <div style="color: white; font-size: 24px; font-weight: 600;">{contract_date}</div>
                                                    </div>
                                                </div>
                                            </div>
                                        </div>
                                    </div>
                                """, unsafe_allow_html=True)
                                
                                # Photo upload section
                                st.markdown("---")
                                col_upload1, col_upload2 = st.columns([2, 4])
                                
                                with col_upload1:
                                    uploaded_player_photo = st.file_uploader(
                                        "📷 Subir Foto del Jugador",
                                        type=['jpg', 'jpeg', 'png'],
                                        key=f"upload_photo_{idx}_{player_name.replace(' ', '_').replace('.', '_')}",
                                        help="Sube una foto del jugador"
                                    )
                                    
                                    if uploaded_player_photo:
                                        if st.button("💾 Guardar Foto", key=f"save_photo_{idx}_{player_name.replace(' ', '_').replace('.', '_')}"):
                                            import os
                                            os.makedirs('player_photos', exist_ok=True)
                                            photo_filename = f"player_photos/{player_name.replace(' ', '_')}.jpg"
                                            with open(photo_filename, 'wb') as f:
                                                f.write(uploaded_player_photo.getbuffer())
                                            st.success("✅ Foto guardada correctamente!")
                                            import time
                                            time.sleep(1)
                                            st.rerun()
                                
                                st.markdown("---")
                                
                                # Match Reports Section with Al Nassr Design
                                if has_match_reports:
                                    st.markdown(f"""
                                        <div style="color: #1a2332; font-size: 20px; font-weight: 700; margin: 20px 0 15px 0;">
                                            {len(player_reports)} Match Report(s)
                                        </div>
                                    """, unsafe_allow_html=True)
                                    
                                    # Display each match report card
                                    for rep_idx, report in player_reports.iterrows():
                                        match_date = report['Date']
                                        match_teams = report['Match']
                                        scout = report['Scout']
                                        phase = report['Phase']
                                        performance = report['Performance']
                                        potential = report['Potential']
                                        conclusion = report.get('Conclusion', 'B - Seguir')
                                        full_report = report.get('Report', '')
                                        
                                        # Conclusion badge color
                                        conclusion_str = str(conclusion).strip().upper()
                                        if 'A -' in conclusion_str or 'A-' in conclusion_str or conclusion_str.startswith('A '):
                                            conclusion_color = '#4CAF50'  # Green for A
                                            conclusion_bg = 'rgba(76, 175, 80, 0.1)'
                                        elif 'B+' in conclusion_str:
                                            conclusion_color = '#4CAF50'  # Green for B+
                                            conclusion_bg = 'rgba(76, 175, 80, 0.1)'
                                        else:
                                            conclusion_color = '#ff8c00'  # Orange for B
                                            conclusion_bg = 'rgba(255, 140, 0, 0.1)'
                                        
                                        # Performance percentage for red gradient bar
                                        perf_percent = (performance / 6) * 100
                                        pot_percent = (potential / 6) * 100
                                        
                                        # Unique key for toggle
                                        toggle_key = f"toggle_report_{idx}_{player_name.replace(' ', '_').replace('.', '_')}_{rep_idx}"
                                        if toggle_key not in st.session_state:
                                            st.session_state[toggle_key] = False
                                        
                                        # Match Report Card with Al Nassr Design
                                        st.markdown(f'''
                                        <div style="border: 2px solid #e0e0e0; border-radius: 8px; margin-bottom: 15px; overflow: hidden;">
                                            <div style="background: #1a2332; padding: 15px 20px; color: white;">
                                                <div style="font-size: 16px; font-weight: 600;">{match_date} | {match_teams}</div>
                                            </div>
                                            <div style="background: #f9f9f9; padding: 12px 20px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 10px;">
                                                <div style="font-size: 13px; color: #666;">
                                                    <strong>Scout:</strong> {scout} | <strong>Phase:</strong> {phase}
                                                </div>
                                                <div style="background: {conclusion_color}; color: white; padding: 4px 12px; border-radius: 12px; font-size: 11px; font-weight: 600; text-transform: uppercase;">
                                                    {conclusion}
                                                </div>
                                            </div>
                                            <div style="padding: 20px; background: white;">
                                                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                                                    <div>
                                                        <div style="font-size: 12px; color: #666; margin-bottom: 8px; font-weight: 600;">PERFORMANCE</div>
                                                        <div style="font-size: 24px; font-weight: 700; color: #1a2332; margin-bottom: 8px;">{performance}/6</div>
                                                        <div style="background: #f0f0f0; border-radius: 10px; height: 8px; overflow: hidden;">
                                                            <div style="background: linear-gradient(90deg, #ff4444, #ff6666); height: 100%; width: {perf_percent}%; border-radius: 10px;"></div>
                                                        </div>
                                                    </div>
                                                    <div>
                                                        <div style="font-size: 12px; color: #666; margin-bottom: 8px; font-weight: 600;">POTENTIAL</div>
                                                        <div style="font-size: 24px; font-weight: 700; color: #1a2332; margin-bottom: 8px;">{potential}/6</div>
                                                        <div style="background: #f0f0f0; border-radius: 10px; height: 8px; overflow: hidden;">
                                                            <div style="background: linear-gradient(90deg, #ff4444, #ff6666); height: 100%; width: {pot_percent}%; border-radius: 10px;"></div>
                                                        </div>
                                                    </div>
                                                </div>
                                            </div>
                                        </div>
                                        ''', unsafe_allow_html=True)
                                        
                                        # Toggle button for full report
                                        if full_report and str(full_report).strip() and str(full_report) != 'nan':
                                            col_btn, col_space = st.columns([1, 5])
                                            with col_btn:
                                                button_text = "▼ Ver Reporte" if not st.session_state[toggle_key] else "▲ Ocultar"
                                                if st.button(button_text, key=f"btn_{toggle_key}"):
                                                    st.session_state[toggle_key] = not st.session_state[toggle_key]
                                                    st.rerun()
                                            
                                            # Show report if toggled
                                            if st.session_state[toggle_key]:
                                                st.markdown(f"""
                                                    <div style="border-top: 3px solid #FFC60A; padding: 20px; background: #f5f5f5; margin-top: -15px; margin-bottom: 15px; border-radius: 0 0 8px 8px;">
                                                        <div style="color: #333; line-height: 1.6;">{full_report}</div>
                                                    </div>
                                                """, unsafe_allow_html=True)
                                
                                # Individual Report Section (if exists)
                                if has_individual_reports:
                                    st.markdown("""
                                        <div style="color: #1a2332; font-size: 20px; font-weight: 700; margin: 30px 0 15px 0;">
                                            Individual Reports
                                        </div>
                                    """, unsafe_allow_html=True)
                                    
                                    # Calcular medias
                                    num_reports = len(player_individual_reports)
                                    avg_rendimiento = player_individual_reports['Rendimiento'].mean()
                                    avg_potencial = player_individual_reports['Potencial'].mean()
                                    
                                    # Obtener lista de scouts
                                    scouts_list = player_individual_reports['Scout'].dropna().unique().tolist()
                                    scouts_text = ", ".join(scouts_list) if scouts_list else "N/A"
                                    
                                    # Mostrar información general
                                    st.markdown(f"📊 **{num_reports} Informe(s) de: {scouts_text}**")
                                    st.markdown("")
                                    
                                    # Cargar logo Al Nassr para las tarjetas
                                    try:
                                        from PIL import Image
                                        import io
                                        logo_img = Image.open('alnassr.png')
                                        logo_img.thumbnail((40, 40))
                                        buffered = io.BytesIO()
                                        if logo_img.mode in ('RGBA', 'LA', 'P'):
                                            logo_img.save(buffered, format="PNG")
                                        else:
                                            logo_img = logo_img.convert('RGB')
                                            logo_img.save(buffered, format="PNG")
                                        logo_str = base64.b64encode(buffered.getvalue()).decode()
                                        logo_html = f'<img src="data:image/png;base64,{logo_str}" style="height:24px; margin-bottom:8px;">'
                                    except:
                                        logo_html = '<div style="font-size:24px; margin-bottom:10px;">🦅</div>'
                                    
                                    # Mostrar valoraciones medias en 2 columnas con estilo minimalista
                                    col_avg1, col_avg2 = st.columns(2)
                                    
                                    with col_avg1:
                                        st.markdown("""
                                            <div style="background: white; padding: 20px 18px; border-radius: 10px; text-align: center; box-shadow: 0 2px 12px rgba(0,0,0,0.06); border: 1.5px solid #FFD700;">
                                                {}
                                                <p style="font-size: 11px; color: #999; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 8px; font-weight: 600;">🎯 RENDIMIENTO</p>
                                                <h1 style="font-size: 48px; font-weight: 700; color: #1B2845; margin: 8px 0; line-height: 1;">{:.1f}</h1>
                                                <p style="font-size: 10px; color: #999; margin-top: 6px;">Sobre 6.0</p>
                                            </div>
                                        """.format(logo_html, avg_rendimiento), unsafe_allow_html=True)
                                    
                                    with col_avg2:
                                        st.markdown("""
                                            <div style="background: white; padding: 20px 18px; border-radius: 10px; text-align: center; box-shadow: 0 2px 12px rgba(0,0,0,0.06); border: 1.5px solid #FFD700;">
                                                {}
                                                <p style="font-size: 11px; color: #999; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 8px; font-weight: 600;">⭐ POTENCIAL</p>
                                                <h1 style="font-size: 48px; font-weight: 700; color: #1B2845; margin: 8px 0; line-height: 1;">{:.1f}</h1>
                                                <p style="font-size: 10px; color: #999; margin-top: 6px;">Sobre 6.0</p>
                                            </div>
                                        """.format(logo_html, avg_potencial), unsafe_allow_html=True)
                                    
                                    # Si solo hay 1 informe, mostrar nota
                                    st.markdown("")
                                    if num_reports == 1:
                                        st.info("ℹ️ Este jugador tiene 1 informe individual.")
                                    else:
                                        st.success(f"✅ Media calculada de {num_reports} informes de diferentes scouts.")
                                    
                                    # Mostrar informes individuales
                                    st.markdown("---")
                                    st.markdown(f"### 📝 Informes Individuales ({num_reports})")
                                    st.markdown("")
                                    
                                    for idx, ind_report in player_individual_reports.iterrows():
                                        scout_name = ind_report.get('Scout', 'Sin nombre')
                                        if pd.isna(scout_name) or str(scout_name).strip() == '' or str(scout_name) == 'nan':
                                            scout_name = "Sin nombre"
                                        
                                        report_date = ind_report.get('Date', 'N/A')
                                        
                                        with st.expander(f"👤 {scout_name} - 📅 {report_date}", expanded=False):
                                            col_r1, col_r2 = st.columns(2)
                                            
                                            with col_r1:
                                                st.metric("🎯 Rendimiento", f"{ind_report['Rendimiento']}/6")
                                            
                                            with col_r2:
                                                st.metric("⭐ Potencial", f"{ind_report['Potencial']}/6")
                                            
                                            # Perfil
                                            perfil_val = ind_report.get('Perfil', 'N/A')
                                            if isinstance(perfil_val, str) and '-' in perfil_val:
                                                st.markdown(f"🏆 **Perfil:** {perfil_val}")
                                            else:
                                                st.markdown(f"🏆 **Perfil:** {perfil_val}/6")
                                            
                                            # Technical comment
                                            tech_comment = ind_report.get('Technical Comment', None)
                                            if tech_comment and str(tech_comment) != 'nan' and str(tech_comment).strip():
                                                st.markdown("💬 **Comentario Técnico:**")
                                                st.info(tech_comment)
                                            
                                            # Conclusion
                                            conclusion_text = ind_report.get('Conclusion', '')
                                            if conclusion_text and str(conclusion_text) != 'nan':
                                                st.markdown(f"✅ **Conclusión:** {conclusion_text}")
                                
                                # Player card display complete
                            
        
        except FileNotFoundError:
            st.error("❌ WorldCupU20playersdata.xlsx file not found. Please add the file to the project directory." if st.session_state.language == 'en' else "❌ ملف WorldCupU20playersdata.xlsx غير موجود. يرجى إضافة الملف إلى المشروع.")
        except Exception as e:
            st.error(f"Error loading player data: {e}" if st.session_state.language == 'en' else f"خطأ في تحميل بيانات اللاعبين: {e}")
    
    # Tab 6: MATCH REPORTS - Al Nassr Professional Dashboard
    with tabs[5]:
        # Al Nassr Match Reports CSS
        st.markdown("""
            <style>
            .stats-card {
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                text-align: center;
                border-left: 4px solid #FFC60A;
            }
            .stats-number {
                font-size: 36px;
                font-weight: 700;
                color: #1a2332;
                margin: 10px 0;
            }
            .stats-label {
                font-size: 12px;
                color: #666;
                text-transform: uppercase;
                letter-spacing: 1px;
                font-weight: 600;
            }
            .match-card {
                background: white;
                border-left: 4px solid #e0e0e0;
                border-radius: 8px;
                padding: 20px;
                margin-bottom: 15px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.08);
                transition: all 0.3s ease;
                cursor: pointer;
            }
            .match-card:hover {
                border-left-color: #FFC60A;
                transform: translateX(5px);
                box-shadow: 0 4px 12px rgba(255,198,10,0.3);
            }
            </style>
        """, unsafe_allow_html=True)
        
        # Title with Al Nassr logo
        try:
            from PIL import Image
            import io
            logo_dashboard = Image.open('alnassr.png')
            logo_dashboard.thumbnail((40, 40))
            buffered_logo = io.BytesIO()
            if logo_dashboard.mode in ('RGBA', 'LA', 'P'):
                logo_dashboard.save(buffered_logo, format="PNG")
            else:
                logo_dashboard = logo_dashboard.convert('RGB')
                logo_dashboard.save(buffered_logo, format="PNG")
            logo_dashboard_str = base64.b64encode(buffered_logo.getvalue()).decode()
            logo_dashboard_html = f'<img src="data:image/png;base64,{logo_dashboard_str}" style="height:32px; vertical-align:middle; margin-right:10px;">'
        except:
            logo_dashboard_html = '⚽'
        
        st.markdown(f"<h2 style='text-align: center; color: #1a2332;'>{logo_dashboard_html} MATCH REPORTS DASHBOARD</h2>", unsafe_allow_html=True)
        st.markdown("<h3 style='text-align: center; color: #666;'>FIFA U20 World Cup</h3>", unsafe_allow_html=True)
        
        # Botón de recarga
        col_reload, col_space = st.columns([1, 5])
        with col_reload:
            if st.button("🔄 Recargar", key="reload_match_reports"):
                st.rerun()
        
        # Scout photo name mapping
        SCOUT_PHOTO_MAP = {
            'Alvaro': 'alvaro',
            'Álvaro': 'alvaro',
            'Juan Gambero': 'juan',
            'Juan': 'juan',
            'Rafa Gil': 'rafa',
            'Rafa': 'rafa',
            'Rafael Gil': 'rafa'
        }
        
        # Country flag mapping - Complete
        COUNTRY_FLAG_MAP = {
            'Arabia Saudita': 'saff.png',
            'Italia': 'italia.png',
            'Estados Unidos': 'usa.png',
            'Francia': 'francia.png',
            'Argentina': 'afa.png',
            'Japón': 'japan.png',
            'Japon': 'japan.png',
            'Noruega': 'noruega.png',
            'Sudáfrica': 'sudafrica.png',
            'Sudafrica': 'sudafrica.png',
            'Nueva Caledonia': 'nuevacaledonia.png',
            'Cuba': 'cuba.png',
            'Panamá': 'panama.png',
            'Panama': 'panama.png',
            'Corea del Sur': 'korea.png',
            'Ucrania': 'ucrania.png',
            'Paraguay': 'paraguay.png',
            'Egipto': 'egipto.png',
            'Chile': 'chile.png',
            'México': 'mexico.png',
            'Mexico': 'mexico.png',
            'Marruecos': 'marruecos.png',
            'Brasil': 'brasil.png',
            'España': 'spain.png',
            'Espana': 'spain.png',
            'Nueva Zelanda': 'nuevazelanda.png',
            'Australia': 'australia.png',
            'Colombia': 'colombiau20.png',
            'Nigeria': 'nigeria.png',
            'Inglaterra': 'inglaterra.png',
            'Inglés': 'inglaterra.png',
            'Ecuador': 'ecuador.png',
            'Venezuela': 'venezuela.png',
            'Uruguay': 'uruguay.png',
            'Perú': 'peru.png',
            'Peru': 'peru.png',
            'Costa Rica': 'costarica.png',
            'Honduras': 'honduras.png',
            'Guatemala': 'guatemala.png',
            'Bolivia': 'bolivia.png'
        }
        
        # Load match reports from local Excel
        try:
            df_reports = pd.read_excel('fifa_u20_player_reports.xlsx')
        except FileNotFoundError:
            df_reports = pd.DataFrame()
        except Exception as e:
            df_reports = pd.DataFrame()
        
        # Load player birth year data from WorldCupU20playersdata.xlsx
        try:
            df_players_data = pd.read_excel('WorldCupU20playersdata.xlsx')
            
            # Detect player column names in both dataframes
            player_col_reports = None
            player_col_data = None
            year_col_data = None
            
            # Check for player column in reports (could be 'Player Name', 'Player', etc.)
            for col in df_reports.columns:
                col_lower = col.lower().replace(' ', '')
                if col_lower in ['playername', 'player', 'jugador', 'nombre', 'name']:
                    player_col_reports = col
                    break
            
            # Check for player column in data (could be 'Nombre', 'Player', etc.)
            for col in df_players_data.columns:
                col_lower = col.lower().replace(' ', '')
                if col_lower in ['nombre', 'player', 'playername', 'jugador', 'name']:
                    player_col_data = col
                    break
            
            # Check for year column (could be 'Año', 'Year', 'Año')
            for col in df_players_data.columns:
                col_lower = col.lower()
                if col_lower in ['año', 'year', 'año', 'ano']:
                    year_col_data = col
                    break
            
            # Try to merge if both have player columns and year exists
            if player_col_reports and player_col_data and year_col_data:
                # Create a temporary dataframe for merge
                df_players_data_temp = df_players_data[[player_col_data, year_col_data]].copy()
                df_players_data_temp.columns = [player_col_reports, 'BirthYear']
                
                # Merge to get birth year
                df_reports = df_reports.merge(
                    df_players_data_temp,
                    on=player_col_reports,
                    how='left'
                )
        except Exception as e:
            pass  # Continue without birth year data
        
        # Detect Player column name
        player_col = None
        for col in df_reports.columns:
            col_lower = col.lower().replace(' ', '')
            if col_lower in ['playername', 'player', 'jugador', 'nombre', 'name']:
                player_col = col
                break
        
        if not player_col and len(df_reports.columns) > 0:
            # If no player column found, try to use first column or a common pattern
            for col in df_reports.columns:
                if 'player' in col.lower() or 'jugador' in col.lower() or 'nombre' in col.lower():
                    player_col = col
                    break
            if not player_col:
                st.error("⚠️ No se encontró la columna de jugadores. Columnas disponibles: " + ", ".join(df_reports.columns.tolist()))
                player_col = df_reports.columns[0]  # Use first column as fallback
        
        if df_reports.empty:
            st.info("📊 No match reports yet. Create match reports to see them here!")
        else:
            # Stats Cards at the top
            st.markdown("---")
            total_matches = df_reports['Match'].nunique()
            total_reports = len(df_reports)
            total_players = df_reports[player_col].nunique()
            total_teams = len(set([team.strip() for match in df_reports['Match'].unique() for team in match.split(' vs ') if ' vs ' in match]))
            
            col_s1, col_s2, col_s3, col_s4 = st.columns(4)
            with col_s1:
                st.markdown(f"""
                    <div class="stats-card">
                        <div class="stats-label">⚽ Partidos</div>
                        <div class="stats-number">{total_matches}</div>
                    </div>
                """, unsafe_allow_html=True)
            
            with col_s2:
                st.markdown(f"""
                    <div class="stats-card">
                        <div class="stats-label">📝 Reportes</div>
                        <div class="stats-number">{total_reports}</div>
                    </div>
                """, unsafe_allow_html=True)
            
            with col_s3:
                st.markdown(f"""
                    <div class="stats-card">
                        <div class="stats-label">👥 Jugadores</div>
                        <div class="stats-number">{total_players}</div>
                    </div>
                """, unsafe_allow_html=True)
            
            with col_s4:
                st.markdown(f"""
                    <div class="stats-card">
                        <div class="stats-label">🏆 Equipos</div>
                        <div class="stats-number">{total_teams}</div>
                    </div>
                """, unsafe_allow_html=True)
            
            # Get scouts list for filtering
            scouts_in_reports = df_reports['Scout'].dropna().unique().tolist()
            
            # Filter options
            st.markdown("---")
            st.markdown("<h3 style='color: #1a2332;'>🔍 FILTROS</h3>", unsafe_allow_html=True)
            col_f1, col_f2, col_f3 = st.columns(3)
            
            with col_f1:
                filter_scout = st.selectbox(
                    "🔍 Filter by Scout" if st.session_state.language == 'en' else "🔍 تصفية حسب الكشاف",
                    [""] + sorted(df_reports['Scout'].dropna().unique().tolist()),
                    key="filter_scout_match_reports"
                )
            
            with col_f2:
                filter_phase = st.selectbox(
                    "🏆 Filter by Phase" if st.session_state.language == 'en' else "🏆 تصفية حسب المرحلة",
                    [""] + sorted(df_reports['Phase'].dropna().unique().tolist()),
                    key="filter_phase_match_reports"
                )
            
            with col_f3:
                filter_conclusion = st.selectbox(
                    "✅ Filter by Conclusion" if st.session_state.language == 'en' else "✅ تصفية حسب الخلاصة",
                    ["", "A - Firmar (Sign)", "B+ - Seguir para Firmar (Follow to Sign)", "B - Seguir (Follow)"],
                    key="filter_conclusion_match_reports"
                )
            
            # Apply filters
            filtered_reports = df_reports.copy()
            if filter_scout:
                filtered_reports = filtered_reports[filtered_reports['Scout'] == filter_scout]
            if filter_phase:
                filtered_reports = filtered_reports[filtered_reports['Phase'] == filter_phase]
            if filter_conclusion:
                # Match exact conclusion from Excel
                filtered_reports = filtered_reports[filtered_reports['Conclusion'] == filter_conclusion]
            
            st.markdown("---")
            st.markdown(f"<p style='color: #666; font-size: 14px;'><strong>📊 Reportes filtrados:</strong> {len(filtered_reports)}</p>", unsafe_allow_html=True)
            st.markdown("")
            
            # Download buttons for Match Reports
            if not filtered_reports.empty:
                st.markdown("### 📥 Descargar Datos / Download Data")
                create_download_buttons(
                    filtered_reports, 
                    filename_base="fifa_u20_match_reports",
                    label_prefix="Descargar / Download"
                )
                st.markdown("---")
            
            # Group by Scout first
            for scout in scouts_in_reports:
                scout_reports = filtered_reports[filtered_reports['Scout'] == scout]
                
                if len(scout_reports) == 0:
                    continue
                
                # Scout Header Section with photo
                scout_filename = SCOUT_PHOTO_MAP.get(scout, None)
                if not scout_filename:
                    # Try to match partial name
                    for key in SCOUT_PHOTO_MAP.keys():
                        if key.lower() in scout.lower() or scout.lower() in key.lower():
                            scout_filename = SCOUT_PHOTO_MAP[key]
                            break
                if not scout_filename:
                    scout_filename = scout.replace(' ', '_').lower()
                
                scout_photo_path = f"{scout_filename}.png"
                scout_photo_html = ''
                try:
                    from PIL import Image
                    import io
                    import os
                    if os.path.exists(scout_photo_path):
                        scout_img = Image.open(scout_photo_path)
                        scout_img.thumbnail((50, 50))
                        buffered = io.BytesIO()
                        if scout_img.mode in ('RGBA', 'LA', 'P'):
                            scout_img.save(buffered, format="PNG")
                        else:
                            scout_img = scout_img.convert('RGB')
                            scout_img.save(buffered, format="PNG")
                        scout_img_str = base64.b64encode(buffered.getvalue()).decode()
                        scout_photo_html = f'<img src="data:image/png;base64,{scout_img_str}" style="width: 50px; height: 50px; border-radius: 50%; object-fit: cover; border: 3px solid #FFC60A;">'
                    else:
                        scout_photo_html = '<div style="width: 50px; height: 50px; border-radius: 50%; background: #1a2332; display: flex; align-items: center; justify-content: center; color: #FFC60A; font-size: 24px; border: 3px solid #FFC60A;">👤</div>'
                except Exception as e:
                    scout_photo_html = '<div style="width: 50px; height: 50px; border-radius: 50%; background: #1a2332; display: flex; align-items: center; justify-content: center; color: #FFC60A; font-size: 24px; border: 3px solid #FFC60A;">👤</div>'
                
                st.markdown(f"""
                    <div style="background: #1a2332; border-left: 6px solid #FFC60A; border-radius: 8px; padding: 15px; margin: 20px 0 15px 0; box-shadow: 0 4px 12px rgba(0,0,0,0.15);">
                        <div style="display: flex; align-items: center; gap: 15px;">
                            {scout_photo_html}
                            <div>
                                <h3 style="color: #FFC60A; margin: 0; font-size: 20px; font-weight: 700;">{scout}</h3>
                                <p style="color: white; margin: 3px 0 0 0; font-size: 11px; opacity: 0.9; letter-spacing: 1px;">SCOUT AL NASSR FC</p>
                            </div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
                
                # Group by match for this scout
                scout_matches = scout_reports['Match'].unique()
                
                for match_name in scout_matches:
                    match_reports = scout_reports[scout_reports['Match'] == match_name]
                    match_date = match_reports.iloc[0]['Date']
                    match_phase = match_reports.iloc[0]['Phase']
                    
                    # Initialize session state for this match card
                    match_key = f"match_card_{scout.replace(' ', '_')}_{match_name.replace(' ', '_')}"
                    if match_key not in st.session_state:
                        st.session_state[match_key] = False
                    
                    # Extract teams from match name
                    if ' vs ' in match_name:
                        team1, team2 = match_name.split(' vs ')
                        team1 = team1.strip()
                        team2 = team2.strip()
                    else:
                        team1 = match_name.strip()
                        team2 = ''
                    
                    # Load team flags with debug
                    team1_flag_html = ''
                    team2_flag_html = ''
                    
                    for team_name, var_name in [(team1, 'team1_flag_html'), (team2, 'team2_flag_html')]:
                        flag_file = COUNTRY_FLAG_MAP.get(team_name, None)
                        if flag_file:
                            try:
                                from PIL import Image
                                import io
                                flag_img = Image.open(flag_file)
                                flag_img.thumbnail((24, 24))
                                buffered = io.BytesIO()
                                if flag_img.mode in ('RGBA', 'LA', 'P'):
                                    flag_img.save(buffered, format="PNG")
                                else:
                                    flag_img = flag_img.convert('RGB')
                                    flag_img.save(buffered, format="PNG")
                                flag_str = base64.b64encode(buffered.getvalue()).decode()
                                flag_html = f'<img src="data:image/png;base64,{flag_str}" style="height: 20px; margin-right: 8px; vertical-align: middle;">'
                                if var_name == 'team1_flag_html':
                                    team1_flag_html = flag_html
                                else:
                                    team2_flag_html = flag_html
                            except Exception as e:
                                # Show emoji flag as fallback
                                if var_name == 'team1_flag_html':
                                    team1_flag_html = '🏁 '
                                else:
                                    team2_flag_html = '🏁 '
                        else:
                            # No flag file found in mapping
                            if var_name == 'team1_flag_html':
                                team1_flag_html = '🏁 '
                            else:
                                team2_flag_html = '🏁 '
                    
                    # Match Header Card
                    match_header_html = f"""
                        <div style="background: white; border-left: 4px solid #FFC60A; border-radius: 8px; padding: 18px; margin-bottom: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                            <h4 style="color: #1a2332; margin: 0 0 8px 0; font-size: 18px; font-weight: 700;">{team1_flag_html}{team1} vs {team2_flag_html}{team2}</h4>
                            <p style="margin: 0; color: #666; font-size: 13px;">{match_date} | {match_phase}</p>
                        </div>
                    """
                    st.markdown(match_header_html, unsafe_allow_html=True)
                    
                    # Toggle button
                    if st.button(f"▼ Ver Jugadores ({len(match_reports)})" if not st.session_state[match_key] else "▲ Ocultar", key=f"btn_{match_key}"):
                        st.session_state[match_key] = not st.session_state[match_key]
                        st.rerun()
                    
                    # Show players if expanded
                    if st.session_state[match_key]:
                        # Group players by their ACTUAL team from Excel
                        teams_in_match = match_reports['Team'].unique() if 'Team' in match_reports.columns else [team1, team2]
                        
                        for team_name in teams_in_match:
                            if pd.isna(team_name) or str(team_name).strip() == '':
                                continue
                            
                            team_name = str(team_name).strip()
                            team_players = match_reports[match_reports['Team'] == team_name] if 'Team' in match_reports.columns else match_reports
                            
                            # Load team flag for section header
                            team_flag_html = ''
                            flag_file = COUNTRY_FLAG_MAP.get(team_name, None)
                            if flag_file:
                                try:
                                    from PIL import Image
                                    import io
                                    team_flag_img = Image.open(flag_file)
                                    team_flag_img.thumbnail((24, 24))
                                    buffered_team = io.BytesIO()
                                    if team_flag_img.mode in ('RGBA', 'LA', 'P'):
                                        team_flag_img.save(buffered_team, format="PNG")
                                    else:
                                        team_flag_img = team_flag_img.convert('RGB')
                                        team_flag_img.save(buffered_team, format="PNG")
                                    team_flag_str = base64.b64encode(buffered_team.getvalue()).decode()
                                    team_flag_html = f'<img src="data:image/png;base64,{team_flag_str}" style="height: 20px; margin-right: 8px; vertical-align: middle;">'
                                except:
                                    team_flag_html = '🏁 '
                            else:
                                team_flag_html = '🏁 '
                            
                            # Team section header
                            st.markdown(f"""
                                <div style="background: #f8f9fa; border-left: 4px solid #1a2332; padding: 12px 15px; margin: 15px 0 10px 0; border-radius: 6px;">
                                    <h5 style="margin: 0; color: #1a2332; font-size: 16px; font-weight: 700;">{team_flag_html}{team_name}</h5>
                                </div>
                            """, unsafe_allow_html=True)
                            
                            # Player rows for this team
                            for p_idx, report in team_players.iterrows():
                                player_name = report[player_col]
                                performance = report['Performance']
                                potential = report['Potential']
                                full_report_text = report.get('Report', '')
                                player_position = report.get('Position', 'N/A')
                                player_number = report.get('Number', '-')
                                player_team = report.get('Team', team_name)
                                player_age = report.get('Age', 'N/A')
                                player_birth_year = report.get('BirthYear', report.get('Year', ''))
                                conclusion = report.get('Conclusion', '')
                                
                                # Determine conclusion badge color and text (order matters: B+ before B!)
                                conclusion_str = str(conclusion).strip()
                                conclusion_upper = conclusion_str.upper()
                                
                                # Check for A - FIRMAR first (exact match format from Excel)
                                if 'A - FIRMAR' in conclusion_upper or 'A -' in conclusion_str:
                                    conclusion_color = '#4CAF50'  # Green
                                    conclusion_text = 'A - FIRMAR'
                                # Check for B+ before B (MUST check B+ first! More specific)
                                elif 'B+' in conclusion_str or 'SEGUIR PARA FIRMAR' in conclusion_upper:
                                    conclusion_color = '#2196F3'  # Blue
                                    conclusion_text = 'B+ - SEGUIR PARA FIRMAR'
                                # Then check for B - SEGUIR (but not B+)
                                elif 'B - SEGUIR' in conclusion_upper and 'B+' not in conclusion_str:
                                    conclusion_color = '#FF9800'  # Orange
                                    conclusion_text = 'B - SEGUIR'
                                elif 'B -' in conclusion_str and '+' not in conclusion_str:
                                    conclusion_color = '#FF9800'  # Orange
                                    conclusion_text = 'B - SEGUIR'
                                else:
                                    conclusion_color = '#9E9E9E'  # Gray for unknown
                                    conclusion_text = conclusion_str if conclusion_str and conclusion_str != 'nan' else 'N/A'
                                
                                # Load player's team flag
                                flag_file = COUNTRY_FLAG_MAP.get(str(player_team).strip(), None)
                                if flag_file:
                                    try:
                                        from PIL import Image
                                        import io
                                        flag_img = Image.open(flag_file)
                                        flag_img.thumbnail((24, 24))
                                        buffered = io.BytesIO()
                                        if flag_img.mode in ('RGBA', 'LA', 'P'):
                                            flag_img.save(buffered, format="PNG")
                                        else:
                                            flag_img = flag_img.convert('RGB')
                                            flag_img.save(buffered, format="PNG")
                                        flag_str = base64.b64encode(buffered.getvalue()).decode()
                                        flag_html = f'<img src="data:image/png;base64,{flag_str}" style="height: 18px; margin-right: 6px; vertical-align: middle;">'
                                    except:
                                        flag_html = ''
                                else:
                                    flag_html = ''
                                
                                # Progress bars (RED)
                                perf_percent = (performance / 6) * 100
                                pot_percent = (potential / 6) * 100
                                
                                # Unique key for expand/collapse
                                player_key = f"player_report_{scout.replace(' ', '_')}_{match_name.replace(' ', '_')}_{p_idx}"
                                if player_key not in st.session_state:
                                    st.session_state[player_key] = False
                                
                                # Build birth year display
                                if player_birth_year and str(player_birth_year) not in ['nan', '', 'None']:
                                    try:
                                        birth_year_int = int(float(player_birth_year))
                                        birth_year_display = f"{birth_year_int} • "
                                    except:
                                        birth_year_display = f"{player_birth_year} • "
                                else:
                                    birth_year_display = ""
                                
                                # Horizontal player card: Nombre • Year • Posición | Rendimiento | Potencial | Conclusion Badge
                                player_card_html = f"""
                                    <div style="background: white; border: 1px solid #e0e0e0; border-radius: 8px; padding: 15px; margin-bottom: 10px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 15px;">
                                        <div style="flex: 1; min-width: 200px;">
                                            <div style="font-size: 15px; font-weight: 700; color: #1a2332; margin-bottom: 2px;">{player_name}</div>
                                            <div style="font-size: 11px; color: #999;">{birth_year_display}{player_position} • {player_number}</div>
                                        </div>
                                        <div style="display: flex; gap: 20px; align-items: center; flex-wrap: wrap;">
                                            <div style="text-align: center; min-width: 85px;">
                                                <div style="font-size: 9px; color: #666; margin-bottom: 2px; text-transform: uppercase; font-weight: 600;">Rendimiento</div>
                                                <div style="font-size: 16px; font-weight: 700; color: #1a2332;">{performance}/6</div>
                                                <div style="background: #f0f0f0; border-radius: 3px; height: 5px; overflow: hidden; width: 65px; margin: 4px auto 0;">
                                                    <div style="background: #ff4444; height: 100%; width: {perf_percent}%;"></div>
                                                </div>
                                            </div>
                                            <div style="text-align: center; min-width: 85px;">
                                                <div style="font-size: 9px; color: #666; margin-bottom: 2px; text-transform: uppercase; font-weight: 600;">Potencial</div>
                                                <div style="font-size: 16px; font-weight: 700; color: #1a2332;">{potential}/6</div>
                                                <div style="background: #f0f0f0; border-radius: 3px; height: 5px; overflow: hidden; width: 65px; margin: 4px auto 0;">
                                                    <div style="background: #ff4444; height: 100%; width: {pot_percent}%;"></div>
                                                </div>
                                            </div>
                                            <div style="background: {conclusion_color}; color: white; padding: 8px 12px; border-radius: 6px; font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; box-shadow: 0 2px 4px rgba(0,0,0,0.2);">
                                                {conclusion_text}
                                            </div>
                                        </div>
                                    </div>
                                """
                                st.markdown(player_card_html, unsafe_allow_html=True)
                                
                                # VER and EDITAR buttons
                                edit_key = f"{player_key}_edit"
                                if edit_key not in st.session_state:
                                    st.session_state[edit_key] = False
                                
                                col_btn1, col_btn2, col_space = st.columns([1, 1, 7])
                                with col_btn1:
                                    if st.button("📝 VER" if not st.session_state[player_key] else "❌ CERRAR", key=f"btn_{player_key}", use_container_width=True):
                                        st.session_state[player_key] = not st.session_state[player_key]
                                        st.rerun()
                                with col_btn2:
                                    if st.button("✏️ EDITAR" if not st.session_state[edit_key] else "❌ CANCELAR", key=f"btn_edit_{player_key}", use_container_width=True):
                                        st.session_state[edit_key] = not st.session_state[edit_key]
                                        st.rerun()
                                
                                # EDIT MODE
                                if st.session_state[edit_key]:
                                    st.markdown("---")
                                    st.markdown("### ✏️ Editar Informe")
                                    
                                    # Create editable fields
                                    col_e1, col_e2, col_e3 = st.columns(3)
                                    
                                    with col_e1:
                                        new_performance = st.number_input(
                                            "Rendimiento (1-6)",
                                            min_value=1.0,
                                            max_value=6.0,
                                            value=float(performance),
                                            step=0.5,
                                            key=f"perf_{player_key}"
                                        )
                                    
                                    with col_e2:
                                        new_potential = st.number_input(
                                            "Potencial (1-6)",
                                            min_value=1.0,
                                            max_value=6.0,
                                            value=float(potential),
                                            step=0.5,
                                            key=f"pot_{player_key}"
                                        )
                                    
                                    with col_e3:
                                        conclusion_options = [
                                            "A - Firmar (Sign)",
                                            "B+ - Seguir para Firmar (Follow to Sign)",
                                            "B - Seguir (Follow)"
                                        ]
                                        current_conclusion_idx = 0
                                        for idx, opt in enumerate(conclusion_options):
                                            if opt in str(conclusion):
                                                current_conclusion_idx = idx
                                                break
                                        
                                        new_conclusion = st.selectbox(
                                            "Conclusión",
                                            conclusion_options,
                                            index=current_conclusion_idx,
                                            key=f"concl_{player_key}"
                                        )
                                    
                                    new_report = st.text_area(
                                        "Report",
                                        value=str(full_report_text) if str(full_report_text) != 'nan' else '',
                                        height=150,
                                        key=f"report_{player_key}"
                                    )
                                    
                                    # Save button
                                    if st.button("💾 GUARDAR CAMBIOS", key=f"save_{player_key}", type="primary"):
                                        try:
                                            # Load current Excel data
                                            df_all_reports = pd.read_excel('fifa_u20_player_reports.xlsx')
                                            
                                            # Find the exact row to update using multiple criteria
                                            mask = (
                                                (df_all_reports['Scout'] == scout) &
                                                (df_all_reports['Match'] == match_name) &
                                                (df_all_reports[player_col] == player_name)
                                            )
                                            
                                            # Update the values
                                            df_all_reports.loc[mask, 'Performance'] = new_performance
                                            df_all_reports.loc[mask, 'Potential'] = new_potential
                                            df_all_reports.loc[mask, 'Conclusion'] = new_conclusion
                                            df_all_reports.loc[mask, 'Report'] = new_report
                                            
                                            # Save back to Excel
                                            df_all_reports.to_excel('fifa_u20_player_reports.xlsx', index=False)
                                            
                                            st.success("✅ Informe actualizado exitosamente!")
                                            st.session_state[edit_key] = False
                                            
                                            # Wait a moment and reload
                                            import time
                                            time.sleep(1)
                                            st.rerun()
                                            
                                        except Exception as e:
                                            st.error(f"❌ Error al guardar: {e}")
                                    
                                    st.markdown("---")
                                
                                # Expandable REPORT section (VIEW MODE)
                                elif st.session_state[player_key] and full_report_text and str(full_report_text) != 'nan':
                                    report_html = f"""
                                        <div style="background: #fffef0; border: 2px solid #FFC60A; border-radius: 8px; padding: 25px; margin-bottom: 15px;">
                                            <h4 style="color: #1a2332; margin: 0 0 15px 0; font-size: 16px; border-bottom: 2px solid #FFC60A; padding-bottom: 10px;">📝 REPORT</h4>
                                            <div style="font-size: 14px; color: #333; line-height: 1.8;">
                                                {full_report_text}
                                            </div>
                                        </div>
                                    """
                                    st.markdown(report_html, unsafe_allow_html=True)
    
    # Tab 7: CAMPOGRAMAS
    with tabs[6]:
        if st.session_state.language == 'en':
            st.markdown("### ⚽ Campogramas (Field Maps)")
            st.info("🚧 This section will provide tactical field analysis and heat maps.")
        else:
            st.markdown("### ⚽ رسوم الملعب")
            st.info("🚧 سيوفر هذا القسم تحليلات تكتيكية للملعب وخرائط حرارية.")
        
        # Placeholder for campogramas
        st.markdown("**Coming soon:** Player positioning, movement patterns, heat maps, tactical formations, etc.")
        st.info("🚧 Campogramas will be implemented in future updates" if st.session_state.language == 'en' else "🚧 سيتم تنفيذ الرسوم التكتيكية في التحديثات القادمة")

# -----------------------------
# LOGIN SYSTEM
# -----------------------------
# User credentials database
USERS = {
    'juangambero': {
        'password': 'juangambero',
        'photo': 'juan.png',
        'name': 'Juan Gambero'
    },
    'alvarolopez': {
        'password': 'alvarolopez',
        'photo': 'alvaro.png',
        'name': 'Alvaro Lopez'
    },
    'rafagil': {
        'password': 'rafagil',
        'photo': 'rafa.png',
        'name': 'Rafa Gil'
    }
}

def show_login_page():
    """Display login page with Al Nassr branding"""
    
    # Center the login form
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # Load and display Al Nassr logo
        try:
            from PIL import Image
            import io
            import base64
            
            logo_img = Image.open('alnassr.png')
            logo_img.thumbnail((200, 200))
            buffered = io.BytesIO()
            if logo_img.mode in ('RGBA', 'LA', 'P'):
                logo_img.save(buffered, format="PNG")
            else:
                logo_img = logo_img.convert('RGB')
                logo_img.save(buffered, format="PNG")
            logo_str = base64.b64encode(buffered.getvalue()).decode()
            logo_html = f'<img src="data:image/png;base64,{logo_str}" style="width: 150px; display: block; margin: 0 auto;">'
        except:
            logo_html = '<div style="text-align: center; font-size: 80px;">⚽</div>'
        
        # Login card
        st.markdown(f"""
            <div style="
                background: white;
                border-radius: 16px;
                padding: 40px;
                box-shadow: 0 8px 24px rgba(0,0,0,0.15);
                margin-top: 80px;
                text-align: center;
                border-top: 6px solid #FFC60A;
            ">
                {logo_html}
                <h1 style="color: #1a2332; margin: 25px 0 10px 0; font-size: 28px; font-weight: 700;">INICIO DE SESIÓN</h1>
                <p style="color: #666; font-size: 14px; margin-bottom: 30px; letter-spacing: 1.5px; text-transform: uppercase;">AL NASSR SCOUTING DEPARTMENT</p>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        
        # Login form
        with st.form("login_form", clear_on_submit=False):
            username = st.text_input(
                "👤 Usuario",
                placeholder="Ingresa tu usuario",
                key="login_username"
            )
            password = st.text_input(
                "🔒 Contraseña",
                type="password",
                placeholder="Ingresa tu contraseña",
                key="login_password"
            )
            
            col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
            with col_btn2:
                submit = st.form_submit_button(
                    "🚀 INICIAR SESIÓN",
                    use_container_width=True
                )
            
            if submit:
                if username in USERS and USERS[username]['password'] == password:
                    # Successful login
                    st.session_state.authenticated = True
                    st.session_state.current_user = username
                    st.session_state.user_name = USERS[username]['name']
                    st.session_state.user_photo = USERS[username]['photo']
                    st.success(f"✅ Bienvenido, {USERS[username]['name']}!")
                    st.balloons()
                    st.rerun()
                else:
                    st.error("❌ Usuario o contraseña incorrectos")
        
        # Footer
        st.markdown("""
            <div style="text-align: center; margin-top: 40px; color: #999; font-size: 12px;">
                <p>© 2025 Al Nassr FC - Scouting Department</p>
                <p style="margin-top: 5px;">🔒 Sistema Seguro</p>
            </div>
        """, unsafe_allow_html=True)

def logout():
    """Logout user and clear session"""
    st.session_state.authenticated = False
    st.session_state.current_user = None
    st.session_state.user_name = None
    st.session_state.user_photo = None
    st.session_state.page = 'home'
    st.rerun()

def show_calendar_schedule(category):
    """Show calendar and match schedule for U21 category"""
    
    # Custom CSS for calendar
    st.markdown("""
    <style>
        .calendar-header {
            background: linear-gradient(135deg, #002B5B 0%, #004080 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            margin-bottom: 25px;
            box-shadow: 0 4px 15px rgba(0,43,91,0.3);
        }
        .match-date-header {
            background: #F9D342;
            color: #002B5B;
            padding: 12px 20px;
            border-radius: 8px;
            font-weight: bold;
            font-size: 16px;
            margin: 25px 0 15px 0;
            display: inline-block;
            box-shadow: 0 2px 8px rgba(249,211,66,0.4);
        }
        .match-card {
            background: white;
            border-left: 4px solid #F9D342;
            border-radius: 8px;
            padding: 18px;
            margin-bottom: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            transition: all 0.3s ease;
        }
        .match-card:hover {
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(249,211,66,0.3);
            border-left-color: #002B5B;
        }
        .team-name {
            font-weight: 600;
            color: #002B5B;
            font-size: 15px;
        }
        .match-score {
            font-size: 20px;
            font-weight: 700;
            color: #F9D342;
            padding: 0 15px;
        }
        .match-info {
            color: #666;
            font-size: 13px;
            margin-top: 8px;
        }
        .week-badge {
            background: #002B5B;
            color: white;
            padding: 4px 12px;
            border-radius: 15px;
            font-size: 11px;
            font-weight: 600;
            display: inline-block;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Header with Jawwy logo
    try:
        from PIL import Image
        import io
        jawwy_logo = Image.open('jawwy.png')
        jawwy_logo.thumbnail((60, 60))
        buffered = io.BytesIO()
        if jawwy_logo.mode in ('RGBA', 'LA', 'P'):
            jawwy_logo.save(buffered, format="PNG")
        else:
            jawwy_logo = jawwy_logo.convert('RGB')
            jawwy_logo.save(buffered, format="PNG")
        jawwy_logo_str = base64.b64encode(buffered.getvalue()).decode()
        jawwy_html = f'<img src="data:image/png;base64,{jawwy_logo_str}" style="height:50px; vertical-align:middle; margin-left:15px;">'
    except:
        jawwy_html = ''
    
    if st.session_state.language == 'en':
        st.markdown(f"""
            <div class="calendar-header">
                <h2 style="margin: 0; font-size: 28px;">📅 Match Calendar & Schedule {jawwy_html}</h2>
                <p style="margin: 10px 0 0 0; font-size: 14px; opacity: 0.9;">Saudi Elite League U-21</p>
            </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
            <div class="calendar-header">
                <h2 style="margin: 0; font-size: 28px;">{jawwy_html} 📅 تقويم ومواعيد المباريات</h2>
                <p style="margin: 10px 0 0 0; font-size: 14px; opacity: 0.9;">الدوري السعودي النخبة تحت 21</p>
            </div>
        """, unsafe_allow_html=True)
    
    # U21 League ID
    league_id = 350
    
    # Normalización: Nombres del scraper → Nombres completos
    TEAM_NAMES_NORMALIZED = {
        'Al Adalah': 'Al-Adalah FC',
        'Al Ahli': 'Al-Ahli Saudi FC',
        'Al Bukiryah': 'Al-Bukayriyah FC',
        'Al Ettifaq': 'Ettifaq FC',
        'Al Fateh': 'Al-Fateh SC',
        'Al Fayha': 'Al-Fayha FC',
        'Al Hazem': 'Al-Hazem FC',
        'Al Hilal': 'Al-Hilal SFC',
        'Al Ittihad': 'Al-Ittihad Club',
        'Al Jabalin': 'Al-Jabalain',
        'Al Khaleej': 'Khaleej FC',
        'Al Kholood': 'Al-Kholood Club',
        'Al Najmah': 'Al-Najmah SC',
        'Al Nassr': 'Al-Nassr FC',
        'Al Okhdood': 'Al-Okhdood FC',
        'Al Orobah': 'Al-Orobah FC',
        'Al Qadisiyah': 'Al-Qadsiah FC',
        'Al Raed': 'Al-Raed FC',
        'Al Riyadh': 'Al-Riyadh SC',
        'Al Shabab': 'Al-Shabab FC',
        'Al Taawoun': 'Al-Taawoun FC',
        'Al Wehda': 'Al-Wehda FC',
        'Damac': 'Damac FC',
        'NEOM': 'NEOM Club'
    }

    # Nombres completos → Archivos de logos
    TEAM_LOGOS = {
        'Al-Wehda FC': 'alwehda.png',
        'Al-Jabalain': 'aljabalain.png',
        'Al-Raed FC': 'alraed.png',
        'Al-Orobah FC': 'AlOrobah.png',
        'Al-Adalah FC': 'aladalahclub.png',
        'Al-Bukayriyah FC': 'albukiryahfc.png',
        'Al-Ahli Saudi FC': 'alahli.png',
        'Al-Nassr FC': 'alnassr.png',
        'Al-Taawoun FC': 'altaawoun.png',
        'Al-Shabab FC': 'alshabab.png',
        'Al-Okhdood FC': 'alokhdood.png',
        'Al-Riyadh SC': 'alriyadh.png',
        'Al-Najmah SC': 'alnajma.png',
        'NEOM Club': 'neom.png',
        'Al-Kholood Club': 'alkholood.png',
        'Al-Fateh SC': 'alfateh.png',
        'Damac FC': 'damac.png',
        'Al-Hazem FC': 'alhazem.png',
        'Al-Fayha FC': 'alfayha.png',
        'Ettifaq FC': 'alettifaq.png',
        'Al-Hilal SFC': 'alhilal.png',
        'Al-Qadsiah FC': 'alqadsiah.png',
        'Khaleej FC': 'alkhaleej.png',
        'Al-Ittihad Club': 'alittihad.png'
    }
    
    def get_team_logo_html(team_name_from_scraper, size=24):
        """
        Obtiene el HTML de un logo de equipo en formato base64
        Normaliza el nombre del equipo antes de buscar el logo
        
        Args:
            team_name_from_scraper: Nombre del equipo tal como viene del scraper
            size: Tamaño del logo en pixels
        
        Returns:
            str: HTML del logo o emoji de fallback
        """
        # Paso 1: Normalizar el nombre (scraper → nombre completo)
        normalized_name = TEAM_NAMES_NORMALIZED.get(team_name_from_scraper, team_name_from_scraper)
        
        # Paso 2: Obtener el archivo del logo
        logo_file = TEAM_LOGOS.get(normalized_name, None)
        
        if logo_file:
            try:
                from PIL import Image
                import io
                import base64
                
                img = Image.open(logo_file)
                img.thumbnail((size, size))
                buffered = io.BytesIO()
                
                if img.mode in ('RGBA', 'LA', 'P'):
                    img.save(buffered, format="PNG")
                else:
                    img = img.convert('RGB')
                    img.save(buffered, format="PNG")
                
                img_str = base64.b64encode(buffered.getvalue()).decode()
                return f'<img src="data:image/png;base64,{img_str}" style="height:{size}px; width:{size}px; object-fit:contain; margin-right:8px; vertical-align:middle;">'
            except Exception as e:
                # Si hay error al cargar, devolver emoji
                return '⚽ '
        else:
            # Si no se encuentra el logo, devolver emoji
            return '⚽ '
    
    # Load schedule with cache
    @st.cache_data(ttl=3600)
    def load_schedule_cached(lid):
        return scrape_schedule(lid)
    
    # Loading state
    loading_text = '🔄 Loading schedule...' if st.session_state.language == 'en' else '🔄 جاري تحميل الجدول...'
    with st.spinner(loading_text):
        matches = load_schedule_cached(league_id)
    
    if not matches:
        error_text = "❌ Could not load schedule. Please try again later." if st.session_state.language == 'en' else "❌ تعذر تحميل الجدول. يرجى المحاولة لاحقاً."
        st.error(error_text)
        return
    
    # Stats summary
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        label1 = "🏆 Total Matches" if st.session_state.language == 'en' else "🏆 إجمالي المباريات"
        st.metric(label1, len(matches))
    
    with col2:
        unique_teams = set()
        for m in matches:
            unique_teams.add(m['home_team'])
            unique_teams.add(m['away_team'])
        label2 = "⚽ Teams" if st.session_state.language == 'en' else "⚽ الفرق"
        st.metric(label2, len(unique_teams))
    
    with col3:
        unique_dates = len(set([m['date'] for m in matches if m['date']]))
        label3 = "📅 Match Days" if st.session_state.language == 'en' else "📅 أيام المباريات"
        st.metric(label3, unique_dates)
    
    with col4:
        played = sum(1 for m in matches if m['score'] and m['score'] != '-')
        label4 = "✅ Played" if st.session_state.language == 'en' else "✅ تم لعبها"
        st.metric(label4, played)
    
    st.markdown("---")
    
    # FILTERS SECTION
    st.markdown("### 🔍 " + ("Filters" if st.session_state.language == 'en' else "الفلاتر"))
    
    # Extract unique values for filters
    all_weeks = sorted(list(set([m['week'] for m in matches if m['week']])))
    all_dates = sorted(list(set([m['date'] for m in matches if m['date']])))
    all_teams = sorted(list(set([m['home_team'] for m in matches] + [m['away_team'] for m in matches])))
    
    # Create filter columns
    col_f1, col_f2, col_f3 = st.columns(3)
    
    with col_f1:
        week_label = "Week" if st.session_state.language == 'en' else "الجولة"
        all_weeks_label = "All Weeks" if st.session_state.language == 'en' else "كل الجولات"
        selected_week = st.selectbox(
            f"📅 {week_label}",
            [all_weeks_label] + all_weeks,
            key="filter_week"
        )
    
    with col_f2:
        date_label = "Date" if st.session_state.language == 'en' else "التاريخ"
        all_dates_label = "All Dates" if st.session_state.language == 'en' else "كل التواريخ"
        selected_date = st.selectbox(
            f"📆 {date_label}",
            [all_dates_label] + all_dates,
            key="filter_date"
        )
    
    with col_f3:
        team_label = "Team" if st.session_state.language == 'en' else "الفريق"
        all_teams_label = "All Teams" if st.session_state.language == 'en' else "كل الفرق"
        selected_team = st.selectbox(
            f"⚽ {team_label}",
            [all_teams_label] + all_teams,
            key="filter_team"
        )
    
    # Apply filters (concatenating logic - AND)
    filtered_matches = matches.copy()
    
    # Filter by week
    if selected_week != (all_weeks_label if st.session_state.language == 'en' else "كل الجولات"):
        filtered_matches = [m for m in filtered_matches if m['week'] == selected_week]
    
    # Filter by date
    if selected_date != (all_dates_label if st.session_state.language == 'en' else "كل التواريخ"):
        filtered_matches = [m for m in filtered_matches if m['date'] == selected_date]
    
    # Filter by team (home or away)
    if selected_team != (all_teams_label if st.session_state.language == 'en' else "كل الفرق"):
        filtered_matches = [m for m in filtered_matches if m['home_team'] == selected_team or m['away_team'] == selected_team]
    
    # Show filtered results count
    if st.session_state.language == 'en':
        st.info(f"📊 Showing **{len(filtered_matches)}** of **{len(matches)}** matches")
    else:
        st.info(f"📊 عرض **{len(filtered_matches)}** من **{len(matches)}** مباراة")
    
    st.markdown("---")
    
    # Download buttons (for filtered data)
    df_matches = pd.DataFrame(filtered_matches)
    download_label = "Download Schedule" if st.session_state.language == 'en' else "تحميل الجدول"
    create_download_buttons(
        df_matches,
        filename_base="U21_calendar_filtered",
        label_prefix=download_label
    )
    
    st.markdown("---")
    
    # Display filtered matches grouped by date
    current_date = ""
    
    for match in filtered_matches:
        # Date header when date changes
        if match['date'] != current_date:
            current_date = match['date']
            st.markdown(f"""
                <div class="match-date-header">
                    📅 {current_date}
                </div>
            """, unsafe_allow_html=True)
        
        # Get team logos (pasando los nombres tal como vienen del scraper)
        home_logo = get_team_logo_html(match['home_team'], size=28)
        away_logo = get_team_logo_html(match['away_team'], size=28)
        
        # Match card
        st.markdown(f"""
            <div class="match-card">
                <div style="display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px;">
                    <div style="flex: 1; min-width: 350px;">
                        <div style="display: flex; align-items: center; gap: 5px;">
                            {home_logo}
                            <span class="team-name">{match['home_team']}</span>
                            <span class="match-score">{match['score']}</span>
                            {away_logo}
                            <span class="team-name">{match['away_team']}</span>
                        </div>
                    </div>
                    <div style="text-align: right;">
                        <span class="week-badge">{match['week']}</span>
                    </div>
                </div>
                <div class="match-info">
                    ⏰ {match['time']} | 🏟️ {match['stadium']}
                </div>
            </div>
        """, unsafe_allow_html=True)

# -----------------------------
# MAIN APPLICATION
# -----------------------------
def main():
    # Initialize authentication state
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    if 'current_user' not in st.session_state:
        st.session_state.current_user = None
    if 'user_name' not in st.session_state:
        st.session_state.user_name = None
    if 'user_photo' not in st.session_state:
        st.session_state.user_photo = None
    
    # Check authentication
    if not st.session_state.authenticated:
        show_login_page()
        return
    
    # Apply styling
    apply_custom_css()
    
    # Language toggle button in sidebar
    with st.sidebar:
        # User info at top
        if st.session_state.user_name:
            try:
                from PIL import Image
                import io
                import base64
                
                user_img = Image.open(st.session_state.user_photo)
                user_img.thumbnail((60, 60))
                buffered = io.BytesIO()
                if user_img.mode in ('RGBA', 'LA', 'P'):
                    user_img.save(buffered, format="PNG")
                else:
                    user_img = user_img.convert('RGB')
                    user_img.save(buffered, format="PNG")
                user_img_str = base64.b64encode(buffered.getvalue()).decode()
                user_photo_html = f'<img src="data:image/png;base64,{user_img_str}" style="width: 50px; height: 50px; border-radius: 50%; border: 3px solid #FFC60A; object-fit: cover;">'
            except:
                user_photo_html = '👤'
            
            st.markdown(f"""
                <div style="
                    background: linear-gradient(135deg, #1a2332 0%, #2d3e50 100%);
                    padding: 15px;
                    border-radius: 10px;
                    text-align: center;
                    margin-bottom: 20px;
                    border: 2px solid #FFC60A;
                ">
                    {user_photo_html}
                    <p style="color: white; margin: 10px 0 0 0; font-weight: 600; font-size: 14px;">{st.session_state.user_name}</p>
                    <p style="color: #FFC60A; margin: 3px 0 0 0; font-size: 10px; text-transform: uppercase; letter-spacing: 1px;">SCOUT</p>
                </div>
            """, unsafe_allow_html=True)
        
        st.markdown("### Settings / الإعدادات")
        if st.button("🇸🇦 عربي" if st.session_state.language == 'en' else "🇬🇧 English",
                    key="lang_toggle",
                    help="Toggle Language",
                    use_container_width=True):
            toggle_language()
        
        st.markdown("---")
        
        # Navigation
        if st.session_state.page != 'home':
            if st.button("🏠 " + ("Home" if st.session_state.language == 'en' else "الرئيسية"), 
                        key="btn_home",
                        use_container_width=True):
                st.session_state.page = 'home'
                st.rerun()
        
        st.markdown("---")
        
        # Logout button
        if st.button("🚪 Cerrar Sesión", key="btn_logout", use_container_width=True, type="secondary"):
            logout()
    
    # Show appropriate page based on session state
    if st.session_state.page == 'home':
        show_home_page()
    elif st.session_state.page == 'categories':
        show_categories_page()
    elif st.session_state.page == 'category_view':
        show_category_view()
    elif st.session_state.page == 'fifa_u20':
        show_fifa_u20_view()
    elif st.session_state.page == 'advanced_data':
        show_advanced_data_analysis(None)

# -----------------------------
# RUN APPLICATION
# -----------------------------
if __name__ == "__main__":
    main()
